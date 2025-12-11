import sqlite3
import datetime
import re
import traceback
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget,
                            QTableWidgetItem, QPushButton, QMessageBox,
                            QFileDialog, QHeaderView, QGroupBox, QGridLayout,
                            QColorDialog, QLineEdit, QComboBox)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog

from database import get_connection
from category_utils import resolve_category_code

class CompteResultatDisplay(QDialog):
    def __init__(self, parent, config_data):
        super().__init__(parent)
        self.parent = parent
        self.config_data = config_data
        
        # Extraction des paramètres de configuration
        self.project_ids = config_data.get('project_ids', [])
        self.years = config_data.get('years', [])
        self.period_type = config_data.get('period_type', 'yearly')  # 'yearly' ou 'monthly'
        self.granularity = config_data.get('granularity', 'yearly')  # 'yearly' ou 'monthly'
        self.cost_type = config_data.get('cost_type', 'cout_production')  # Type de coût sélectionné
        
        # Vérifier si au moins un projet a le CIR activé
        self.has_cir_projects = self.check_cir_projects()
        
        self.setWindowTitle("Compte de Résultat")
        self.setMinimumSize(1000, 700)
        
        self.init_ui()
        self.load_data()
    
    def load_export_settings(self):
        """Charge les paramètres d'export depuis la base de données"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM export_settings WHERE id = 1')
            result = cursor.fetchone()
            
            if result:
                settings = type('Settings', (), {})()
                settings.title_color = result[1]
                settings.header_color = result[2]
                settings.total_color = result[3]
                settings.result_color = result[4]
                settings.logo_path = result[5]
                settings.logo_position = result[6]
            else:
                # Paramètres par défaut
                settings = type('Settings', (), {})()
                settings.title_color = '#2c3e50'
                settings.header_color = '#34495e'
                settings.total_color = '#3498db'
                settings.result_color = '#2ecc71'
                settings.logo_path = ''
                settings.logo_position = 'Haut gauche'
            
            conn.close()
            return settings
            
        except Exception:
            # En cas d'erreur, utiliser les paramètres par défaut
            settings = type('Settings', (), {})()
            settings.title_color = '#2c3e50'
            settings.header_color = '#34495e'
            settings.total_color = '#3498db'
            settings.result_color = '#2ecc71'
            settings.logo_path = ''
            settings.logo_position = 'Haut gauche'
            return settings
    
    def check_cir_projects(self):
        """Vérifie si au moins un projet a le CIR activé"""
        conn = get_connection()
        cursor = conn.cursor()
        
        try:
            # Récupérer les projets avec CIR activé
            placeholders = ','.join(['?'] * len(self.project_ids))
            cursor.execute(f"""
                SELECT COUNT(*) FROM projets 
                WHERE id IN ({placeholders}) AND cir = 1
            """, self.project_ids)
            
            count = cursor.fetchone()[0] or 0
            return count > 0
            
        except sqlite3.OperationalError:
            # Table n'existe pas ou colonne CIR manquante
            return False
        finally:
            conn.close()
    
    def format_currency(self, value, with_decimals=False):
        """Formate une valeur monétaire avec le formatage français :
        - Séparateur des milliers : espace
        - Arrondissement à l'entier le plus proche par défaut
        """
        if value == 0:
            return ""
        
        if with_decimals:
            # Formatage avec 2 décimales
            formatted = f"{value:,.2f}"
            # Remplacer la virgule (séparateur des milliers) par un espace
            # et le point (séparateur des décimales) par une virgule
            formatted = formatted.replace(",", "TEMP").replace(".", ",").replace("TEMP", " ")
        else:
            # Formatage sans décimales - arrondir à l'entier le plus proche
            rounded_value = round(value)
            formatted = f"{rounded_value:,}"
            # Remplacer la virgule (séparateur des milliers) par un espace
            formatted = formatted.replace(",", " ")
        
        return formatted
    
    def init_ui(self):
        """Initialise l'interface utilisateur"""
        layout = QVBoxLayout()
        
        # En-tête avec titre et informations
        header_layout = self.create_header()
        layout.addLayout(header_layout)
        
        # Pas de section de configuration
        
        # Tableau principal du compte de résultat
        self.table = QTableWidget()
        self.setup_table()
        layout.addWidget(self.table)
        
        # Boutons d'actions
        buttons_layout = self.create_buttons()
        layout.addLayout(buttons_layout)
        
        self.setLayout(layout)
    
    def create_header(self):
        """Crée l'en-tête avec titre et informations"""
        header_layout = QHBoxLayout()
        
        # Titre principal
        title = QLabel("COMPTE DE RÉSULTAT")
        title.setStyleSheet("font-size: 18pt; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Informations sur la sélection
        info_text = self.get_selection_info()
        info_label = QLabel(info_text)
        info_label.setStyleSheet("font-size: 10pt; color: #7f8c8d;")
        header_layout.addWidget(info_label)
        
        return header_layout
    
    def get_selection_info(self):
        """Génère le texte d'information sur la sélection"""
        info_lines = []
        
        # Projets
        if len(self.project_ids) == 1:
            project_name = self.get_project_name(self.project_ids[0])
            info_lines.append(f"Projet: {project_name}")
        else:
            info_lines.append(f"Projets: {len(self.project_ids)} sélectionnés")
        
        # Années
        if len(self.years) == 1:
            info_lines.append(f"Année: {self.years[0]}")
        else:
            years_str = f"{min(self.years)} - {max(self.years)}" if len(self.years) > 1 else str(self.years[0])
            info_lines.append(f"Période: {years_str}")
        
        # Granularité
        granularity_text = "Mensuel" if self.granularity == 'monthly' else "Annuel"
        info_lines.append(f"Granularité: {granularity_text}")
        
        # Type de coût
        cost_type_mapping = {
            'montant_charge': 'Montant chargé',
            'cout_production': 'Coût de production',
            'cout_complet': 'Coût complet'
        }
        cost_type_text = cost_type_mapping.get(self.cost_type, self.cost_type)
        info_lines.append(f"Type de coût: {cost_type_text}")
        
        return " | ".join(info_lines)
    
    def get_project_name(self, project_id):
        """Récupère le nom d'un projet"""
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT code, nom FROM projets WHERE id = ?", (project_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            code, nom = result
            return f"{code} - {nom}"
        return "Projet inconnu"
    
    def get_active_months_for_year(self, year):
        """Détermine les mois actifs pour une année donnée en fonction des dates du projet"""
        conn = get_connection()
        cursor = conn.cursor()
        
        try:
            # Récupérer les dates de début et fin de tous les projets
            active_months = set()
            has_project_without_dates = False
            
            for project_id in self.project_ids:
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                project_info = cursor.fetchone()
                
                if not project_info or not project_info[0] or not project_info[1]:
                    # Au moins un projet n'a pas de dates : on inclura tous les mois
                    has_project_without_dates = True
                    continue
                
                try:
                    debut_projet = datetime.datetime.strptime(project_info[0], '%m/%Y')
                    fin_projet = datetime.datetime.strptime(project_info[1], '%m/%Y')
                    
                    # Si l'année demandée est dans la période du projet
                    if debut_projet.year <= year <= fin_projet.year:
                        # Déterminer le mois de début pour cette année
                        start_month = debut_projet.month if year == debut_projet.year else 1
                        # Déterminer le mois de fin pour cette année
                        end_month = fin_projet.month if year == fin_projet.year else 12
                        
                        # Ajouter tous les mois actifs pour ce projet
                        for month in range(start_month, end_month + 1):
                            active_months.add(month)
                            
                except (ValueError, TypeError):
                    continue
            
            # Si au moins un projet n'a pas de dates, inclure tous les mois de l'année
            if has_project_without_dates:
                for month in range(1, 13):
                    active_months.add(month)
            
            return sorted(active_months)
            
        finally:
            conn.close()
    
    def setup_table(self):
        """Configure le tableau du compte de résultat"""
        # Le nombre de colonnes dépend de la granularité et des années
        if self.granularity == 'monthly':
            # Une colonne par mois actif pour chaque année
            columns = ["Poste"]
            for year in sorted(self.years):
                active_months = self.get_active_months_for_year(year)
                for month in active_months:
                    columns.append(f"{month:02d}/{year}")
            columns.append("TOTAL")
        else:
            # Une colonne par année
            columns = ["Poste"] + [str(year) for year in sorted(self.years)]
            if len(self.years) > 1:
                columns.append("TOTAL")
        
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        
        # Configuration de l'apparence
        # Définir une largeur minimum pour la première colonne (libellés)
        self.table.setColumnWidth(0, 250)  # Largeur fixe de 250 pixels pour les libellés
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        # Les autres colonnes s'adaptent automatiquement
        for col in range(1, len(columns)):
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.setAlternatingRowColors(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Rendre le tableau en lecture seule
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
    
    def create_buttons(self):
        """Crée les boutons d'actions"""
        buttons_layout = QHBoxLayout()
        
        # Export Excel
        excel_btn = QPushButton("Export Excel")
        excel_btn.clicked.connect(self.export_to_excel)
        excel_btn.setStyleSheet("QPushButton { background-color: #27ae60; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(excel_btn)
        
        # Export PDF
        pdf_btn = QPushButton("Export PDF")
        pdf_btn.clicked.connect(self.export_to_pdf)
        pdf_btn.setStyleSheet("QPushButton { background-color: #e74c3c; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(pdf_btn)
        
        # Paramètres Export
        settings_btn = QPushButton("Paramètres Export")
        settings_btn.clicked.connect(self.open_export_settings)
        settings_btn.setStyleSheet("QPushButton { background-color: #9b59b6; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(settings_btn)
        
        # Imprimer
        print_btn = QPushButton("Imprimer")
        print_btn.clicked.connect(self.print_compte_resultat)
        print_btn.setStyleSheet("QPushButton { background-color: #3498db; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(print_btn)
        
        buttons_layout.addStretch()
        
        # Fermer
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.accept)
        buttons_layout.addWidget(close_btn)
        
        return buttons_layout
    
    def open_export_settings(self):
        """Ouvre le dialogue des paramètres d'export"""
        dialog = ExportSettingsDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            QMessageBox.information(self, "Paramètres", "Paramètres d'export sauvegardés avec succès!")
    
    def load_data(self):
        """Charge et affiche les données du compte de résultat"""
        try:
            # Collecter toutes les données
            data = self.collect_financial_data()
            
            # Afficher dans le tableau
            self.populate_table(data)
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement des données: {str(e)}")
    
    def collect_financial_data(self):
        """Collecte toutes les données financières"""
        conn = get_connection()
        cursor = conn.cursor()
        
        try:
            data = {}
            
            for year in self.years:
                if self.granularity == 'monthly':
                    active_months = self.get_active_months_for_year(year)
                    for month in active_months:
                        period_key = f"{month:02d}/{year}"
                        data[period_key] = self.collect_period_data(cursor, year, month)
                else:
                    data[str(year)] = self.collect_period_data(cursor, year)
            
            return data
            
        finally:
            conn.close()
    
    def collect_period_data(self, cursor, year, month=None):
        """Collecte les données pour une période spécifique"""
        # Conditions de filtrage et définition des noms de mois
        month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                      "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        
        if month:
            # Filtrage mensuel - utilise les noms de mois français
            month_condition = f"AND mois = '{month_names[month-1]}'"
            year_condition = f"AND annee = {year}"
        else:
            # Filtrage annuel
            month_condition = ""
            year_condition = f"AND annee = {year}"
        
        project_condition = f"AND projet_id IN ({','.join(['?'] * len(self.project_ids))})"
        
        data = {
            # RECETTES
            'recettes': 0,
            'subventions': 0,
            
            # DÉPENSES
            'achats_sous_traitance': 0,
            'autres_achats': 0,
            'cout_direct': 0,
            'nb_jours_total': 0,
            'cout_moyen_par_jour': 0,
            'dotation_amortissements': 0,
            
            # CHARGES
            'charges_financieres': 0,
            'credit_impot': 0
        }
        
        # 1. RECETTES - NOUVELLE LOGIQUE avec redistribution automatique
        try:
            data['recettes'] = self.calculate_redistributed_recettes(cursor, self.project_ids, year, month)
        except sqlite3.OperationalError:
            # Table n'existe pas encore
            data['recettes'] = 0
        
        # 2. SUBVENTIONS - NOUVELLE LOGIQUE avec redistribution automatique
        try:
            subventions_total = 0
            # Importer la nouvelle méthode de calcul depuis SubventionDialog
            from subvention_dialog import SubventionDialog
            
            for project_id in self.project_ids:
                # Récupérer les informations du projet (dates, etc.)
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                projet_info = cursor.fetchone()
                
                # Ne pas ignorer les projets sans dates - la fonction de calcul gère ces cas
                # if not projet_info or not projet_info[0] or not projet_info[1]:
                #     continue  # Pas de dates de projet, skip
                
                # Calculer la subvention pour ce projet et cette période
                subvention_periode = self.calculate_smart_distributed_subvention(
                    cursor, project_id, year, month, projet_info
                )
                subventions_total += subvention_periode
            
            data['subventions'] = subventions_total
        except Exception as e:
            data['subventions'] = 0
        
        # 3. ACHATS ET SOUS-TRAITANCE - NOUVELLE LOGIQUE avec redistribution automatique
        try:
            achats_total = 0
            for project_id in self.project_ids:
                # Utiliser la nouvelle méthode de redistribution automatique
                montant_periode = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'depenses'
                )
                achats_total += montant_periode
            
            data['achats_sous_traitance'] = achats_total
        except Exception as e:
            data['achats_sous_traitance'] = 0
        
        # 4. AUTRES ACHATS - NOUVELLE LOGIQUE avec redistribution automatique
        try:
            autres_achats_total = 0
            for project_id in self.project_ids:
                # Utiliser la nouvelle méthode de redistribution automatique
                montant_periode = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'autres_depenses'
                )
                autres_achats_total += montant_periode
            
            data['autres_achats'] = autres_achats_total
        except Exception as e:
            data['autres_achats'] = 0
        
        # 5. COÛT DIRECT - temps_travail * (type de coût sélectionné) AVEC REDISTRIBUTION AUTOMATIQUE
        try:
            # NOUVELLE LOGIQUE : Utiliser la redistribution automatique pour chaque projet
            cout_direct_total = 0
            nb_jours_total = 0
            
            for project_id in self.project_ids:
                # Utiliser la nouvelle méthode avec redistribution automatique
                cout_project = self.calculate_redistributed_temps_travail(
                    cursor, project_id, year, month, self.cost_type
                )
                cout_direct_total += cout_project
                
                # Calculer aussi le nombre de jours avec la même logique
                nb_jours_project = self.calculate_redistributed_temps_travail_jours(
                    cursor, project_id, year, month
                )
                nb_jours_total += nb_jours_project
            
            data['cout_direct'] = cout_direct_total
            data['nb_jours_total'] = nb_jours_total
            
            # Calculer le coût moyen par jour
            if data['nb_jours_total'] > 0:
                data['cout_moyen_par_jour'] = data['cout_direct'] / data['nb_jours_total']
            else:
                data['cout_moyen_par_jour'] = 0
            
            # Vérification pour debug : si le coût direct est 0, vérifier pourquoi
            if data['cout_direct'] == 0:
                # Compter les entrées de temps de travail
                count_query = f"""
                    SELECT COUNT(*) FROM temps_travail t
                    WHERE t.annee = {year} {month_condition.replace('mois', 't.mois') if month_condition else ''} 
                    AND t.projet_id IN ({','.join(['?'] * len(self.project_ids))})
                """
                cursor.execute(count_query, self.project_ids)
                temps_count = cursor.fetchone()[0]
                
                # Compter les catégories de coût disponibles
                cursor.execute(f"SELECT COUNT(*) FROM categorie_cout WHERE annee = {year}")
                cat_count = cursor.fetchone()[0]
                if temps_count > 0:
                    # Vérifier les catégories qui ne matchent pas
                    cursor.execute(f"""
                        SELECT DISTINCT t.categorie 
                        FROM temps_travail t 
                        WHERE t.annee = {year} AND t.projet_id IN ({','.join(['?'] * len(self.project_ids))})
                        AND t.categorie NOT IN (SELECT libelle FROM categorie_cout WHERE annee = {year})
                    """, self.project_ids)
                    missing_cats = [row[0] for row in cursor.fetchall()]
                        
        except sqlite3.OperationalError as e:
            data['cout_direct'] = 0
            data['nb_jours_total'] = 0
            data['cout_moyen_par_jour'] = 0
        
        # 6. DOTATION AUX AMORTISSEMENTS
        try:
            amortissements_total = 0
            
            # Calculer les amortissements pour chaque projet
            for project_id in self.project_ids:
                amortissement_projet = self.calculate_amortissement_for_period(
                    cursor, project_id, year, month)
                amortissements_total += amortissement_projet
            
            data['dotation_amortissements'] = amortissements_total
        except sqlite3.OperationalError:
            data['dotation_amortissements'] = 0
        
        # 7. CRÉDIT D'IMPÔT RECHERCHE (CIR) - Calculé avec répartition équitable simple
        try:
            data['credit_impot'] = 0
            data['credit_impot_note'] = ""
            
            if self.has_cir_projects:
                # Calculer le CIR avec la nouvelle méthode de répartition équitable
                cir_result = self.calculate_distributed_cir(cursor, year, month)
                
                if cir_result == "CIR_NON_APPLICABLE":
                    data['credit_impot'] = 0
                    data['credit_impot_note'] = "CIR non applicable (subventions > dépenses éligibles)"
                else:
                    data['credit_impot'] = cir_result
                
        except sqlite3.OperationalError as e:
            # Erreur avec les tables - message d'alerte mais continuer
            if not hasattr(self, '_cir_error_shown'):
                QMessageBox.warning(self.parent, "Erreur CIR", 
                                  f"Erreur lors du calcul du CIR : {str(e)}\n"
                                  f"Le calcul du CIR sera ignoré.")
                self._cir_error_shown = True
            data['credit_impot'] = 0
        except Exception as e:
            data['credit_impot'] = 0
        
        return data
    
    def get_cost_type_label(self):
        """Retourne le libellé du type de coût sélectionné"""
        cost_type_mapping = {
            'montant_charge': 'Salaire (montant chargé)',
            'cout_production': 'Salaire (coût de production)',
            'cout_complet': 'Salaire (coût complet)'
        }
        return cost_type_mapping.get(self.cost_type, 'Salaire (coût direct)')
    
    def generate_filename(self, extension):
        """Génère un nom de fichier basé sur la configuration"""
        # 1. Partie projets
        if len(self.project_ids) <= 3:
            # Récupérer les codes des projets
            project_codes = []
            conn = get_connection()
            cursor = conn.cursor()
            try:
                for project_id in self.project_ids:
                    cursor.execute("SELECT code FROM projets WHERE id = ?", (project_id,))
                    result = cursor.fetchone()
                    if result and result[0]:
                        # Nettoyer le code projet pour le nom de fichier
                        clean_code = re.sub(r'[^\w\-_]', '', result[0])
                        project_codes.append(clean_code)
                    else:
                        project_codes.append(f"PROJ{project_id}")
            finally:
                conn.close()
            
            projects_part = "_".join(project_codes) if project_codes else "projets"
        else:
            projects_part = "multi_projets"
        
        # 2. Partie années
        if len(self.years) == 1:
            years_part = str(self.years[0])
        else:
            years_sorted = sorted(self.years)
            years_part = f"{years_sorted[0]}_{years_sorted[-1]}"
        
        # 3. Partie granularité
        granularity_part = "mensuel" if self.granularity == 'monthly' else "annuel"
        
        # 4. Partie type de coût (version courte pour le nom de fichier)
        cost_type_mapping = {
            'montant_charge': 'montant_charge',
            'cout_production': 'cout_production', 
            'cout_complet': 'cout_complet'
        }
        cost_part = cost_type_mapping.get(self.cost_type, 'cout_production')
        
        # 5. Assembler le nom de fichier
        filename = f"compte_resultat_{projects_part}_{years_part}_{granularity_part}_{cost_part}.{extension}"
        
        # 6. Nettoyer le nom de fichier pour Windows
        # Remplacer les caractères problématiques
        filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
        # Limiter la longueur (255 caractères max pour Windows)
        if len(filename) > 255:
            base_name = filename[:255-len(extension)-1]
            filename = f"{base_name}.{extension}"
        
        return filename

    def calculate_distributed_cir(self, cursor, target_year, target_month=None):
        """
        Calcule le CIR directement pour la période demandée selon la formule :
        CIR = [(salaire × k1) + (amortissements × k2) - subventions] × k3
        
        NOUVELLE LOGIQUE DIRECTE :
        - Calcul direct du CIR pour la période demandée (année ou mois)
        - Assiette éligible = (temps_travail × k1) + (amortissements × k2) - subventions
        - CIR = assiette_éligible × k3 (si assiette > 0, sinon CIR = 0)
        
        Retourne:
        - Un nombre positif si CIR applicable (c'est un produit)
        - 0 si pas de CIR ou si assiette éligible ≤ 0
        """
        try:
            # Récupérer les projets CIR
            cir_project_ids = []
            for project_id in self.project_ids:
                cursor.execute("SELECT cir FROM projets WHERE id = ? AND cir = 1", (project_id,))
                if cursor.fetchone():
                    cir_project_ids.append(project_id)
            
            if not cir_project_ids:
                return 0
            
            # Récupérer les coefficients CIR pour l'année cible
            cursor.execute('SELECT k1, k2, k3 FROM cir_coeffs WHERE annee = ?', (target_year,))
            cir_coeffs = cursor.fetchone()
            
            if not cir_coeffs:
                # Essayer de récupérer les coefficients d'une autre année si pas trouvé
                cursor.execute("SELECT k1, k2, k3 FROM cir_coeffs LIMIT 1")
                cir_coeffs = cursor.fetchone()
                
            if not cir_coeffs:
                return 0
            
            k1, k2, k3 = cir_coeffs
            
            # Calculer les éléments de l'assiette éligible pour la période
            period_montant_charge = 0
            period_amortissements = 0
            period_subventions = 0
            
            for project_id in cir_project_ids:
                # Vérifier que le projet est valide pour l'année cible
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                result = cursor.fetchone()
                if result and result[0] and result[1]:
                    try:
                        debut_year = int(result[0].split('/')[1])
                        fin_year = int(result[1].split('/')[1])
                        if not (debut_year <= target_year <= fin_year):
                            continue  # Projet pas actif cette année
                    except (ValueError, IndexError):
                        pass  # Si erreur de parsing, continuer quand même
                
                # 1. Temps de travail avec redistribution
                temps_travail = self.calculate_redistributed_temps_travail(
                    cursor, project_id, target_year, target_month, 'montant_charge'
                )
                period_montant_charge += temps_travail
                
                # 2. Amortissements avec redistribution
                if target_month:
                    # Mode mensuel
                    amortissements = self.calculate_amortissement_for_period(
                        cursor, project_id, target_year, target_month
                    )
                else:
                    # Mode annuel
                    projet_info = (result[0], result[1]) if result else None
                    amortissements = self.calculate_amortissement_for_year(
                        cursor, project_id, target_year, None, projet_info
                    )
                period_amortissements += amortissements
                
                # 3. Subventions pour la période
                try:
                    # Récupérer les informations du projet
                    cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                    projet_info = cursor.fetchone()
                    
                    if projet_info and projet_info[0] and projet_info[1]:
                        # Calculer les subventions pour cette période
                        subventions_periode = self.calculate_smart_distributed_subvention(
                            cursor, project_id, target_year, target_month, projet_info
                        )
                        period_subventions += subventions_periode
                except Exception:
                    pass  # Ignorer les erreurs de subvention
            
            # Calculer l'assiette éligible
            assiette_eligible = (period_montant_charge * k1) + (period_amortissements * k2) - period_subventions
            
            # Calculer le CIR
            if assiette_eligible > 0:
                cir = assiette_eligible * k3
                return max(0, cir)
            else:
                return 0
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return 0

    def populate_table(self, data):
        """Remplit le tableau avec les données"""
        # Structure du compte de résultat selon vos spécifications
        structure = [
            ("CHARGES", "header"),
            ("Achats et sous-traitance", "achats_sous_traitance"),
            ("Autres achats", "autres_achats"),
            ("Dotation aux amortissements", "dotation_amortissements"),
            (self.get_cost_type_label(), "cout_direct"),  # Nom dynamique selon le type de coût
            ("  - Nombre de jours TOTAL", "nb_jours_total"),
            ("  - Coût moyen par jour", "cout_moyen_par_jour"),
            ("TOTAL CHARGES", "total_charges"),
            ("", "separator"),
            ("PRODUITS", "header"),
            ("Chiffre d'affaires", "recettes"),
            ("Subventions", "subventions"),
            ("TOTAL PRODUITS", "total_produits"),
        ]
        
        # Ajouter la ligne CIR APRÈS le total produits si au moins un projet a le CIR activé
        if self.has_cir_projects:
            structure.append(("", "separator"))  # Ligne vide avant le CIR
            structure.append(("Crédit d'impôt recherche", "credit_impot"))
        
        structure.extend([
            ("", "separator"),
            ("RÉSULTAT FINANCIER", "resultat_financier")
        ])
        
        self.table.setRowCount(len(structure))
        
        # Colonnes de données (exclut la première colonne "Poste")
        data_columns = list(data.keys())
        if len(data_columns) > 1:
            data_columns.append("TOTAL")  # Ajouter une colonne total si plusieurs périodes
        
        for row, (label, data_key) in enumerate(structure):
            # Colonne des postes
            item = QTableWidgetItem(label)
            
            if data_key == "header":
                font = QFont()
                font.setBold(True)
                item.setFont(font)
                item.setBackground(QColor(52, 73, 94))
                item.setForeground(QColor(255, 255, 255))
            elif data_key.startswith("total_") or data_key.startswith("resultat_"):
                font = QFont()
                font.setBold(True)
                item.setFont(font)
                if "resultat" in data_key:
                    # Ligne RÉSULTAT FINANCIER entièrement en rouge
                    item.setBackground(QColor(231, 76, 60))
                    item.setForeground(QColor(255, 255, 255))
                elif data_key.startswith("total_"):
                    # Lignes TOTAL CHARGES et TOTAL PRODUITS en bleu
                    item.setBackground(QColor(52, 152, 219))
                    item.setForeground(QColor(255, 255, 255))
            
            self.table.setItem(row, 0, item)
            
            # Colonnes de données
            # Trier les clés chronologiquement pour la granularité mensuelle
            if self.granularity == 'monthly':
                # Créer une liste de tuples (année, mois, clé) pour un tri chronologique correct
                period_tuples = []
                for key in data.keys():
                    if '/' in key:  # Format "MM/YYYY"
                        parts = key.split('/')
                        month_num = int(parts[0])
                        year_num = int(parts[1])
                        period_tuples.append((year_num, month_num, key))
                
                # Trier par année puis par mois
                period_tuples.sort(key=lambda x: (x[0], x[1]))
                sorted_periods = [t[2] for t in period_tuples]
            else:
                # Pour la granularité annuelle, le tri alphabétique fonctionne
                sorted_periods = sorted(data.keys())
            
            for col, period in enumerate(sorted_periods, 1):
                if data_key == "separator" or data_key == "header":
                    # Lignes vides ou en-têtes
                    self.table.setItem(row, col, QTableWidgetItem(""))
                elif data_key.startswith("total_") or data_key.startswith("resultat_"):
                    # Calculs des totaux
                    value = self.calculate_total(data[period], data_key)
                    item = QTableWidgetItem(self.format_currency(value) if value != 0 else "")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                    
                    if "resultat_financier" in data_key:
                        # Ligne RÉSULTAT FINANCIER entièrement en rouge
                        item.setBackground(QColor(231, 76, 60))
                        item.setForeground(QColor(255, 255, 255))
                    elif "total_" in data_key:
                        item.setBackground(QColor(52, 152, 219))
                        item.setForeground(QColor(255, 255, 255))
                    
                    self.table.setItem(row, col, item)
                else:
                    # Données simples
                    value = data[period].get(data_key, 0)
                    
                    # Formatage spécial pour les nouveaux indicateurs
                    if data_key == "nb_jours_total":
                        # Afficher le nombre de jours sans décimales
                        item = QTableWidgetItem(f"{self.format_currency(value, False)} jours" if value != 0 else "")
                    elif data_key == "cout_moyen_par_jour":
                        # Afficher le coût moyen par jour avec 2 décimales et €/jour
                        item = QTableWidgetItem(f"{self.format_currency(value)} €/jour" if value != 0 else "")
                    elif data_key == "credit_impot":
                        # Formatage spécial pour le CIR avec gestion de la note explicative
                        note = data[period].get('credit_impot_note', "")
                        if note:
                            item = QTableWidgetItem(note)
                        else:
                            # Le CIR est affiché en NÉGATIF mais calculé positivement dans le résultat
                            cir_value = value
                            if cir_value != 0:
                                # Afficher en négatif pour indiquer que c'est un crédit
                                item = QTableWidgetItem(f"-{self.format_currency(abs(cir_value))}")
                            else:
                                item = QTableWidgetItem("")
                    else:
                        # Formatage normal pour les autres données
                        formatted_value = self.format_currency(value) if value != 0 else ""
                        item = QTableWidgetItem(formatted_value)
                    
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    
                    # Style spécial pour les indicateurs (légèrement en retrait visuellement)
                    if data_key in ["nb_jours_total", "cout_moyen_par_jour"]:
                        item.setForeground(QColor(100, 100, 100))  # Couleur légèrement grisée
                    
                    self.table.setItem(row, col, item)
            
            # Colonne TOTAL si plusieurs périodes
            if len(data) > 1 and col < self.table.columnCount() - 1:
                total_value = self.calculate_row_total(data, data_key)
                if total_value is not None:
                    # Formatage spécial pour les nouveaux indicateurs dans la colonne TOTAL
                    if data_key == "nb_jours_total":
                        item = QTableWidgetItem(f"{self.format_currency(total_value, False)} jours" if total_value != 0 else "")
                    elif data_key == "cout_moyen_par_jour":
                        item = QTableWidgetItem(f"{self.format_currency(total_value)} €/jour" if total_value != 0 else "")
                    elif data_key == "credit_impot":
                        # CIR affiché en négatif dans la colonne TOTAL aussi
                        if total_value != 0:
                            item = QTableWidgetItem(f"-{self.format_currency(abs(total_value))}")
                        else:
                            item = QTableWidgetItem("")
                    else:
                        item = QTableWidgetItem(self.format_currency(total_value) if total_value != 0 else "")
                    
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    
                    # Toute la colonne TOTAL en gras
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                    
                    # Style spécial pour les indicateurs
                    if data_key in ["nb_jours_total", "cout_moyen_par_jour"]:
                        item.setForeground(QColor(100, 100, 100))
                    elif data_key.startswith("total_") or data_key.startswith("resultat_"):
                        if "resultat_financier" in data_key:
                            # Ligne RÉSULTAT FINANCIER entièrement en rouge
                            item.setBackground(QColor(231, 76, 60))
                            item.setForeground(QColor(255, 255, 255))
                        elif data_key.startswith("total_"):
                            # Lignes TOTAL CHARGES et TOTAL PRODUITS en bleu
                            item.setBackground(QColor(52, 152, 219))
                            item.setForeground(QColor(255, 255, 255))
                    
                    self.table.setItem(row, self.table.columnCount() - 1, item)
    
    def calculate_total(self, period_data, total_type):
        """Calcule les totaux selon le type"""
        if total_type == "total_produits":
            # TOTAL PRODUITS = uniquement recettes + subventions (sans le CIR)
            total_produits = period_data.get('recettes', 0) + period_data.get('subventions', 0)
            return total_produits
        elif total_type == "total_charges":
            # TOTAL CHARGES = uniquement les vraies charges, sans le CIR
            total_charges = (period_data.get('achats_sous_traitance', 0) + 
                           period_data.get('autres_achats', 0) + 
                           period_data.get('cout_direct', 0) + 
                           period_data.get('dotation_amortissements', 0))
            
            # Ne pas inclure le CIR dans le total des charges
            return total_charges
        elif total_type == "resultat_financier":
            total_produits = self.calculate_total(period_data, "total_produits")
            total_charges = self.calculate_total(period_data, "total_charges")
            
            # Le résultat = produits - charges + CIR (le CIR améliore le résultat)
            resultat = total_produits - total_charges
            if self.has_cir_projects:
                cir_value = period_data.get('credit_impot', 0)
                resultat += abs(cir_value)  # Ajouter le CIR au résultat final
            
            return resultat
        
        return 0
    
    def calculate_row_total(self, all_data, data_key):
        """Calcule le total d'une ligne sur toutes les périodes"""
        if data_key == "separator" or data_key == "header":
            return None
        
        if data_key.startswith("total_") or data_key.startswith("resultat_"):
            # Pour les totaux calculés, sommer les résultats de chaque période
            total = 0
            for period_data in all_data.values():
                total += self.calculate_total(period_data, data_key)
            return total
        elif data_key == "cout_moyen_par_jour":
            # Pour le coût moyen par jour, recalculer la moyenne globale
            total_cout = 0
            total_jours = 0
            for period_data in all_data.values():
                total_cout += period_data.get('cout_direct', 0)
                total_jours += period_data.get('nb_jours_total', 0)
            
            return total_cout / total_jours if total_jours > 0 else 0
        else:
            # Pour les données simples, sommer directement
            total = 0
            for period_data in all_data.values():
                total += period_data.get(data_key, 0)
            return total
    
    def export_to_excel(self):
        """Exporte vers Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Charger les paramètres d'export
            settings = self.load_export_settings()
            
            # Générer le nom de fichier basé sur la configuration
            default_filename = self.generate_filename("xlsx")
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exporter le compte de résultat",
                default_filename,
                "Fichiers Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Compte de Résultat"
            
            # Convertir les couleurs hex en couleurs openpyxl (sans le #)
            def hex_to_openpyxl(hex_color):
                return hex_color.lstrip('#').upper()
            
            # Déterminer la structure selon la présence du logo
            if settings.logo_path and settings.logo_path.strip():
                # Titre
                title_cell = ws.cell(row=1, column=1)
                title_cell.value = f"COMPTE DE RÉSULTAT"
                title_color = hex_to_openpyxl(settings.title_color)
                title_cell.font = Font(bold=True, size=16, color=title_color)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=self.table.columnCount())
                
                # Sous-titre avec informations
                subtitle_cell = ws.cell(row=2, column=1)
                subtitle_cell.value = self.get_selection_info()
                subtitle_cell.font = Font(size=10, color="666666")
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=self.table.columnCount())
                
                # Essayer d'insérer le logo comme image
                try:
                    import os
                    from openpyxl.drawing.image import Image
                    
                    if os.path.exists(settings.logo_path):
                        logo_img = Image(settings.logo_path)
                        
                        # Redimensionner le logo (max 100px de hauteur)
                        max_height = 100
                        if logo_img.height > max_height:
                            ratio = max_height / logo_img.height
                            logo_img.height = max_height
                            logo_img.width = int(logo_img.width * ratio)
                        
                        # Position selon le paramètre
                        if settings.logo_position == "Haut droite":
                            logo_img.anchor = f"{chr(65 + self.table.columnCount() - 1)}3"  # Dernière colonne
                        elif settings.logo_position == "Haut centre":
                            middle_col = self.table.columnCount() // 2
                            logo_img.anchor = f"{chr(65 + middle_col)}3"
                        else:  # Haut gauche par défaut
                            logo_img.anchor = "A3"
                        
                        ws.add_image(logo_img)
                        
                        # Ajuster la hauteur des lignes pour le logo
                        ws.row_dimensions[3].height = max(60, logo_img.height * 0.75)
                        ws.row_dimensions[4].height = 20
                        
                        header_row = 5
                    else:
                        # Si le fichier n'existe pas, juste noter le nom
                        logo_note_cell = ws.cell(row=3, column=1)
                        logo_name = os.path.basename(settings.logo_path)
                        logo_note_cell.value = f"Logo configuré: {logo_name} (Position: {settings.logo_position})"
                        logo_note_cell.font = Font(italic=True, size=9, color="999999")
                        header_row = 4
                        
                except ImportError:
                    # Si openpyxl.drawing.image n'est pas disponible
                    logo_note_cell = ws.cell(row=3, column=1)
                    logo_note_cell.value = f"Logo: {os.path.basename(settings.logo_path)} (Position: {settings.logo_position})"
                    logo_note_cell.font = Font(italic=True, size=9, color="999999")
                    header_row = 4
                except Exception:
                    # Autre erreur avec le logo
                    header_row = 3
            else:
                header_row = 1
            
            # Exporter les en-têtes de colonnes à la bonne ligne
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                cell = ws.cell(row=header_row, column=col+1)
                cell.value = header_item.text() if header_item else ""
                cell.font = Font(bold=True, color="FFFFFF")
                header_color = hex_to_openpyxl(settings.header_color)
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Exporter les données du tableau
            data_start_row = header_row + 1
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    cell = ws.cell(row=data_start_row + row, column=col+1)
                    
                    if item and item.text():
                        # Pour la première colonne (postes), garder le texte
                        if col == 0:
                            cell.value = item.text()
                            # Appliquer le style selon le type de ligne
                            if "PRODUITS" in item.text() or "CHARGES" in item.text():
                                cell.font = Font(bold=True, color="FFFFFF")
                                header_color = hex_to_openpyxl(settings.header_color)
                                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                            elif "TOTAL" in item.text():
                                cell.font = Font(bold=True, color="FFFFFF")
                                total_color = hex_to_openpyxl(settings.total_color)
                                cell.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")
                            elif "RÉSULTAT" in item.text():
                                cell.font = Font(bold=True, color="FFFFFF")
                                result_color = hex_to_openpyxl(settings.result_color)
                                cell.fill = PatternFill(start_color=result_color, end_color=result_color, fill_type="solid")
                        else:
                            # Pour les colonnes de données, convertir en nombre si possible
                            try:
                                # Enlever les espaces et virgules pour la conversion
                                # Le texte contient maintenant des espaces comme séparateurs de milliers et des virgules comme décimales
                                value_str = item.text().replace(" ", "").replace(",", ".")
                                if value_str and value_str not in ["jours", "€/jour"]:
                                    # Enlever les unités si présentes
                                    value_str = value_str.replace(" jours", "").replace(" €/jour", "")
                                    cell.value = float(value_str)
                                    # Format français : espace pour milliers, virgule pour décimales
                                    cell.number_format = '# ##0,00'
                                else:
                                    cell.value = ""
                            except ValueError:
                                cell.value = item.text()
                            
                            cell.alignment = Alignment(horizontal='right')
                            
                            # Appliquer les couleurs aux colonnes de données selon le type de ligne
                            first_col_item = self.table.item(row, 0)
                            if first_col_item:
                                first_col_text = first_col_item.text()
                                if "PRODUITS" in first_col_text or "CHARGES" in first_col_text:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    header_color = hex_to_openpyxl(settings.header_color)
                                    cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                                elif "TOTAL" in first_col_text:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    total_color = hex_to_openpyxl(settings.total_color)
                                    cell.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")
                                elif "RÉSULTAT" in first_col_text:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    result_color = hex_to_openpyxl(settings.result_color)
                                    cell.fill = PatternFill(start_color=result_color, end_color=result_color, fill_type="solid")
            
            # Ajuster la largeur des colonnes
            try:
                for col_num in range(1, self.table.columnCount() + 1):
                    max_length = 0
                    for row_num in range(1, ws.max_row + 1):
                        try:
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell.value and not hasattr(cell, '_merge_parent'):  # Éviter les cellules fusionnées
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except:
                            continue
                    
                    # Calculer la lettre de colonne
                    if col_num <= 26:
                        column_letter = chr(64 + col_num)  # A, B, C, etc.
                    else:
                        column_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                    
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                # Si l'ajustement automatique échoue, utiliser des largeurs fixes
                for i in range(self.table.columnCount()):
                    column_letter = chr(65 + i)  # A, B, C, etc.
                    ws.column_dimensions[column_letter].width = 20
            
            wb.save(file_path)
            QMessageBox.information(self, "Export réussi", f"Fichier exporté: {file_path}")
            
        except ImportError:
            QMessageBox.critical(self, "Erreur", "Le module openpyxl n'est pas installé.\nInstallez-le avec: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export Excel: {str(e)}")
    
    def export_to_pdf(self):
        """Exporte vers PDF"""
        try:
            from PyQt6.QtGui import QTextDocument
            
            # Générer le nom de fichier basé sur la configuration
            default_filename = self.generate_filename("pdf")
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exporter le compte de résultat",
                default_filename,
                "Fichiers PDF (*.pdf)"
            )
            
            if not file_path:
                return
            
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            
            # Ne pas définir de marges personnalisées pour éviter les problèmes de compatibilité
            
            html_content = self.generate_html_content()
            
            document = QTextDocument()
            document.setHtml(html_content)
            document.print(printer)
            
            QMessageBox.information(self, "Export réussi", f"Fichier exporté: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export PDF: {str(e)}")
    
    def print_compte_resultat(self):
        """Imprime le compte de résultat"""
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            
            dialog = QPrintDialog(printer, self)
            if dialog.exec() != QPrintDialog.DialogCode.Accepted:
                return
            
            html_content = self.generate_html_content()
            
            from PyQt6.QtGui import QTextDocument
            document = QTextDocument()
            document.setHtml(html_content)
            document.print(printer)
            
            QMessageBox.information(self, "Impression", "Document envoyé à l'imprimante.")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'impression", f"Erreur lors de l'impression: {str(e)}")
    
    def generate_html_content(self):
        """Génère le contenu HTML pour l'export"""
        # Charger les paramètres d'export
        settings = self.load_export_settings()
        
        # Gérer le logo
        logo_html = ""
        if settings.logo_path and settings.logo_path.strip():
            import os
            import base64
            
            try:
                if os.path.exists(settings.logo_path):
                    with open(settings.logo_path, 'rb') as image_file:
                        encoded_string = base64.b64encode(image_file.read()).decode()
                    
                    # Déterminer le type MIME
                    ext = os.path.splitext(settings.logo_path)[1].lower()
                    mime_type = {
                        '.png': 'image/png',
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.gif': 'image/gif',
                        '.bmp': 'image/bmp'
                    }.get(ext, 'image/png')
                    
                    # Position du logo
                    position_style = {
                        'Haut gauche': 'float: left;',
                        'Haut droite': 'float: right;',
                        'Haut centre': 'display: block; margin: 0 auto;'
                    }.get(settings.logo_position, 'float: left;')
                    
                    logo_html = f"""
                    <div style="margin-bottom: 20px;">
                        <img src="data:{mime_type};base64,{encoded_string}" 
                             style="max-height: 80px; max-width: 200px; {position_style}" 
                             alt="Logo">
                    </div>
                    """
            except Exception:
                # Si le logo ne peut pas être chargé, on continue sans
                pass
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ text-align: center; color: {settings.title_color}; }}
                h2 {{ text-align: center; color: #7f8c8d; font-size: 12pt; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #bdc3c7; padding: 8px; text-align: left; }}
                th {{ background-color: {settings.header_color}; color: white; font-weight: bold; }}
                th.total-header {{ background-color: {settings.header_color}; color: white; font-weight: bold; border: 1px solid #bdc3c7; }}
                .header {{ background-color: {settings.header_color}; color: white; font-weight: bold; }}
                .total {{ background-color: {settings.total_color}; color: white; font-weight: bold; }}
                .result {{ background-color: {settings.result_color}; color: white; font-weight: bold; }}
                .amount {{ text-align: right; }}
                .total-column {{ font-weight: bold; text-align: right; }}
                .logo-container {{ overflow: auto; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            {logo_html}
            <div style="clear: both;"></div>
            <h1>COMPTE DE RÉSULTAT</h1>
            <h2>{self.get_selection_info()}</h2>
            <table>
        """
        
        # En-têtes de colonnes
        html += "<tr>"
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else ""
            # Utiliser une classe CSS pour l'en-tête de la colonne TOTAL
            if header_text == "TOTAL":
                html += f"<th class='total-header'>{header_text}</th>"
            else:
                html += f"<th>{header_text}</th>"
        html += "</tr>"
        
        # Données
        for row in range(self.table.rowCount()):
            html += "<tr>"
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                value = item.text() if item else ""
                
                # Vérifier si on est dans la colonne TOTAL
                header_item = self.table.horizontalHeaderItem(col)
                is_total_column = header_item and header_item.text() == "TOTAL"
                
                # Déterminer la classe CSS selon le contenu et la position
                css_class = ""
                first_col_item = self.table.item(row, 0)
                first_col_text = first_col_item.text() if first_col_item else ""
                
                if col == 0:  # Première colonne (libellés)
                    if first_col_text == "PRODUITS" or first_col_text == "CHARGES":
                        css_class = "header"
                    elif "TOTAL" in first_col_text:
                        css_class = "total"
                    elif "RÉSULTAT" in first_col_text:
                        css_class = "result"
                else:  # Colonnes de données
                    # Pour les lignes d'en-tête de sections simples (CHARGES/PRODUITS uniquement), 
                    # ne pas appliquer la couleur de fond aux cellules de données
                    if first_col_text == "PRODUITS" or first_col_text == "CHARGES":
                        # Cellules de données des lignes d'en-tête simples : style normal
                        css_class = "total-column" if is_total_column else "amount"
                    elif "TOTAL" in first_col_text:
                        # Lignes TOTAL : toute la ligne colorée
                        css_class = "total"
                    elif "RÉSULTAT" in first_col_text:
                        # Lignes RÉSULTAT : toute la ligne colorée
                        css_class = "result"
                    else:
                        # Autres lignes normales
                        css_class = "total-column" if is_total_column else "amount"
                
                html += f'<td class="{css_class}">{value}</td>'
            html += "</tr>"
        
        html += """
            </table>
        </body>
        </html>
        """
        
        return html
    
    def calculate_smart_distributed_subvention(self, cursor, project_id, year, month, projet_info):
        """
        Calcule les subventions avec une logique de répartition adaptée au compte de résultat :
        - Si dates de subvention définies : utilise la méthode SubventionDialog
        - Sinon : répartit intelligemment selon la granularité (mensuelle ou annuelle)
        """
        try:
            # Récupérer toutes les subventions pour ce projet
            cursor.execute('''
                SELECT nom, mode_simplifie, montant_forfaitaire, depenses_temps_travail, coef_temps_travail, 
                       depenses_externes, coef_externes, depenses_autres_achats, coef_autres_achats, 
                       depenses_dotation_amortissements, coef_dotation_amortissements, cd, taux,
                       date_debut_subvention, date_fin_subvention, montant_subvention_max, depenses_eligibles_max
                FROM subventions WHERE projet_id = ?
            ''', (project_id,))
            
            subventions_config = cursor.fetchall()
            if not subventions_config:
                return 0
            
            subvention_total_periode = 0
            
            # Importer SubventionDialog pour utiliser sa méthode de référence
            from subvention_dialog import SubventionDialog
            
            for subvention in subventions_config:
                # Construire le dictionnaire de données de subvention
                subvention_data = {
                    'nom': subvention[0],
                    'mode_simplifie': subvention[1] or 0,
                    'montant_forfaitaire': subvention[2] or 0,
                    'depenses_temps_travail': subvention[3] or 0,
                    'coef_temps_travail': subvention[4] or 1,
                    'depenses_externes': subvention[5] or 0,
                    'coef_externes': subvention[6] or 1,
                    'depenses_autres_achats': subvention[7] or 0,
                    'coef_autres_achats': subvention[8] or 1,
                    'depenses_dotation_amortissements': subvention[9] or 0,
                    'coef_dotation_amortissements': subvention[10] or 1,
                    'cd': subvention[11] or 1,
                    'taux': subvention[12] or 100,
                    'date_debut_subvention': subvention[13],
                    'date_fin_subvention': subvention[14],
                    'montant_subvention_max': subvention[15],
                    'depenses_eligibles_max': subvention[16]
                }
                
                # Utiliser la logique de redistribution du compte de résultat pour les subventions
                if subvention[13] and subvention[14]:  # dates de subvention définies
                    # Utiliser la méthode adaptée au compte de résultat qui comprend la redistribution
                    subvention_periode = self.calculate_subvention_with_redistribution(
                        cursor, project_id, subvention_data, year, month, subvention[13], subvention[14]
                    )
                else:
                    # Pas de dates de subvention définies : utiliser une répartition adaptée au compte de résultat
                    if month is not None:
                        # Granularité mensuelle : répartir la subvention annuelle sur les mois actifs
                        subvention_periode = self.calculate_monthly_subvention_fallback(
                            cursor, project_id, subvention_data, year, month, projet_info
                        )
                    else:
                        # Granularité annuelle : calculer pour toute l'année
                        subvention_periode = SubventionDialog.calculate_distributed_subvention(
                            project_id, subvention_data, year, None
                        )
                
                subvention_total_periode += subvention_periode
            
            return subvention_total_periode
            
        except Exception as e:
            # En cas d'erreur, afficher l'erreur pour débogage mais continuer
            print(f"Erreur calcul subvention projet {project_id}: {str(e)}")
            return 0
    
    def calculate_subvention_with_redistribution(self, cursor, project_id, subvention_data, year, month, date_debut, date_fin):
        """
        Calcule la subvention en utilisant la même logique de redistribution que le compte de résultat.
        
        Cette méthode :
        1. Vérifie si le mois est dans la période de subvention
        2. Utilise la redistribution automatique pour calculer les dépenses éligibles du mois
        3. Calcule la subvention proportionnellement à ces dépenses redistribuées
        """
        try:
            import datetime
            
            # Vérifier si le mois demandé est dans la période de subvention
            try:
                debut_subv = datetime.datetime.strptime(date_debut, '%m/%Y')
                fin_subv = datetime.datetime.strptime(date_fin, '%m/%Y')
                
                if month is not None:
                    target_date = datetime.datetime(year, month, 1)
                    if target_date < debut_subv or target_date > fin_subv:
                        return 0  # Mois hors période de subvention
                else:
                    # Mode annuel : vérifier qu'au moins un mois de l'année est dans la période
                    year_start = datetime.datetime(year, 1, 1)
                    year_end = datetime.datetime(year, 12, 1)
                    if year_end < debut_subv or year_start > fin_subv:
                        return 0
                        
            except ValueError:
                return 0
            
            # Calculer le montant total de la subvention sur toute la période
            montant_total_subvention = self.calculate_total_subvention_amount_with_redistribution(
                cursor, project_id, subvention_data, date_debut, date_fin
            )
            
            if montant_total_subvention <= 0:
                return 0
            
            # MODE SIMPLIFIÉ : Répartition temporelle (par nombre de mois)
            if subvention_data.get('mode_simplifie', 0):
                # Calculer le nombre total de mois de la période de subvention
                total_mois_subvention = (fin_subv.year - debut_subv.year) * 12 + (fin_subv.month - debut_subv.month) + 1
                
                # Déterminer combien de mois de la période cible sont dans la période de subvention
                if month is not None:
                    # Un seul mois demandé
                    target_date = datetime.datetime(year, month, 1)
                    if debut_subv <= target_date <= fin_subv:
                        mois_couverts = 1
                    else:
                        return 0
                else:
                    # Toute l'année demandée : compter les mois couverts
                    mois_couverts = 0
                    for mois in range(1, 13):
                        mois_date = datetime.datetime(year, mois, 1)
                        if debut_subv <= mois_date <= fin_subv:
                            mois_couverts += 1
                    
                    if mois_couverts == 0:
                        return 0
                
                # Répartir proportionnellement au nombre de mois
                proportion = mois_couverts / total_mois_subvention if total_mois_subvention > 0 else 0
                montant_reparti = montant_total_subvention * proportion
                
                return montant_reparti
            
            # MODE DÉTAILLÉ : Répartition proportionnelle aux dépenses éligibles
            # Calculer les dépenses éligibles totales sur la période de subvention (avec redistribution)
            depenses_eligibles_totales = self.calculate_total_eligible_expenses_with_redistribution(
                cursor, project_id, subvention_data, date_debut, date_fin
            )
            
            if depenses_eligibles_totales <= 0:
                return 0
            
            # Calculer les dépenses éligibles pour la période demandée (avec redistribution)
            depenses_eligibles_periode = self.calculate_period_eligible_expenses_with_redistribution(
                cursor, project_id, subvention_data, year, month
            )
            
            # Calculer la proportion et retourner le montant réparti
            proportion = depenses_eligibles_periode / depenses_eligibles_totales
            montant_reparti = montant_total_subvention * proportion
            
            return montant_reparti
            
        except Exception as e:
            return 0
    
    def calculate_total_subvention_amount_with_redistribution(self, cursor, project_id, subvention_data, date_debut, date_fin):
        """Calcule le montant total de la subvention en utilisant les dépenses redistribuées"""
        try:
            import datetime
            
            # Mode de calcul
            if subvention_data.get('mode_simplifie', 0):
                return float(subvention_data.get('montant_forfaitaire', 0))
            
            # Mode détaillé : calculer sur toute la période de subvention
            debut_subv = datetime.datetime.strptime(date_debut, '%m/%Y')
            fin_subv = datetime.datetime.strptime(date_fin, '%m/%Y')
            
            assiette_totale = 0
            
            # Parcourir tous les mois de la période de subvention
            current_date = debut_subv
            while current_date <= fin_subv:
                # Calculer les dépenses éligibles pour ce mois avec redistribution
                depenses_mois = self.calculate_period_eligible_expenses_with_redistribution(
                    cursor, project_id, subvention_data, current_date.year, current_date.month
                )
                assiette_totale += depenses_mois
                
                # Passer au mois suivant
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            # Appliquer les plafonds si définis
            depenses_eligibles_max = subvention_data.get('depenses_eligibles_max', 0)
            if depenses_eligibles_max and depenses_eligibles_max > 0:
                assiette_totale = min(assiette_totale, depenses_eligibles_max)
            
            # Calculer la subvention
            taux = subvention_data.get('taux', 100) / 100
            montant_subvention = assiette_totale * taux
            
            # Appliquer le plafond du montant max si défini
            montant_subvention_max = subvention_data.get('montant_subvention_max', 0)
            if montant_subvention_max and montant_subvention_max > 0:
                montant_subvention = min(montant_subvention, montant_subvention_max)
            
            return montant_subvention
            
        except Exception as e:
            return 0
    
    def calculate_total_eligible_expenses_with_redistribution(self, cursor, project_id, subvention_data, date_debut, date_fin):
        """Calcule les dépenses éligibles totales avec redistribution"""
        try:
            import datetime
            
            debut_subv = datetime.datetime.strptime(date_debut, '%m/%Y')
            fin_subv = datetime.datetime.strptime(date_fin, '%m/%Y')
            
            depenses_totales = 0
            
            # Parcourir tous les mois de la période de subvention
            current_date = debut_subv
            while current_date <= fin_subv:
                # Utiliser la méthode auxiliaire qui ne fait pas de vérification de dates
                depenses_mois = self._calculate_eligible_expenses_for_month(
                    cursor, project_id, subvention_data, current_date.year, current_date.month
                )
                depenses_totales += depenses_mois
                
                # Passer au mois suivant
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            return depenses_totales
            
        except Exception as e:
            return 0
    
    def calculate_period_eligible_expenses_with_redistribution(self, cursor, project_id, subvention_data, year, month):
        """Calcule les dépenses éligibles pour une période en utilisant la redistribution du compte de résultat"""
        try:
            import datetime
            
            # Récupérer les dates de subvention pour filtrer correctement
            date_debut_subv = subvention_data.get('date_debut_subvention')
            date_fin_subv = subvention_data.get('date_fin_subvention')
            
            if not date_debut_subv or not date_fin_subv:
                # Si pas de dates de subvention, utiliser les dates du projet
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id=?", (project_id,))
                projet_dates = cursor.fetchone()
                if not projet_dates or not projet_dates[0] or not projet_dates[1]:
                    return 0
                date_debut_subv, date_fin_subv = projet_dates[0], projet_dates[1]
            
            try:
                debut_subv = datetime.datetime.strptime(date_debut_subv, '%m/%Y')
                fin_subv = datetime.datetime.strptime(date_fin_subv, '%m/%Y')
            except ValueError:
                return 0
            
            depenses_eligibles = 0
            
            if month is not None:
                # Mode mensuel : calculer pour le mois demandé s'il est dans la période de subvention
                target_date = datetime.datetime(year, month, 1)
                if target_date < debut_subv or target_date > fin_subv:
                    return 0  # Mois hors période de subvention
                    
                # Calculer les dépenses éligibles pour ce mois
                depenses_eligibles = self._calculate_eligible_expenses_for_month(
                    cursor, project_id, subvention_data, year, month
                )
            else:
                # Mode annuel : calculer pour tous les mois de l'année qui sont dans la période de subvention
                year_start = datetime.datetime(year, 1, 1)
                year_end = datetime.datetime(year, 12, 1)
                
                # Calculer les bornes d'intersection entre l'année et la période de subvention
                mois_debut = max(1, debut_subv.month if debut_subv.year == year else 1)
                mois_fin = min(12, fin_subv.month if fin_subv.year == year else 12)
                
                # Si l'année n'intersecte pas avec la période de subvention
                if year < debut_subv.year or year > fin_subv.year:
                    return 0
                
                # Si l'année intersecte partiellement, ajuster les bornes
                if year == debut_subv.year:
                    mois_debut = debut_subv.month
                if year == fin_subv.year:
                    mois_fin = fin_subv.month
                
                # Calculer la somme des dépenses éligibles pour tous les mois concernés
                for mois in range(mois_debut, mois_fin + 1):
                    depenses_mois = self._calculate_eligible_expenses_for_month(
                        cursor, project_id, subvention_data, year, mois
                    )
                    depenses_eligibles += depenses_mois
            
            return depenses_eligibles
            
        except Exception as e:
            return 0
    
    def _calculate_eligible_expenses_for_month(self, cursor, project_id, subvention_data, year, month):
        """Calcule les dépenses éligibles pour un mois donné"""
        try:
            depenses_eligibles = 0
            
            # Temps de travail éligible avec redistribution
            if subvention_data.get('depenses_temps_travail', 0):
                coef_temps = subvention_data.get('coef_temps_travail', 1.0)
                cd = subvention_data.get('cd', 1.0)
                
                # Utiliser la même logique de redistribution que le compte de résultat
                cout_temps_travail = self.calculate_redistributed_temps_travail(
                    cursor, project_id, year, month, 'montant_charge'
                )
                depenses_eligibles += cout_temps_travail * coef_temps * cd
            
            # Dépenses externes éligibles avec redistribution
            if subvention_data.get('depenses_externes', 0):
                coef_externes = subvention_data.get('coef_externes', 1.0)
                depenses_externes = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'depenses'
                )
                depenses_eligibles += depenses_externes * coef_externes
            
            # Autres dépenses éligibles avec redistribution
            if subvention_data.get('depenses_autres_achats', 0):
                coef_autres = subvention_data.get('coef_autres_achats', 1.0)
                autres_depenses = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'autres_depenses'
                )
                depenses_eligibles += autres_depenses * coef_autres
            
            # Amortissements éligibles
            if subvention_data.get('depenses_dotation_amortissements', 0):
                coef_amort = subvention_data.get('coef_dotation_amortissements', 1.0)
                amortissements = self.calculate_amortissement_for_period(
                    cursor, project_id, year, month
                )
                depenses_eligibles += amortissements * coef_amort
            
            return depenses_eligibles
            
        except Exception as e:
            return 0
    
    def calculate_monthly_subvention_fallback(self, cursor, project_id, subvention_data, year, month, projet_info):
        """
        Méthode de fallback pour la répartition mensuelle des subventions 
        quand les dates de subvention ne sont pas définies.
        
        Stratégie :
        1. Calculer le montant total annuel de la subvention
        2. Déterminer les mois actifs du projet pour cette année
        3. Répartir équitablement sur tous les mois actifs
        """
        try:
            # Importer SubventionDialog pour utiliser ses méthodes
            from subvention_dialog import SubventionDialog
            
            # 1. Calculer le montant total annuel de cette subvention
            subvention_annuelle = SubventionDialog.calculate_distributed_subvention(
                project_id, subvention_data, year, None
            )
            
            if subvention_annuelle <= 0:
                return 0
            
            # 2. Déterminer les mois actifs pour cette année
            # Si le projet n'a pas de dates, considérer tous les mois de l'année comme actifs
            if not projet_info or not projet_info[0] or not projet_info[1]:
                # Pas de dates de projet : répartir sur tous les 12 mois de l'année
                active_months = list(range(1, 13))
            else:
                # Projet avec dates : utiliser la méthode normale
                active_months = self.get_active_months_for_year(year)
            
            if not active_months or month not in active_months:
                return 0
            
            # 3. Répartir équitablement sur tous les mois actifs
            subvention_mensuelle = subvention_annuelle / len(active_months)
            
            return subvention_mensuelle
            
        except Exception as e:
            print(f"Erreur calcul monthly subvention fallback: {str(e)}")
            return 0
    
    def calculate_proportional_subvention_detailed(self, cursor, project_id, subvention, year, month, projet_info):
        """
        Calcule la subvention en mode détaillé avec répartition proportionnelle basée sur les vraies dépenses
        """
        try:
            # Extraire les données de la subvention
            depenses_temps_travail = subvention[3] or 0
            coef_temps_travail = subvention[4] or 1
            depenses_externes = subvention[5] or 0
            coef_externes = subvention[6] or 1
            depenses_autres_achats = subvention[7] or 0
            coef_autres_achats = subvention[8] or 1
            depenses_dotation_amortissements = subvention[9] or 0
            coef_dotation_amortissements = subvention[10] or 1
            cd = subvention[11] or 1
            taux = subvention[12] or 100
            date_debut_subv = subvention[13] or projet_info[0]
            date_fin_subv = subvention[14] or projet_info[1]
            montant_subvention_max = subvention[15] or None
            depenses_eligibles_max = subvention[16] or None
            
            
            # 1. Calculer le montant total de la subvention sur la période de subvention
            montant_total_subvention = self.calculate_total_subvention_amount(
                cursor, project_id, subvention, date_debut_subv, date_fin_subv, montant_subvention_max, depenses_eligibles_max
            )
            
            if montant_total_subvention <= 0:
                return 0
            
            # 2. Calculer les dépenses éligibles totales sur la période de subvention
            depenses_eligibles_totales = self.calculate_total_eligible_expenses(
                cursor, project_id, subvention, date_debut_subv, date_fin_subv
            )
            
            if depenses_eligibles_totales <= 0:
                return 0
            
            # 3. Calculer les dépenses éligibles pour la période demandée
            depenses_eligibles_periode = self.calculate_period_eligible_expenses(
                cursor, project_id, subvention, year, month
            )
            
            # 4. Calculer la proportion et répartir
            proportion = depenses_eligibles_periode / depenses_eligibles_totales
            subvention_repartie = montant_total_subvention * proportion
            
            return subvention_repartie
            
        except Exception as e:
            return 0
    
    def calculate_total_subvention_amount(self, cursor, project_id, subvention, date_debut_subv, date_fin_subv, montant_subvention_max=None, depenses_eligibles_max=None):
        """Calcule le montant total de la subvention sur sa période"""
        try:
            # Extraire les paramètres
            depenses_temps_travail = subvention[3] or 0
            coef_temps_travail = subvention[4] or 1
            depenses_externes = subvention[5] or 0
            coef_externes = subvention[6] or 1
            depenses_autres_achats = subvention[7] or 0
            coef_autres_achats = subvention[8] or 1
            depenses_dotation_amortissements = subvention[9] or 0
            coef_dotation_amortissements = subvention[10] or 1
            cd = subvention[11] or 1
            taux = subvention[12] or 100
            
            # Calculer l'assiette totale
            assiette_totale = 0
            
            # Temps de travail
            if depenses_temps_travail:
                cout_temps_total = self.calculate_temps_travail_for_period(
                    cursor, project_id, date_debut_subv, date_fin_subv
                )
                assiette_totale += cout_temps_total * coef_temps_travail
            
            # Dépenses externes
            if depenses_externes:
                depenses_ext_total = self.calculate_depenses_for_period(
                    cursor, project_id, 'depenses', date_debut_subv, date_fin_subv
                )
                assiette_totale += depenses_ext_total * coef_externes
            
            # Autres dépenses
            if depenses_autres_achats:
                autres_dep_total = self.calculate_depenses_for_period(
                    cursor, project_id, 'autres_depenses', date_debut_subv, date_fin_subv
                )
                assiette_totale += autres_dep_total * coef_autres_achats
            
            # Amortissements
            if depenses_dotation_amortissements:
                amort_total = self.calculate_amortissements_for_period(
                    cursor, project_id, date_debut_subv, date_fin_subv
                )
                assiette_totale += amort_total * coef_dotation_amortissements
            
            # Appliquer le plafond du coût éligible max à l'assiette totale
            assiette_plafonnee = assiette_totale
            if depenses_eligibles_max and depenses_eligibles_max > 0:
                assiette_plafonnee = min(assiette_totale, depenses_eligibles_max)
            
            # Calculer la subvention avec cd et taux sur l'assiette plafonnée
            montant_subvention = assiette_plafonnee * cd * (taux / 100)
            
            # Appliquer le plafond du montant max si défini
            if montant_subvention_max and montant_subvention_max > 0:
                montant_subvention = min(montant_subvention, montant_subvention_max)
            
            return montant_subvention
            
        except Exception as e:
            return 0
    
    def calculate_total_eligible_expenses(self, cursor, project_id, subvention, date_debut_subv, date_fin_subv):
        """Calcule les dépenses éligibles totales sur la période de subvention"""
        try:
            # Extraire les paramètres
            depenses_temps_travail = subvention[3] or 0
            coef_temps_travail = subvention[4] or 1
            depenses_externes = subvention[5] or 0
            coef_externes = subvention[6] or 1
            depenses_autres_achats = subvention[7] or 0
            coef_autres_achats = subvention[8] or 1
            depenses_dotation_amortissements = subvention[9] or 0
            coef_dotation_amortissements = subvention[10] or 1
            
            depenses_eligibles = 0
            
            # Temps de travail éligible
            if depenses_temps_travail:
                cout_temps = self.calculate_temps_travail_for_period(
                    cursor, project_id, date_debut_subv, date_fin_subv
                )
                depenses_eligibles += cout_temps * coef_temps_travail
            
            # Dépenses externes éligibles
            if depenses_externes:
                depenses_ext = self.calculate_depenses_for_period(
                    cursor, project_id, 'depenses', date_debut_subv, date_fin_subv
                )
                depenses_eligibles += depenses_ext * coef_externes
            
            # Autres dépenses éligibles
            if depenses_autres_achats:
                autres_dep = self.calculate_depenses_for_period(
                    cursor, project_id, 'autres_depenses', date_debut_subv, date_fin_subv
                )
                depenses_eligibles += autres_dep * coef_autres_achats
            
            # Amortissements éligibles
            if depenses_dotation_amortissements:
                amort = self.calculate_amortissements_for_period(
                    cursor, project_id, date_debut_subv, date_fin_subv
                )
                depenses_eligibles += amort * coef_dotation_amortissements
            
            return depenses_eligibles
            
        except Exception as e:
            return 0
    
    def calculate_period_eligible_expenses(self, cursor, project_id, subvention, year, month):
        """Calcule les dépenses éligibles pour la période demandée (mois ou année)"""
        try:
            # Extraire les paramètres
            depenses_temps_travail = subvention[3] or 0
            coef_temps_travail = subvention[4] or 1
            depenses_externes = subvention[5] or 0
            coef_externes = subvention[6] or 1
            depenses_autres_achats = subvention[7] or 0
            coef_autres_achats = subvention[8] or 1
            depenses_dotation_amortissements = subvention[9] or 0
            coef_dotation_amortissements = subvention[10] or 1
            
            depenses_eligibles = 0
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            # DEBUG pour projet test
            if project_id == 4:
                print(f"    CALCUL DÉPENSES ÉLIGIBLES PÉRIODE - {year}/{month if month else 'annuel'}")
            
            # Temps de travail éligible pour la période
            if depenses_temps_travail:
                if month is not None:
                    # Mois spécifique
                    cursor.execute(f"""
                        SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                        FROM temps_travail t
                        JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                        WHERE t.annee = ? AND t.mois = ? AND t.projet_id = ?
                    """, (year, month_names[month-1], project_id))
                else:
                    # Année complète
                    cursor.execute(f"""
                        SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                        FROM temps_travail t
                        JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                        WHERE t.annee = ? AND t.projet_id = ?
                    """, (year, project_id))
                
                cout_temps = cursor.fetchone()[0] or 0
                cout_temps_eligible = cout_temps * coef_temps_travail
                depenses_eligibles += cout_temps_eligible
                
                if project_id == 4:
                    print(f"      Temps travail: {cout_temps} × {coef_temps_travail} = {cout_temps_eligible}")
            
            # Dépenses externes éligibles pour la période
            if depenses_externes:
                depenses_ext = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'depenses'
                )
                depenses_ext_eligible = depenses_ext * coef_externes
                depenses_eligibles += depenses_ext_eligible
                
                if project_id == 4:
                    print(f"      Dépenses externes: {depenses_ext} × {coef_externes} = {depenses_ext_eligible}")
            
            # Autres dépenses éligibles pour la période
            if depenses_autres_achats:
                autres_dep = self.calculate_redistributed_expenses(
                    cursor, project_id, year, month, 'autres_depenses'
                )
                autres_dep_eligible = autres_dep * coef_autres_achats
                depenses_eligibles += autres_dep_eligible
                
                if project_id == 4:
                    print(f"      Autres dépenses: {autres_dep} × {coef_autres_achats} = {autres_dep_eligible}")
            
            # Amortissements éligibles pour la période
            if depenses_dotation_amortissements:
                amort = self.calculate_amortissement_for_period(cursor, project_id, year, month)
                amort_eligible = amort * coef_dotation_amortissements
                depenses_eligibles += amort_eligible
                
                if project_id == 4:
                    print(f"      Amortissements: {amort} × {coef_dotation_amortissements} = {amort_eligible}")
            
            if project_id == 4:
                print(f"      TOTAL ÉLIGIBLE PÉRIODE: {depenses_eligibles}")
            
            return depenses_eligibles
            
        except Exception as e:
            print(f"Erreur dans calculate_period_eligible_expenses: {e}")
            return 0
    
    def calculate_temps_travail_for_period(self, cursor, project_id, date_debut, date_fin):
        """Calcule le coût du temps de travail pour une période donnée"""
        try:
            debut = datetime.datetime.strptime(date_debut, '%m/%Y')
            fin = datetime.datetime.strptime(date_fin, '%m/%Y')
            
            cout_total = 0
            current_date = debut
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            while current_date <= fin:
                cursor.execute(f"""
                    SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                    FROM temps_travail t
                    JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                    WHERE t.annee = ? AND t.mois = ? AND t.projet_id = ?
                """, (current_date.year, month_names[current_date.month-1], project_id))
                
                cout_mois = cursor.fetchone()[0] or 0
                cout_total += cout_mois
                
                # Passer au mois suivant
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            return cout_total
            
        except Exception as e:
            print(f"Erreur dans calculate_temps_travail_for_period: {e}")
            return 0
    
    def calculate_depenses_for_period(self, cursor, project_id, table_name, date_debut, date_fin):
        """Calcule les dépenses pour une période donnée"""
        try:
            debut = datetime.datetime.strptime(date_debut, '%m/%Y')
            fin = datetime.datetime.strptime(date_fin, '%m/%Y')
            
            total_depenses = 0
            current_date = debut
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            while current_date <= fin:
                # Utiliser la méthode de redistribution automatique
                depenses_mois = self.calculate_redistributed_expenses(
                    cursor, project_id, current_date.year, current_date.month, table_name
                )
                total_depenses += depenses_mois
                
                # Passer au mois suivant
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            return total_depenses
            
        except Exception as e:
            print(f"Erreur dans calculate_depenses_for_period: {e}")
            return 0
    
    def calculate_amortissements_for_period(self, cursor, project_id, date_debut, date_fin):
        """Calcule les amortissements pour une période donnée"""
        try:
            debut = datetime.datetime.strptime(date_debut, '%m/%Y')
            fin = datetime.datetime.strptime(date_fin, '%m/%Y')
            
            total_amort = 0
            current_date = debut
            
            while current_date <= fin:
                amort_mois = self.calculate_amortissement_for_period(
                    cursor, project_id, current_date.year, current_date.month
                )
                total_amort += amort_mois
                
                # Passer au mois suivant
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
            
            return total_amort
            
        except Exception as e:
            return 0
    
    def calculate_monthly_distributed_subvention(self, cursor, project_id, subvention_data, year, month, projet_info):
        """
        Calcule la subvention pour un mois en appliquant une répartition intelligente :
        - Si la subvention a des dépenses spécifiques dans ce mois : utilise la méthode proportionnelle
        - Sinon : répartit équitablement sur tous les mois actifs de l'année
        """
        try:
            # D'abord essayer la méthode de distribution proportionnelle
            from subvention_dialog import SubventionDialog
            subvention_proportionnelle = SubventionDialog.calculate_distributed_subvention(
                project_id, subvention_data, year, month
            )
            
            # Si on obtient un résultat avec la méthode proportionnelle, l'utiliser
            if subvention_proportionnelle > 0:
                return subvention_proportionnelle
            
            # Sinon, appliquer la répartition équitable sur les mois actifs
            # Calculer le total annuel de cette subvention
            subvention_annuelle = SubventionDialog.calculate_distributed_subvention(
                project_id, subvention_data, year, None
            )
            
            if subvention_annuelle <= 0:
                return 0
                
            # Calculer les mois actifs pour cette année
            active_months = self.get_active_months_for_year(year)
            if not active_months or month not in active_months:
                return 0
                
            # Répartir équitablement sur tous les mois actifs
            subvention_mensuelle = subvention_annuelle / len(active_months)
            return subvention_mensuelle
            
        except Exception as e:
            return 0
    
    def calculate_smart_distributed_cir(self, cursor, year, month):
        """
        Calcule le CIR avec une répartition intelligente :
        - Si mode mensuel : répartit le CIR proportionnellement aux dépenses éligibles du mois
        - Si mode annuel : retourne le CIR total de l'année
        """
        try:
            # Récupérer les projets CIR
            cir_project_ids = []
            for project_id in self.project_ids:
                cursor.execute("SELECT cir FROM projets WHERE id = ? AND cir = 1", (project_id,))
                if cursor.fetchone():
                    cir_project_ids.append(project_id)
            
            if not cir_project_ids:
                return 0
            
            # Calculer les coefficients CIR
            k1, k2, k3 = None, None, None
            for test_year in sorted(self.years):
                cursor.execute('SELECT k1, k2, k3 FROM cir_coeffs WHERE annee = ?', (test_year,))
                cir_coeffs = cursor.fetchone()
                if cir_coeffs:
                    k1, k2, k3 = cir_coeffs
                    break
            
            if not k1:
                return 0
            
            if month is not None:
                # Mode mensuel : calculer la répartition basée sur les dépenses éligibles du mois
                return self.calculate_monthly_distributed_cir(cursor, cir_project_ids, year, month, k1, k2, k3)
            else:
                # Mode annuel : calculer le CIR total pour l'année
                return self.calculate_annual_distributed_cir(cursor, cir_project_ids, year, k1, k2, k3)
                
        except Exception as e:
            return 0
    
    def calculate_monthly_distributed_cir(self, cursor, cir_project_ids, year, month, k1, k2, k3):
        """Calcule le CIR pour un mois spécifique en répartissant intelligemment"""
        try:
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            # 1. Calculer le CIR total annuel pour tous les projets CIR
            cir_annuel_total = self.calculate_annual_distributed_cir(cursor, cir_project_ids, year, k1, k2, k3)
            
            if cir_annuel_total <= 0:
                return 0
            
            # 2. Calculer les dépenses éligibles totales de l'année
            depenses_eligibles_annuelles = 0
            for project_id in cir_project_ids:
                # Vérifier que le projet est actif pour cette année
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                projet_info = cursor.fetchone()
                if not projet_info or not projet_info[0] or not projet_info[1]:
                    continue
                    
                try:
                    debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
                    fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
                    
                    if year < debut_projet.year or year > fin_projet.year:
                        continue
                        
                    # Coûts directs (temps de travail) de l'année
                    cursor.execute(f"""
                        SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                        FROM temps_travail t
                        JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                        WHERE t.annee = ? AND t.projet_id = ?
                    """, (year, project_id))
                    cout_temps_annuel = cursor.fetchone()[0] or 0
                    
                    # Amortissements de l'année
                    amort_annuel = self.calculate_amortissement_for_period(cursor, project_id, year)
                    
                    # Dépenses éligibles = (temps * k1) + (amortissements * k2)
                    depenses_eligibles_annuelles += (cout_temps_annuel * k1) + (amort_annuel * k2)
                    
                except Exception as e:
                    continue
            
            if depenses_eligibles_annuelles <= 0:
                return 0
            
            # 3. Calculer les dépenses éligibles du mois cible
            depenses_eligibles_mois = 0
            for project_id in cir_project_ids:
                # Vérifier que le projet est actif pour ce mois
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                projet_info = cursor.fetchone()
                if not projet_info or not projet_info[0] or not projet_info[1]:
                    continue
                    
                try:
                    debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
                    fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
                    
                    target_date = datetime.datetime(year, month, 1)
                    if target_date < debut_projet or target_date > fin_projet:
                        continue
                        
                    # Coûts directs (temps de travail) du mois
                    cursor.execute(f"""
                        SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                        FROM temps_travail t
                        JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                        WHERE t.annee = ? AND t.mois = ? AND t.projet_id = ?
                    """, (year, month_names[month-1], project_id))
                    cout_temps_mois = cursor.fetchone()[0] or 0
                    
                    # Amortissements du mois
                    amort_mois = self.calculate_amortissement_for_period(cursor, project_id, year, month)
                    
                    # Dépenses éligibles = (temps * k1) + (amortissements * k2)
                    depenses_eligibles_mois += (cout_temps_mois * k1) + (amort_mois * k2)
                    
                except Exception as e:
                    continue
            
            # 4. Calculer la proportion et répartir le CIR
            if depenses_eligibles_annuelles > 0:
                proportion = depenses_eligibles_mois / depenses_eligibles_annuelles
                cir_mois = cir_annuel_total * proportion
                return cir_mois
            else:
                # Si pas de dépenses éligibles annuelles, répartir équitablement sur les mois actifs
                active_months = self.get_active_months_for_year(year)
                if active_months and month in active_months:
                    return cir_annuel_total / len(active_months)
                else:
                    return 0
                    
        except Exception as e:
            return 0
    
    def calculate_annual_distributed_cir(self, cursor, cir_project_ids, year, k1, k2, k3):
        """Calcule le CIR total pour une année"""
        try:
            # Calculer les totaux pour l'année
            total_montant_charge = 0
            total_amortissements = 0
            total_subventions = 0
            
            for project_id in cir_project_ids:
                # Vérifier que le projet est actif pour cette année
                cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
                projet_info = cursor.fetchone()
                if not projet_info or not projet_info[0] or not projet_info[1]:
                    continue
                    
                try:
                    debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
                    fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
                    
                    if year < debut_projet.year or year > fin_projet.year:
                        continue
                        
                    # Coûts directs (temps de travail) de l'année
                    cursor.execute(f"""
                        SELECT COALESCE(SUM(t.jours * c.montant_charge), 0)
                        FROM temps_travail t
                        JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                        WHERE t.annee = ? AND t.projet_id = ?
                    """, (year, project_id))
                    cout_temps = cursor.fetchone()[0] or 0
                    total_montant_charge += cout_temps
                    
                    # Amortissements de l'année
                    amort = self.calculate_amortissement_for_period(cursor, project_id, year)
                    total_amortissements += amort
                    
                    # Subventions de l'année (pour déduction)
                    subv = self.calculate_smart_distributed_subvention(cursor, project_id, year, None, projet_info)
                    total_subventions += subv
                    
                except Exception as e:
                    continue
            
            # Calculer le CIR
            montant_eligible_total = (total_montant_charge * k1) + (total_amortissements * k2)
            montant_net_eligible_total = montant_eligible_total - total_subventions
            
            if montant_net_eligible_total <= 0:
                return "CIR_NON_APPLICABLE"
                
            cir_total = montant_net_eligible_total * k3
            return max(0, cir_total)  # Le CIR ne peut pas être négatif
            
        except Exception as e:
            return 0

    def calculate_proportional_distributed_subvention(self, cursor, project_id, year, month, projet_info):
        """
        Calcule la subvention répartie proportionnellement aux dépenses éligibles de la période.
        
        NOUVELLE LOGIQUE DE RÉPARTITION :
        - La subvention totale est calculée sur l'ensemble du projet selon les paramètres configurés
        - Elle est ensuite répartie proportionnellement aux dépenses éligibles (selon les coefficients) 
          de la période demandée (mois ou année)
        - Cela permet une répartition mensuelle cohérente basée sur l'activité réelle de chaque période
        
        UTILISE LA NOUVELLE MÉTHODE DE SubventionDialog avec dates de début/fin de subvention
        """
        
        try:
            # Récupérer TOUS les paramètres de subvention pour ce projet (y compris les nouvelles dates)
            cursor.execute('''
                SELECT nom, mode_simplifie, montant_forfaitaire, depenses_temps_travail, coef_temps_travail, 
                       depenses_externes, coef_externes, depenses_autres_achats, coef_autres_achats, 
                       depenses_dotation_amortissements, coef_dotation_amortissements, cd, taux,
                       date_debut_subvention, date_fin_subvention, montant_subvention_max, depenses_eligibles_max
                FROM subventions WHERE projet_id = ?
            ''', (project_id,))
            
            subventions_config = cursor.fetchall()
            if not subventions_config:
                return 0
            
            subvention_total_periode = 0
            
            # Importer SubventionDialog pour utiliser sa méthode statique
            from subvention_dialog import SubventionDialog
            
            for subvention in subventions_config:
                (nom, mode_simplifie, montant_forfaitaire, dep_temps, coef_temps, dep_ext, coef_ext, 
                 dep_autres, coef_autres, dep_amort, coef_amort, cd, taux, date_debut_subv, date_fin_subv) = subvention
                
                # Construire le dictionnaire de données de subvention
                subvention_data = {
                    'nom': nom,
                    'mode_simplifie': mode_simplifie,
                    'montant_forfaitaire': montant_forfaitaire or 0,
                    'depenses_temps_travail': dep_temps or 0,
                    'coef_temps_travail': coef_temps or 1,
                    'depenses_externes': dep_ext or 0,
                    'coef_externes': coef_ext or 1,
                    'depenses_autres_achats': dep_autres or 0,
                    'coef_autres_achats': coef_autres or 1,
                    'depenses_dotation_amortissements': dep_amort or 0,
                    'coef_dotation_amortissements': coef_amort or 1,
                    'cd': cd or 1,
                    'taux': taux or 100,
                    'date_debut_subvention': date_debut_subv,
                    'date_fin_subvention': date_fin_subv
                }
                
                # Utiliser la nouvelle méthode de calcul distribué
                montant_subvention_periode = SubventionDialog.calculate_distributed_subvention(
                    project_id, subvention_data, year, month
                )
                
                subvention_total_periode += montant_subvention_periode
            
            return subvention_total_periode
            
        except Exception as e:
            traceback.print_exc()
            return 0
    
    def calculate_simple_distributed_subvention(self, cursor, project_id, year, month, projet_info):
        """Calcule la subvention répartie équitablement selon la règle simple :
        Subvention totale / Nb mois total du projet
        UTILISE LA MÊME LOGIQUE QUE subvention_dialog.py"""
        import datetime
        
        try:
            # Récupérer les paramètres de subvention pour ce projet
            cursor.execute('''
                SELECT mode_simplifie, montant_forfaitaire, depenses_temps_travail, coef_temps_travail, 
                       depenses_externes, coef_externes, depenses_autres_achats, coef_autres_achats, 
                       depenses_dotation_amortissements, coef_dotation_amortissements, cd, taux,
                       date_debut_subvention, date_fin_subvention, montant_subvention_max, depenses_eligibles_max
                FROM subventions WHERE projet_id = ?
            ''', (project_id,))
            
            subventions_config = cursor.fetchall()
            if not subventions_config:
                return 0
            
            # Calculer les dates du projet
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Vérifier si l'année demandée est dans la période du projet
            if year < debut_projet.year or year > fin_projet.year:
                return 0  # L'année n'est pas dans le projet
            
            # Calculer le montant total de subvention sur tout le projet
            montant_total_projet = 0
            
            for subvention in subventions_config:
                (mode_simplifie, montant_forfaitaire, dep_temps, coef_temps, dep_ext, coef_ext, dep_autres, coef_autres, 
                 dep_amort, coef_amort, cd, taux, date_debut_subv, date_fin_subv, montant_subvention_max, depenses_eligibles_max) = subvention
                
                if mode_simplifie:
                    # Mode simplifié : utiliser directement le montant forfaitaire
                    montant_subvention = montant_forfaitaire
                    # Appliquer le plafond du montant max si défini
                    if montant_subvention_max and montant_subvention_max > 0:
                        montant_subvention = min(montant_subvention, montant_subvention_max)
                    montant_total_projet += montant_subvention
                else:
                    # Mode détaillé : calculer selon les coefficients
                    montant_subvention_config = 0
                    
                    # 1. TEMPS DE TRAVAIL - MÊME LOGIQUE QUE subvention_dialog.py
                    if dep_temps and coef_temps:
                        cout_total_temps = self.calculate_temps_travail_total(cursor, project_id)
                        temps_travail = cout_total_temps * cd
                        montant_subvention_config += coef_temps * temps_travail
                    
                    # 2. DÉPENSES EXTERNES - MÊME LOGIQUE
                    if dep_ext and coef_ext:
                        cursor.execute('''
                            SELECT COALESCE(SUM(montant), 0) FROM depenses 
                            WHERE projet_id = ?
                        ''', (project_id,))
                        depenses_ext_total = cursor.fetchone()[0] or 0
                        montant_subvention_config += coef_ext * depenses_ext_total
                    
                    # 3. AUTRES ACHATS - MÊME LOGIQUE
                    if dep_autres and coef_autres:
                        cursor.execute('''
                            SELECT COALESCE(SUM(montant), 0) FROM autres_depenses 
                            WHERE projet_id = ?
                        ''', (project_id,))
                        autres_achats_total = cursor.fetchone()[0] or 0
                        montant_subvention_config += coef_autres * autres_achats_total
                    
                    # 4. AMORTISSEMENTS - MÊME LOGIQUE QUE subvention_dialog.py
                    if dep_amort and coef_amort:
                        amortissements_total = self.calculate_amortissements_total_subvention_style(
                            cursor, project_id, projet_info)
                        montant_subvention_config += coef_amort * amortissements_total
                    
                    # Calculer l'assiette éligible (avant application du taux)
                    assiette_eligible = montant_subvention_config
                    
                    # Appliquer le plafond du coût éligible max à l'assiette
                    assiette_plafonnee = assiette_eligible
                    if depenses_eligibles_max and depenses_eligibles_max > 0:
                        assiette_plafonnee = min(assiette_eligible, depenses_eligibles_max)
                    
                    # Appliquer le taux de subvention sur l'assiette plafonnée
                    montant_subvention_config = assiette_plafonnee * (taux / 100)
                    
                    # Appliquer le plafond du montant max si défini
                    if montant_subvention_max and montant_subvention_max > 0:
                        montant_subvention_config = min(montant_subvention_config, montant_subvention_max)
                    
                    montant_total_projet += montant_subvention_config
            
            # Calculer le nombre total de mois du projet
            nb_mois_total = (fin_projet.year - debut_projet.year) * 12 + (fin_projet.month - debut_projet.month) + 1
            
            if month:
                # Pour un mois : Subvention totale / Nombre total de mois du projet
                return montant_total_projet / nb_mois_total
            else:
                # Pour une année complète : Calculer combien de mois actifs dans cette année
                start_month = debut_projet.month if year == debut_projet.year else 1
                end_month = fin_projet.month if year == fin_projet.year else 12
                nb_mois_annee = end_month - start_month + 1
                
                # Subvention = (Subvention totale / Nb mois total) * Nb mois dans cette année
                return (montant_total_projet / nb_mois_total) * nb_mois_annee
            
        except Exception as e:
            return 0
    
    def calculate_temps_travail_total(self, cursor, project_id):
        """Calcule le temps de travail total exactement comme dans subvention_dialog.py"""
        try:
            cursor.execute("""
                SELECT tt.annee, tt.categorie, tt.mois, tt.jours 
                FROM temps_travail tt 
                WHERE tt.projet_id = ?
            """, (project_id,))
            
            temps_travail_rows = cursor.fetchall()
            cout_total_temps = 0
            
            for annee, categorie, mois, jours in temps_travail_rows:
                categorie_code = resolve_category_code(categorie)
                if not categorie_code:
                    continue
                    
                # Récupérer le montant chargé pour cette catégorie et cette année
                cursor.execute("""
                    SELECT montant_charge 
                    FROM categorie_cout 
                    WHERE categorie = ? AND annee = ?
                """, (categorie_code, annee))
                
                cout_row = cursor.fetchone()
                if cout_row and cout_row[0]:
                    montant_charge = float(cout_row[0])
                    cout_total_temps += jours * montant_charge
                else:
                    # Si pas de coût pour cette année/catégorie, utiliser une valeur par défaut
                    cout_total_temps += jours * 500  # 500€ par jour par défaut
            
            return cout_total_temps
            
        except Exception as e:
            return 0
    
    def calculate_amortissements_total_subvention_style(self, cursor, project_id, projet_info):
        """Calcule les amortissements totaux exactement comme dans subvention_dialog.py"""
        import datetime
        
        try:
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
            
            # Convertir les dates MM/yyyy en objets datetime - MÊME LOGIQUE
            try:
                debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
                fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            except ValueError:
                return 0
            
            cursor.execute("""
                SELECT montant, date_achat, duree 
                FROM investissements 
                WHERE projet_id = ?
            """, (project_id,))
            
            amortissements_total = 0
            
            for montant, date_achat, duree in cursor.fetchall():
                try:
                    # MÊME LOGIQUE QUE subvention_dialog.py lignes 177-213
                    # Convertir la date d'achat en datetime
                    achat_date = datetime.datetime.strptime(date_achat, '%m/%Y')
                    
                    # La dotation commence le mois suivant l'achat
                    debut_amort = achat_date.replace(day=1)
                    debut_amort = datetime.datetime(debut_amort.year, debut_amort.month, 1) + datetime.timedelta(days=32)
                    debut_amort = debut_amort.replace(day=1)
                    
                    # La fin de l'amortissement est soit la fin du projet, soit la fin de la période d'amortissement
                    fin_amort = achat_date.replace(day=1)
                    # Ajouter durée années à la date d'achat
                    fin_amort = datetime.datetime(fin_amort.year + int(duree), fin_amort.month, 1)
                    
                    # Prendre la date la plus proche entre fin du projet et fin d'amortissement
                    fin_effective = min(fin_projet, fin_amort)
                    
                    # Si le début d'amortissement est après la fin du projet, pas d'amortissement
                    if debut_amort > fin_projet:
                        continue
                        
                    # Calculer le nombre de mois d'amortissement effectif
                    mois_amort = (fin_effective.year - debut_amort.year) * 12 + fin_effective.month - debut_amort.month + 1
                    
                    # Calculer la dotation mensuelle (montant / durée en mois)
                    dotation_mensuelle = float(montant) / (int(duree) * 12)
                    
                    # Ajouter au total des amortissements
                    amortissements_total += dotation_mensuelle * mois_amort
                except Exception:
                    # En cas d'erreur dans le calcul, ignorer cet investissement
                    continue
            
            return amortissements_total
            
        except Exception as e:
            return 0
    
    
    def calculate_amortissement_for_period(self, cursor, project_id, year, month=None):
        """
        Calcule les amortissements pour une période donnée selon la règle :
        - Amortissement mensuel = montant total / (durée amortissement × 12)
        - Début d'amortissement = mois suivant le mois d'achat
        - CORRECTION: Respecte les dates de début et fin du projet
        """
        import datetime
        
        try:
            # Récupérer les dates du projet
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            projet_info = cursor.fetchone()
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
            
            # Convertir les dates du projet
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Récupérer tous les investissements du projet
            cursor.execute('''
                SELECT montant, date_achat, duree 
                FROM investissements 
                WHERE projet_id = ?
            ''', (project_id,))
            
            amortissements_total = 0
            
            for montant_inv, date_achat, duree in cursor.fetchall():
                try:
                    # Convertir la date d'achat (format 'MM/YYYY')
                    achat_date = datetime.datetime.strptime(date_achat, '%m/%Y')
                    
                    # La dotation commence le mois suivant l'achat (comme dans project_details_dialog.py)
                    debut_amort = achat_date.replace(day=1)
                    debut_amort = datetime.datetime(debut_amort.year, debut_amort.month, 1) + datetime.timedelta(days=32)
                    debut_amort = debut_amort.replace(day=1)
                    
                    # La fin de l'amortissement est soit la fin du projet, soit la fin de la période d'amortissement
                    fin_amort = achat_date.replace(day=1)
                    # Ajouter durée années à la date d'achat
                    fin_amort = datetime.datetime(fin_amort.year + int(duree), fin_amort.month, 1)
                    
                    # Prendre la date la plus proche entre fin du projet et fin d'amortissement
                    fin_effective = min(fin_projet, fin_amort)
                    
                    # Si le début d'amortissement est après la fin du projet, pas d'amortissement
                    if debut_amort > fin_projet:
                        continue
                    
                    # Calculer la dotation mensuelle (montant / durée en mois)
                    dotation_mensuelle = float(montant_inv) / (int(duree) * 12)
                    
                    if month:
                        # Calcul pour un mois spécifique
                        mois_demande = datetime.datetime(year, month, 1)
                        
                        # Vérifier si ce mois tombe dans la période d'amortissement ET dans le projet
                        if debut_amort <= mois_demande <= fin_effective:
                            amortissements_total += dotation_mensuelle
                    
                    else:
                        # Calcul pour une année complète
                        # Calculer le nombre de mois d'amortissement effectif dans cette année
                        debut_annee = datetime.datetime(year, 1, 1)
                        fin_annee = datetime.datetime(year, 12, 31)
                        
                        # Intersection entre [début_amort, fin_effective] et [début_annee, fin_annee]
                        debut_periode = max(debut_amort, debut_annee)
                        fin_periode = min(fin_effective, fin_annee)
                        
                        if debut_periode <= fin_periode:
                            # Calculer le nombre de mois dans cette intersection
                            mois_amort_annee = (fin_periode.year - debut_periode.year) * 12 + fin_periode.month - debut_periode.month + 1
                            amortissements_total += dotation_mensuelle * mois_amort_annee
                
                except (ValueError, TypeError) as e:
                    continue
            
            return amortissements_total
            
        except Exception as e:
            return 0
    
    def calculate_amortissement_for_year(self, cursor, project_id, year, month, projet_info):
        """Calcule les amortissements pour une année donnée"""
        import datetime
        
        try:
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
                
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Récupérer tous les investissements du projet
            cursor.execute('''
                SELECT montant, date_achat, duree 
                FROM investissements 
                WHERE projet_id = ?
            ''', (project_id,))
            
            amortissements_total = 0
            
            for montant_inv, date_achat, duree in cursor.fetchall():
                try:
                    # Convertir la date d'achat en datetime
                    achat_date = datetime.datetime.strptime(date_achat, '%m/%Y')
                    
                    # La dotation commence le mois suivant l'achat
                    debut_amort = achat_date.replace(day=1)
                    debut_amort = datetime.datetime(debut_amort.year, debut_amort.month, 1) + datetime.timedelta(days=32)
                    debut_amort = debut_amort.replace(day=1)
                    
                    # La fin de l'amortissement
                    fin_amort = datetime.datetime(achat_date.year + int(duree), achat_date.month, 1)
                    fin_effective = min(fin_projet, fin_amort)
                    
                    # Si le début d'amortissement est après la fin du projet, pas d'amortissement
                    if debut_amort > fin_projet:
                        continue
                    
                    # Pour l'année spécifique, calculer la part d'amortissement
                    year_start = datetime.datetime(year, 1, 1)
                    year_end = datetime.datetime(year, 12, 31)
                    
                    # Intersection avec l'année demandée
                    period_start = max(debut_amort, year_start)
                    period_end = min(fin_effective, year_end)
                    
                    if period_start <= period_end:
                        # Calculer le nombre de mois dans cette année
                        if month:
                            # Pour un mois spécifique
                            month_start = datetime.datetime(year, month, 1)
                            month_end = datetime.datetime(year, month, 28)  # Approximation
                            if period_start <= month_end and period_end >= month_start:
                                mois_amort = 1
                            else:
                                mois_amort = 0
                        else:
                            # Pour toute l'année
                            mois_amort = (period_end.year - period_start.year) * 12 + period_end.month - period_start.month + 1
                        
                        # Calculer la dotation mensuelle
                        dotation_mensuelle = float(montant_inv) / (int(duree) * 12)
                        amortissements_total += dotation_mensuelle * mois_amort
                
                except Exception:
                    continue
            
            return amortissements_total
            
        except Exception as e:
            return 0
    
    def calculate_redistributed_expenses(self, cursor, project_id, year, month, table_name):
        """
        Calcule les dépenses avec la logique de redistribution automatique.
        Si une seule dépense existe dans l'année, elle est redistribuée sur tous les mois actifs.
        
        Args:
            cursor: Curseur de base de données
            project_id: ID du projet
            year: Année cible
            month: Mois cible (None pour toute l'année)
            table_name: 'depenses' ou 'autres_depenses'
        """
        try:
            import datetime
            
            # Récupérer les dates du projet
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            projet_info = cursor.fetchone()
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
            
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Vérifier si l'année demandée est dans la période du projet
            if year < debut_projet.year or year > fin_projet.year:
                return 0
            
            # Noms des mois français
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            # Récupérer toutes les dépenses de l'année pour ce projet
            cursor.execute(f"""
                SELECT mois, SUM(montant) as total_montant
                FROM {table_name}
                WHERE projet_id = ? AND annee = ?
                GROUP BY mois
            """, (project_id, year))
            
            depenses_par_mois = cursor.fetchall()
            
            # Si pas de dépenses, retourner 0
            if not depenses_par_mois:
                return 0
            
            # Si une seule entrée dans l'année, appliquer la redistribution automatique
            if len(depenses_par_mois) == 1:
                mois_unique, montant_total = depenses_par_mois[0]
                
                # Calculer les mois actifs du projet dans cette année
                mois_actifs = []
                for m in range(1, 13):
                    mois_date = datetime.datetime(year, m, 1)
                    if debut_projet <= mois_date <= fin_projet:
                        mois_actifs.append(m)
                
                if not mois_actifs:
                    return 0
                
                # Répartir le montant sur tous les mois actifs
                montant_par_mois = montant_total / len(mois_actifs)
                
                if month:
                    # Cas mensuel : retourner la part du mois demandé
                    if month in mois_actifs:
                        return montant_par_mois
                    else:
                        return 0
                else:
                    # Cas annuel : retourner le total redistribué
                    return montant_total
            
            else:
                # Cas normal : plusieurs dépenses dans l'année
                if month:
                    # Cas mensuel : chercher la dépense du mois exact
                    month_name = month_names[month - 1]
                    for mois_nom, montant in depenses_par_mois:
                        if mois_nom == month_name:
                            return montant
                    return 0
                else:
                    # Cas annuel : sommer toutes les dépenses
                    return sum(montant for _, montant in depenses_par_mois)
                    
        except Exception as e:
            return 0
    
    def calculate_redistributed_temps_travail(self, cursor, project_id, year, month, cost_type):
        """
        Calcule le temps de travail avec la logique de redistribution automatique.
        
        RÈGLE DE REDISTRIBUTION :
        - Vérifier que TOUS les couples (membre_id, categorie) du projet/année n'ont qu'une seule entrée de mois
        - Si OUI : redistribuer chaque couple individuellement sur tous les mois actifs du projet
        - Si NON : utiliser les données réelles (pas de redistribution)
        
        Args:
            cursor: Curseur de base de données
            project_id: ID du projet
            year: Année cible
            month: Mois cible (None pour toute l'année)
            cost_type: Type de coût ('montant_charge', 'cout_production', 'cout_complet')
        
        Returns:
            float: Coût du temps de travail pour la période demandée
        """
        try:
            # Récupérer les dates du projet pour calculer les mois actifs
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            projet_info = cursor.fetchone()
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
                
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Vérifier si l'année est dans la période du projet
            if year < debut_projet.year or year > fin_projet.year:
                return 0
            
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            # 1. VÉRIFICATION : Est-ce que tous les couples (membre_id, categorie) n'ont qu'une seule entrée dans l'année ?
            cursor.execute("""
                SELECT membre_id, categorie, COUNT(*) as nb_entries
                FROM temps_travail 
                WHERE projet_id = ? AND annee = ?
                GROUP BY membre_id, categorie
            """, (project_id, year))
            
            couples_entries = cursor.fetchall()
            
            # Si aucune donnée, retourner 0
            if not couples_entries:
                return 0
            
            # Vérifier si TOUS les couples n'ont qu'une seule entrée
            all_single_entry = all(nb_entries == 1 for _, _, nb_entries in couples_entries)
            
            if not all_single_entry:
                # Au moins un couple a plusieurs entrées → PAS DE REDISTRIBUTION
                # Utiliser les données réelles
                return self._calculate_real_temps_travail(cursor, project_id, year, month, cost_type)
            
            # 2. REDISTRIBUTION : Tous les couples n'ont qu'une seule entrée
            # Calculer les mois actifs du projet dans cette année
            mois_actifs = []
            for m in range(1, 13):
                mois_date = datetime.datetime(year, m, 1)
                if debut_projet <= mois_date <= fin_projet:
                    mois_actifs.append(m)
            
            if not mois_actifs:
                return 0
            
            # Récupérer toutes les données de temps de travail de l'année
            cursor.execute("""
                SELECT membre_id, categorie, mois, jours
                FROM temps_travail 
                WHERE projet_id = ? AND annee = ?
            """, (project_id, year))
            
            temps_travail_data = cursor.fetchall()
            
            # Calculer le coût total redistributed
            cout_total = 0
            
            for membre_id, categorie, mois_original, jours_total in temps_travail_data:
                # Récupérer le coût unitaire pour cette catégorie/année
                cursor.execute(f"""
                    SELECT {cost_type} 
                    FROM categorie_cout 
                    WHERE libelle = ? AND annee = ?
                """, (categorie, year))
                
                cout_unitaire = cursor.fetchone()
                if not cout_unitaire or not cout_unitaire[0]:
                    continue
                    
                cout_unitaire = cout_unitaire[0]
                
                # Redistribuer les jours sur tous les mois actifs
                jours_par_mois = jours_total / len(mois_actifs)
                
                if month is not None:
                    # Calcul pour un mois spécifique
                    if month in mois_actifs:
                        cout_total += jours_par_mois * cout_unitaire
                else:
                    # Calcul pour toute l'année
                    cout_total += jours_total * cout_unitaire
            
            return cout_total
                    
        except Exception as e:
            import traceback
            traceback.print_exc()
            return 0
    
    def _calculate_real_temps_travail(self, cursor, project_id, year, month, cost_type):
        """
        Calcule le temps de travail réel sans redistribution
        """
        try:
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            if month is not None:
                # Calcul pour un mois spécifique
                cursor.execute(f"""
                    SELECT COALESCE(SUM(t.jours * c.{cost_type}), 0)
                    FROM temps_travail t
                    JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                    WHERE t.projet_id = ? AND t.annee = ? AND t.mois = ?
                """, (project_id, year, month_names[month-1]))
            else:
                # Calcul pour toute l'année
                cursor.execute(f"""
                    SELECT COALESCE(SUM(t.jours * c.{cost_type}), 0)
                    FROM temps_travail t
                    JOIN categorie_cout c ON t.categorie = c.libelle AND t.annee = c.annee
                    WHERE t.projet_id = ? AND t.annee = ?
                """, (project_id, year))
            
            result = cursor.fetchone()
            return result[0] if result else 0
            
        except Exception as e:
            return 0
    
    def calculate_redistributed_temps_travail_jours(self, cursor, project_id, year, month):
        """
        Calcule le nombre de jours de temps de travail avec la même logique de redistribution
        """
        try:
            # Récupérer les dates du projet pour calculer les mois actifs
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            projet_info = cursor.fetchone()
            if not projet_info or not projet_info[0] or not projet_info[1]:
                return 0
                
            debut_projet = datetime.datetime.strptime(projet_info[0], '%m/%Y')
            fin_projet = datetime.datetime.strptime(projet_info[1], '%m/%Y')
            
            # Vérifier si l'année est dans la période du projet
            if year < debut_projet.year or year > fin_projet.year:
                return 0
            
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            # 1. VÉRIFICATION : Est-ce que tous les couples (membre_id, categorie) n'ont qu'une seule entrée dans l'année ?
            cursor.execute("""
                SELECT membre_id, categorie, COUNT(*) as nb_entries
                FROM temps_travail 
                WHERE projet_id = ? AND annee = ?
                GROUP BY membre_id, categorie
            """, (project_id, year))
            
            couples_entries = cursor.fetchall()
            
            # Si aucune donnée, retourner 0
            if not couples_entries:
                return 0
            
            # Vérifier si TOUS les couples n'ont qu'une seule entrée
            all_single_entry = all(nb_entries == 1 for _, _, nb_entries in couples_entries)
            
            if not all_single_entry:
                # Au moins un couple a plusieurs entrées → PAS DE REDISTRIBUTION
                # Utiliser les données réelles
                return self._calculate_real_temps_travail_jours(cursor, project_id, year, month)
            
            # 2. REDISTRIBUTION : Tous les couples n'ont qu'une seule entrée
            # Calculer les mois actifs du projet dans cette année
            mois_actifs = []
            for m in range(1, 13):
                mois_date = datetime.datetime(year, m, 1)
                if debut_projet <= mois_date <= fin_projet:
                    mois_actifs.append(m)
            
            if not mois_actifs:
                return 0
            
            # Récupérer toutes les données de temps de travail de l'année
            cursor.execute("""
                SELECT membre_id, categorie, mois, jours
                FROM temps_travail 
                WHERE projet_id = ? AND annee = ?
            """, (project_id, year))
            
            temps_travail_data = cursor.fetchall()
            
            # Calculer le nombre total de jours redistributed
            jours_total = 0
            
            for membre_id, categorie, mois_original, jours_original in temps_travail_data:
                # Redistribuer les jours sur tous les mois actifs
                jours_par_mois = jours_original / len(mois_actifs)
                
                if month is not None:
                    # Calcul pour un mois spécifique
                    if month in mois_actifs:
                        jours_total += jours_par_mois
                else:
                    # Calcul pour toute l'année
                    jours_total += jours_original
            
            return jours_total
                    
        except Exception as e:
            import traceback
            traceback.print_exc()
            return 0
    
    def _calculate_real_temps_travail_jours(self, cursor, project_id, year, month):
        """
        Calcule le nombre de jours de temps de travail réel sans redistribution
        """
        try:
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            
            if month is not None:
                # Calcul pour un mois spécifique
                cursor.execute("""
                    SELECT COALESCE(SUM(jours), 0)
                    FROM temps_travail
                    WHERE projet_id = ? AND annee = ? AND mois = ?
                """, (project_id, year, month_names[month-1]))
            else:
                # Calcul pour toute l'année
                cursor.execute("""
                    SELECT COALESCE(SUM(jours), 0)
                    FROM temps_travail
                    WHERE projet_id = ? AND annee = ?
                """, (project_id, year))
            
            result = cursor.fetchone()
            return result[0] if result else 0
            
        except Exception as e:
            return 0

    def calculate_redistributed_recettes(self, cursor, project_ids, year, month):
        """
        Calcule les recettes avec la logique de redistribution automatique.
        Si une seule recette existe dans l'année, elle est redistribuée sur tous les mois actifs.
        
        Args:
            cursor: Curseur de base de données
            project_ids: Liste des IDs des projets
            year: Année cible
            month: Mois cible (None pour toute l'année)
        
        Returns:
            float: Montant des recettes pour la période demandée
        """
        try:
            total_recettes = 0
            
            for project_id in project_ids:
                # Vérifier les mois actifs du projet pour cette année
                active_months = self.get_active_months_for_year(year)
                if not active_months:
                    continue
                
                # 1. Récupérer toutes les recettes du projet pour cette année
                cursor.execute("""
                    SELECT DISTINCT mois FROM recettes 
                    WHERE projet_id = ? AND annee = ?
                """, (project_id, year))
                
                mois_with_data = [row[0] for row in cursor.fetchall()]
                
                # 2. Déterminer si redistribution nécessaire
                # Si une seule entrée de mois ET que ce mois est différent de tous les mois actifs
                # OU si une seule entrée et on veut une granularité mensuelle
                should_redistribute = (
                    len(mois_with_data) == 1 and 
                    len(active_months) > 1
                )
                
                if should_redistribute:
                    # REDISTRIBUTION : récupérer le montant total et le répartir
                    cursor.execute("""
                        SELECT COALESCE(SUM(montant), 0) FROM recettes 
                        WHERE projet_id = ? AND annee = ?
                    """, (project_id, year))
                    
                    montant_total = cursor.fetchone()[0] or 0
                    
                    if montant_total > 0:
                        if month is not None:
                            # Mode mensuel : répartir le montant total sur tous les mois actifs
                            if month in active_months:
                                recettes_project = montant_total / len(active_months)
                            else:
                                recettes_project = 0
                        else:
                            # Mode annuel : retourner le montant total
                            recettes_project = montant_total
                    else:
                        recettes_project = 0
                        
                else:
                    # PAS DE REDISTRIBUTION : utiliser les données réelles
                    if month is not None:
                        # Mode mensuel
                        cursor.execute("""
                            SELECT COALESCE(SUM(montant), 0) FROM recettes 
                            WHERE projet_id = ? AND annee = ? AND mois = ?
                        """, (project_id, year, month))
                    else:
                        # Mode annuel
                        cursor.execute("""
                            SELECT COALESCE(SUM(montant), 0) FROM recettes 
                            WHERE projet_id = ? AND annee = ?
                        """, (project_id, year))
                    
                    recettes_project = cursor.fetchone()[0] or 0
                
                total_recettes += recettes_project
            
            return total_recettes
            
        except Exception as e:
            return 0

def show_compte_resultat(parent, config_data):
    """Fonction pour afficher le compte de résultat"""
    dialog = CompteResultatDisplay(parent, config_data)
    return dialog.exec()


class ExportSettingsDialog(QDialog):
    """Dialogue pour configurer les paramètres d'export PDF et Excel"""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("Paramètres d'Export")
        self.setMinimumSize(500, 600)
        
        # Charger les paramètres existants
        self.settings = self.load_export_settings()
        
        self.init_ui()
        self.load_current_settings()
        
    def init_ui(self):
        """Initialise l'interface utilisateur"""
        layout = QVBoxLayout()
        
        # Titre
        title = QLabel("Configuration des Exports PDF et Excel")
        title.setStyleSheet("font-size: 14pt; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(title)
        
        # Section couleurs
        colors_group = self.create_colors_section()
        layout.addWidget(colors_group)
        
        # Section logo
        logo_group = self.create_logo_section()
        layout.addWidget(logo_group)
        
        # Aperçu des couleurs
        preview_group = self.create_preview_section()
        layout.addWidget(preview_group)
        
        # Boutons
        buttons_layout = self.create_buttons()
        layout.addLayout(buttons_layout)
        
        self.setLayout(layout)
    
    def create_colors_section(self):
        """Crée la section de configuration des couleurs"""
        group = QGroupBox("Configuration des Couleurs")
        layout = QGridLayout()
        
        # Couleur du titre du document
        layout.addWidget(QLabel("Couleur du titre du document:"), 0, 0)
        self.title_color_btn = QPushButton()
        self.title_color_btn.setFixedSize(80, 30)
        self.title_color_btn.clicked.connect(lambda: self.select_color('title'))
        layout.addWidget(self.title_color_btn, 0, 1)
        
        # Couleur des en-têtes
        layout.addWidget(QLabel("Couleur des en-têtes:"), 1, 0)
        self.header_color_btn = QPushButton()
        self.header_color_btn.setFixedSize(80, 30)
        self.header_color_btn.clicked.connect(lambda: self.select_color('header'))
        layout.addWidget(self.header_color_btn, 1, 1)
        
        # Couleur des totaux
        layout.addWidget(QLabel("Couleur des totaux:"), 2, 0)
        self.total_color_btn = QPushButton()
        self.total_color_btn.setFixedSize(80, 30)
        self.total_color_btn.clicked.connect(lambda: self.select_color('total'))
        layout.addWidget(self.total_color_btn, 2, 1)
        
        # Couleur du résultat financier
        layout.addWidget(QLabel("Couleur du résultat financier:"), 3, 0)
        self.result_color_btn = QPushButton()
        self.result_color_btn.setFixedSize(80, 30)
        self.result_color_btn.clicked.connect(lambda: self.select_color('result'))
        layout.addWidget(self.result_color_btn, 3, 1)
        
        group.setLayout(layout)
        return group
    
    def create_logo_section(self):
        """Crée la section de configuration du logo"""
        group = QGroupBox("Configuration du Logo")
        layout = QGridLayout()
        
        # Chemin du logo
        layout.addWidget(QLabel("Fichier logo:"), 0, 0)
        self.logo_path = QLineEdit()
        self.logo_path.setPlaceholderText("Sélectionnez un fichier image...")
        layout.addWidget(self.logo_path, 0, 1)
        
        browse_btn = QPushButton("Parcourir")
        browse_btn.clicked.connect(self.browse_logo)
        layout.addWidget(browse_btn, 0, 2)
        
        # Position du logo
        layout.addWidget(QLabel("Position du logo:"), 1, 0)
        self.logo_position = QComboBox()
        self.logo_position.addItems(["Haut gauche", "Haut droite", "Haut centre"])
        layout.addWidget(self.logo_position, 1, 1)
        
        group.setLayout(layout)
        return group
    
    def create_preview_section(self):
        """Crée la section d'aperçu des couleurs"""
        group = QGroupBox("Aperçu des Couleurs")
        layout = QVBoxLayout()
        
        # Tableau d'aperçu
        self.preview_table = QTableWidget(6, 2)
        self.preview_table.setHorizontalHeaderLabels(["Élément", "Valeur"])
        self.preview_table.setMaximumHeight(200)
        
        # Données d'exemple
        preview_data = [
            ("COMPTE DE RÉSULTAT", "title"),
            ("Poste", "header"),
            ("Recettes", "100 000,00"),
            ("Subventions", "50 000,00"),
            ("TOTAL PRODUITS", "total"),
            ("RÉSULTAT FINANCIER", "result")
        ]
        
        for row, (label, data_type) in enumerate(preview_data):
            # Colonne label
            item_label = QTableWidgetItem(label)
            if data_type == "title":
                item_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
            elif data_type == "header":
                item_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
            elif data_type in ["total", "result"]:
                item_label.setFont(QFont("Arial", 10, QFont.Weight.Bold))
            
            self.preview_table.setItem(row, 0, item_label)
            
            # Colonne valeur
            if data_type in ["title", "header", "total", "result"]:
                item_value = QTableWidgetItem("")
            else:
                item_value = QTableWidgetItem(data_type)
                item_value.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            
            self.preview_table.setItem(row, 1, item_value)
        
        self.preview_table.resizeColumnsToContents()
        layout.addWidget(self.preview_table)
        
        group.setLayout(layout)
        return group
    
    def create_buttons(self):
        """Crée les boutons de validation"""
        layout = QHBoxLayout()
        
        # Bouton Réinitialiser
        reset_btn = QPushButton("Réinitialiser")
        reset_btn.clicked.connect(self.reset_to_defaults)
        layout.addWidget(reset_btn)
        
        layout.addStretch()
        
        # Bouton Annuler
        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(self.reject)
        layout.addWidget(cancel_btn)
        
        # Bouton Sauvegarder
        save_btn = QPushButton("Sauvegarder")
        save_btn.clicked.connect(self.save_settings)
        save_btn.setStyleSheet("QPushButton { background-color: #27ae60; color: white; font-weight: bold; padding: 8px; }")
        layout.addWidget(save_btn)
        
        return layout
    
    def select_color(self, color_type):
        """Ouvre le sélecteur de couleur"""
        current_color = getattr(self.settings, f'{color_type}_color', '#000000')
        color = QColorDialog.getColor(QColor(current_color), self, f"Choisir la couleur - {color_type}")
        
        if color.isValid():
            setattr(self.settings, f'{color_type}_color', color.name())
            self.update_color_button(color_type, color.name())
            self.update_preview()
    
    def update_color_button(self, color_type, color_hex):
        """Met à jour l'apparence du bouton de couleur"""
        button = getattr(self, f'{color_type}_color_btn')
        button.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999999;")
    
    def browse_logo(self):
        """Ouvre le dialogue de sélection de fichier pour le logo"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Sélectionner un logo",
            "",
            "Images (*.png *.jpg *.jpeg *.bmp *.gif)"
        )
        
        if file_path:
            self.logo_path.setText(file_path)
            self.settings.logo_path = file_path
    
    def update_preview(self):
        """Met à jour l'aperçu des couleurs"""
        color_mapping = {
            0: ('title_color', '#2c3e50'),  # Titre
            1: ('header_color', '#34495e'),  # En-tête
            4: ('total_color', '#3498db'),   # Total
            5: ('result_color', '#2ecc71')   # Résultat
        }
        
        for row, (color_attr, default_color) in color_mapping.items():
            color_hex = getattr(self.settings, color_attr, default_color)
            
            # Mettre à jour la couleur de fond des cellules
            for col in range(2):
                item = self.preview_table.item(row, col)
                if item:
                    item.setBackground(QColor(color_hex))
                    item.setForeground(QColor('#ffffff'))
    
    def load_current_settings(self):
        """Charge les paramètres actuels dans l'interface"""
        # Mettre à jour les boutons de couleur
        colors = ['title', 'header', 'total', 'result']
        defaults = ['#2c3e50', '#34495e', '#3498db', '#2ecc71']
        
        for color_type, default in zip(colors, defaults):
            color_hex = getattr(self.settings, f'{color_type}_color', default)
            self.update_color_button(color_type, color_hex)
        
        # Mettre à jour le logo
        if hasattr(self.settings, 'logo_path') and self.settings.logo_path:
            self.logo_path.setText(self.settings.logo_path)
        
        if hasattr(self.settings, 'logo_position') and self.settings.logo_position:
            index = self.logo_position.findText(self.settings.logo_position)
            if index >= 0:
                self.logo_position.setCurrentIndex(index)
        
        # Mettre à jour l'aperçu
        self.update_preview()
    
    def reset_to_defaults(self):
        """Remet les paramètres par défaut"""
        self.settings = self.get_default_settings()
        self.load_current_settings()
    
    def save_settings(self):
        """Sauvegarde les paramètres dans la base de données"""
        try:
            # Récupérer les valeurs actuelles
            self.settings.logo_path = self.logo_path.text()
            self.settings.logo_position = self.logo_position.currentText()
            
            # Sauvegarder dans la base de données
            self.save_export_settings(self.settings)
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la sauvegarde: {str(e)}")
    
    def load_export_settings(self):
        """Charge les paramètres d'export depuis la base de données"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Créer la table si elle n'existe pas
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS export_settings (
                    id INTEGER PRIMARY KEY,
                    title_color TEXT DEFAULT '#2c3e50',
                    header_color TEXT DEFAULT '#34495e',
                    total_color TEXT DEFAULT '#3498db',
                    result_color TEXT DEFAULT '#2ecc71',
                    logo_path TEXT DEFAULT '',
                    logo_position TEXT DEFAULT 'Haut gauche'
                )
            ''')
            
            # Récupérer les paramètres
            cursor.execute('SELECT * FROM export_settings WHERE id = 1')
            result = cursor.fetchone()
            
            if result:
                settings = type('Settings', (), {})()
                settings.title_color = result[1]
                settings.header_color = result[2]
                settings.total_color = result[3]
                settings.result_color = result[4]
                settings.logo_path = result[5]
                settings.logo_position = result[6]
            else:
                settings = self.get_default_settings()
            
            conn.close()
            return settings
            
        except Exception as e:
            QMessageBox.warning(self, "Avertissement", f"Impossible de charger les paramètres: {str(e)}")
            return self.get_default_settings()
    
    def get_default_settings(self):
        """Retourne les paramètres par défaut"""
        settings = type('Settings', (), {})()
        settings.title_color = '#2c3e50'
        settings.header_color = '#34495e'
        settings.total_color = '#3498db'
        settings.result_color = '#2ecc71'
        settings.logo_path = ''
        settings.logo_position = 'Haut gauche'
        return settings
    
    def save_export_settings(self, settings):
        """Sauvegarde les paramètres dans la base de données"""
        conn = get_connection()
        cursor = conn.cursor()
        
        try:
            # Créer la table si elle n'existe pas
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS export_settings (
                    id INTEGER PRIMARY KEY,
                    title_color TEXT DEFAULT '#2c3e50',
                    header_color TEXT DEFAULT '#34495e',
                    total_color TEXT DEFAULT '#3498db',
                    result_color TEXT DEFAULT '#2ecc71',
                    logo_path TEXT DEFAULT '',
                    logo_position TEXT DEFAULT 'Haut gauche'
                )
            ''')
            
            # Insérer ou mettre à jour
            cursor.execute('''
                INSERT OR REPLACE INTO export_settings 
                (id, title_color, header_color, total_color, result_color, logo_path, logo_position)
                VALUES (1, ?, ?, ?, ?, ?, ?)
            ''', (settings.title_color, settings.header_color, settings.total_color,
                  settings.result_color, settings.logo_path, settings.logo_position))
            
            conn.commit()
            
        finally:
            conn.close()
