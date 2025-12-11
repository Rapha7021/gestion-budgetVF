"""
Générateur de modèle Excel pour l'import de données de projet
Ce script crée un fichier Excel avec 4 feuilles pour saisir manuellement :
- Temps de travail
- Dépenses externes
- Autres dépenses
- Recettes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from database import get_connection


def creer_modele_excel(nom_fichier="Modele_Import_Projet.xlsx", projet_id=None):
    """
    Crée un fichier Excel modèle avec les feuilles et colonnes nécessaires
    Si projet_id est fourni, ajoute des listes déroulantes basées sur les données du projet
    """
    # Récupérer les données du projet si un ID est fourni
    projet_data = None
    if projet_id:
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # Récupérer les infos du projet
            cursor.execute("""
                SELECT code, nom, date_debut, date_fin 
                FROM projets WHERE id = ?
            """, (projet_id,))
            projet_info = cursor.fetchone()
            
            if not projet_info:
                conn.close()
                return
            
            # Récupérer l'équipe du projet depuis la table equipe
            cursor.execute("""
                SELECT direction, type, nombre
                FROM equipe 
                WHERE projet_id = ?
                ORDER BY direction, type
            """, (projet_id,))
            equipe_info = cursor.fetchall()
            
            # Générer les membres depuis la table equipe
            directions = set()
            categories = set()
            membres = []
            
            for direction, type_, nombre in equipe_info:
                if direction:
                    directions.add(direction)
                if type_:
                    categories.add(type_)
                
                # Générer les IDs de membres
                if direction and type_ and nombre:
                    for i in range(int(nombre)):
                        if int(nombre) > 1:
                            membre_id = f"{direction}_{type_}_{i+1}"
                        else:
                            membre_id = f"{direction}_{type_}"
                        membres.append(membre_id)
            
            # Récupérer aussi depuis temps_travail (pour les membres déjà existants)
            cursor.execute("""
                SELECT DISTINCT direction, categorie, membre_id 
                FROM temps_travail 
                WHERE projet_id = ?
                ORDER BY direction, categorie, membre_id
            """, (projet_id,))
            temps_travail_data = cursor.fetchall()
            
            for direction, categorie, membre_id in temps_travail_data:
                if direction:
                    directions.add(direction)
                if categorie:
                    categories.add(categorie)
                if membre_id and membre_id not in membres:
                    membres.append(membre_id)
            
            conn.close()
            
            projet_data = {
                'code': projet_info[0],
                'nom': projet_info[1],
                'date_debut': projet_info[2],
                'date_fin': projet_info[3],
                'directions': sorted(list(directions)),
                'categories': sorted(list(categories)),
                'membres': sorted(membres)
            }
        except Exception as e:
            print(f"Avertissement: Impossible de récupérer les données du projet: {e}")
            projet_data = None
    
    wb = Workbook()
    
    # Supprime la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Styles pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style pour les exemples
    example_font = Font(italic=True, color="666666")
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ==================== FEUILLE 1: TEMPS DE TRAVAIL ====================
    ws_temps = wb.create_sheet("Temps de Travail")
    
    # En-têtes
    headers_temps = [
        "Année",
        "Direction", 
        "Catégorie",
        "Membre ID",
        "Mois",
        "Jours"
    ]
    
    descriptions_temps = [
        "Année du projet (ex: 2024)",
        "Direction (ex: R&D, Production, etc.)",
        "Catégorie (ex: Ingénieur, Technicien, etc.)",
        "Identifiant du membre (ex: EMP001)",
        "Mois en français (ex: Janvier)",
        "Nombre de jours travaillés (ex: 15,5)"
    ]
    
    for col, (header, desc) in enumerate(zip(headers_temps, descriptions_temps), 1):
        cell = ws_temps.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Ajouter la description en commentaire
        comment_cell = ws_temps.cell(row=2, column=col, value=desc)
        comment_cell.font = example_font
        comment_cell.fill = example_fill
        comment_cell.border = thin_border
    
    # Exemple de données
    exemple_temps = [
        [2024, "R&D", "Ingénieur", "EMP001", "Janvier", 15.5],
        [2024, "R&D", "Technicien", "EMP002", "Janvier", 20],
        [2024, "Production", "Chef de projet", "EMP003", "Février", 10],
    ]
    
    for row_idx, row_data in enumerate(exemple_temps, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_temps.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border
    
    # Ajuster les largeurs de colonnes
    ws_temps.column_dimensions['A'].width = 12
    ws_temps.column_dimensions['B'].width = 20
    ws_temps.column_dimensions['C'].width = 20
    ws_temps.column_dimensions['D'].width = 15
    ws_temps.column_dimensions['E'].width = 15
    ws_temps.column_dimensions['F'].width = 12
    
    # Variable pour stocker la liste des mois (utilisée dans toutes les feuilles)
    mois_liste = []
    annees_liste = []
    
    # Ajouter des validations de données si projet_data est disponible
    if projet_data:
        # Liste déroulante pour Direction (colonne B)
        if projet_data['directions']:
            directions_str = ','.join(projet_data['directions'])
            dv_direction = DataValidation(type="list", formula1=f'"{directions_str}"', allow_blank=True)
            dv_direction.error = 'Direction invalide'
            dv_direction.errorTitle = 'Entrée invalide'
            dv_direction.prompt = 'Sélectionnez une direction'
            dv_direction.promptTitle = 'Direction'
            ws_temps.add_data_validation(dv_direction)
            dv_direction.add('B3:B1000')
        
        # Liste déroulante pour Catégorie (colonne C)
        if projet_data['categories']:
            categories_str = ','.join(projet_data['categories'])
            dv_categorie = DataValidation(type="list", formula1=f'"{categories_str}"', allow_blank=True)
            dv_categorie.error = 'Catégorie invalide'
            dv_categorie.errorTitle = 'Entrée invalide'
            dv_categorie.prompt = 'Sélectionnez une catégorie'
            dv_categorie.promptTitle = 'Catégorie'
            ws_temps.add_data_validation(dv_categorie)
            dv_categorie.add('C3:C1000')
        
        # Liste déroulante pour Membre ID (colonne D)
        if projet_data['membres']:
            membres_str = ','.join(projet_data['membres'])
            dv_membre = DataValidation(type="list", formula1=f'"{membres_str}"', allow_blank=True)
            dv_membre.error = 'Membre invalide'
            dv_membre.errorTitle = 'Entrée invalide'
            dv_membre.prompt = 'Sélectionnez un membre de l\'équipe'
            dv_membre.promptTitle = 'Membre'
            ws_temps.add_data_validation(dv_membre)
            dv_membre.add('D3:D1000')
        
        # Générer la liste des mois entre date_debut et date_fin
        if projet_data['date_debut'] and projet_data['date_fin']:
            try:
                # Parser les dates (format MM/YYYY ou YYYY-MM-DD)
                date_debut_str = projet_data['date_debut']
                date_fin_str = projet_data['date_fin']
                
                # Essayer différents formats
                debut = None
                fin = None
                for fmt in ['%m/%Y', '%Y-%m-%d', '%d/%m/%Y']:
                    try:
                        debut = datetime.strptime(date_debut_str, fmt)
                        fin = datetime.strptime(date_fin_str, fmt)
                        break
                    except:
                        continue
                
                # Générer tous les mois entre les deux dates si les dates sont valides
                if debut and fin:
                    # Générer la liste des années
                    annees_liste = list(range(debut.year, fin.year + 1))
                    
                    # Liste des mois en français
                    mois_fr = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                              "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                    
                    # Créer la liste des mois couverts par le projet
                    mois_liste = []
                    current = debut
                    while current <= fin:
                        mois_liste.append(mois_fr[current.month - 1])
                        # Passer au mois suivant
                        if current.month == 12:
                            current = current.replace(year=current.year + 1, month=1)
                        else:
                            current = current.replace(month=current.month + 1)
                    
                    # Supprimer les doublons en préservant l'ordre
                    mois_liste_unique = []
                    for mois in mois_liste:
                        if mois not in mois_liste_unique:
                            mois_liste_unique.append(mois)
                    mois_liste = mois_liste_unique
                    
                    if mois_liste:
                        mois_str = ','.join(mois_liste)
                        dv_mois = DataValidation(type="list", formula1=f'"{mois_str}"', allow_blank=True)
                        dv_mois.error = 'Mois invalide ou hors période du projet'
                        dv_mois.errorTitle = 'Entrée invalide'
                        dv_mois.prompt = 'Sélectionnez un mois de la période du projet'
                        dv_mois.promptTitle = 'Mois'
                        ws_temps.add_data_validation(dv_mois)
                        dv_mois.add('E3:E1000')
                    
                    if annees_liste:
                        annees_str = ','.join(map(str, annees_liste))
                        dv_annee = DataValidation(type="list", formula1=f'"{annees_str}"', allow_blank=True)
                        dv_annee.error = 'Année invalide ou hors période du projet'
                        dv_annee.errorTitle = 'Entrée invalide'
                        dv_annee.prompt = 'Sélectionnez une année de la période du projet'
                        dv_annee.promptTitle = 'Année'
                        ws_temps.add_data_validation(dv_annee)
                        dv_annee.add('A3:A1000')
            except Exception as e:
                print(f"Impossible de générer la liste des mois: {e}")
    
    # ==================== FEUILLE 2: DÉPENSES EXTERNES ====================
    ws_depenses = wb.create_sheet("Dépenses Externes")
    
    headers_depenses = [
        "Année",
        "Mois",
        "Libellé",
        "Montant"
    ]
    
    descriptions_depenses = [
        "Année (ex: 2024)",
        "Mois en français (ex: Janvier)",
        "Description de la dépense",
        "Montant en euros (ex: 1500,50)"
    ]
    
    for col, (header, desc) in enumerate(zip(headers_depenses, descriptions_depenses), 1):
        cell = ws_depenses.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        comment_cell = ws_depenses.cell(row=2, column=col, value=desc)
        comment_cell.font = example_font
        comment_cell.fill = example_fill
        comment_cell.border = thin_border
    
    # Exemples
    exemple_depenses = [
        [2024, "Janvier", "Achat matériel informatique", 2500.00],
        [2024, "Février", "Prestations externes", 5000.00],
        [2024, "Mars", "Location équipement", 800.50],
    ]
    
    for row_idx, row_data in enumerate(exemple_depenses, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_depenses.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border
            # Formater les montants avec un point comme séparateur décimal
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = '0.00'
    
    ws_depenses.column_dimensions['A'].width = 12
    ws_depenses.column_dimensions['B'].width = 15
    ws_depenses.column_dimensions['C'].width = 40
    ws_depenses.column_dimensions['D'].width = 15
    
    # Ajouter validation de mois pour Dépenses Externes
    if mois_liste:
        mois_str = ','.join(mois_liste)
        dv_mois_dep = DataValidation(type="list", formula1=f'"{mois_str}"', allow_blank=True)
        dv_mois_dep.error = 'Mois invalide ou hors période du projet'
        dv_mois_dep.errorTitle = 'Entrée invalide'
        dv_mois_dep.prompt = 'Sélectionnez un mois de la période du projet'
        dv_mois_dep.promptTitle = 'Mois'
        ws_depenses.add_data_validation(dv_mois_dep)
        dv_mois_dep.add('B3:B1000')
    
    # Ajouter validation d'année pour Dépenses Externes
    if annees_liste:
        annees_str = ','.join(map(str, annees_liste))
        dv_annee_dep = DataValidation(type="list", formula1=f'"{annees_str}"', allow_blank=True)
        dv_annee_dep.error = 'Année invalide ou hors période du projet'
        dv_annee_dep.errorTitle = 'Entrée invalide'
        dv_annee_dep.prompt = 'Sélectionnez une année de la période du projet'
        dv_annee_dep.promptTitle = 'Année'
        ws_depenses.add_data_validation(dv_annee_dep)
        dv_annee_dep.add('A3:A1000')
    
    # ==================== FEUILLE 3: AUTRES DÉPENSES ====================
    ws_autres = wb.create_sheet("Autres Dépenses")
    
    headers_autres = [
        "Année",
        "Mois",
        "Libellé",
        "Montant"
    ]
    
    descriptions_autres = [
        "Année (ex: 2024)",
        "Mois en français (ex: Janvier)",
        "Description de la dépense",
        "Montant en euros (ex: 1500,50)"
    ]
    
    for col, (header, desc) in enumerate(zip(headers_autres, descriptions_autres), 1):
        cell = ws_autres.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        comment_cell = ws_autres.cell(row=2, column=col, value=desc)
        comment_cell.font = example_font
        comment_cell.fill = example_fill
        comment_cell.border = thin_border
    
    # Exemples
    exemple_autres = [
        [2024, "Janvier", "Fournitures bureau", 300.00],
        [2024, "Février", "Frais de déplacement", 450.00],
        [2024, "Mars", "Formation personnel", 1200.00],
    ]
    
    for row_idx, row_data in enumerate(exemple_autres, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_autres.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border
            # Formater les montants avec un point comme séparateur décimal
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = '0.00'
    
    ws_autres.column_dimensions['A'].width = 12
    ws_autres.column_dimensions['B'].width = 15
    ws_autres.column_dimensions['C'].width = 40
    ws_autres.column_dimensions['D'].width = 15
    
    # Ajouter validation de mois pour Autres Dépenses
    if mois_liste:
        mois_str = ','.join(mois_liste)
        dv_mois_autres = DataValidation(type="list", formula1=f'"{mois_str}"', allow_blank=True)
        dv_mois_autres.error = 'Mois invalide ou hors période du projet'
        dv_mois_autres.errorTitle = 'Entrée invalide'
        dv_mois_autres.prompt = 'Sélectionnez un mois de la période du projet'
        dv_mois_autres.promptTitle = 'Mois'
        ws_autres.add_data_validation(dv_mois_autres)
        dv_mois_autres.add('B3:B1000')
    
    # Ajouter validation d'année pour Autres Dépenses
    if annees_liste:
        annees_str = ','.join(map(str, annees_liste))
        dv_annee_autres = DataValidation(type="list", formula1=f'"{annees_str}"', allow_blank=True)
        dv_annee_autres.error = 'Année invalide ou hors période du projet'
        dv_annee_autres.errorTitle = 'Entrée invalide'
        dv_annee_autres.prompt = 'Sélectionnez une année de la période du projet'
        dv_annee_autres.promptTitle = 'Année'
        ws_autres.add_data_validation(dv_annee_autres)
        dv_annee_autres.add('A3:A1000')
    
    # ==================== FEUILLE 4: RECETTES ====================
    ws_recettes = wb.create_sheet("Recettes")
    
    headers_recettes = [
        "Année",
        "Mois",
        "Libellé",
        "Montant"
    ]
    
    descriptions_recettes = [
        "Année (ex: 2024)",
        "Mois en français (ex: Janvier)",
        "Description de la recette",
        "Montant en euros (ex: 10000,00)"
    ]
    
    for col, (header, desc) in enumerate(zip(headers_recettes, descriptions_recettes), 1):
        cell = ws_recettes.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        comment_cell = ws_recettes.cell(row=2, column=col, value=desc)
        comment_cell.font = example_font
        comment_cell.fill = example_fill
        comment_cell.border = thin_border
    
    # Exemples
    exemple_recettes = [
        [2024, "Janvier", "Subvention recherche", 15000.00],
        [2024, "Mars", "Vente de licence", 5000.00],
        [2024, "Juin", "Partenariat", 8000.00],
    ]
    
    for row_idx, row_data in enumerate(exemple_recettes, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_recettes.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border
            # Formater les montants avec un point comme séparateur décimal
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = '0.00'
    
    ws_recettes.column_dimensions['A'].width = 12
    ws_recettes.column_dimensions['B'].width = 15
    ws_recettes.column_dimensions['C'].width = 40
    ws_recettes.column_dimensions['D'].width = 15
    
    # Ajouter validation de mois pour Recettes
    if mois_liste:
        mois_str = ','.join(mois_liste)
        dv_mois_rec = DataValidation(type="list", formula1=f'"{mois_str}"', allow_blank=True)
        dv_mois_rec.error = 'Mois invalide ou hors période du projet'
        dv_mois_rec.errorTitle = 'Entrée invalide'
        dv_mois_rec.prompt = 'Sélectionnez un mois de la période du projet'
        dv_mois_rec.promptTitle = 'Mois'
        ws_recettes.add_data_validation(dv_mois_rec)
        dv_mois_rec.add('B3:B1000')
    
    # Ajouter validation d'année pour Recettes
    if annees_liste:
        annees_str = ','.join(map(str, annees_liste))
        dv_annee_rec = DataValidation(type="list", formula1=f'"{annees_str}"', allow_blank=True)
        dv_annee_rec.error = 'Année invalide ou hors période du projet'
        dv_annee_rec.errorTitle = 'Entrée invalide'
        dv_annee_rec.prompt = 'Sélectionnez une année de la période du projet'
        dv_annee_rec.promptTitle = 'Année'
        ws_recettes.add_data_validation(dv_annee_rec)
        dv_annee_rec.add('A3:A1000')
    
    # ==================== FEUILLE 5: INSTRUCTIONS ====================
    ws_instructions = wb.create_sheet("Instructions", 0)  # Première position
    
    ws_instructions.merge_cells('A1:D1')
    title_cell = ws_instructions['A1']
    title_cell.value = "MODÈLE D'IMPORT DE DONNÉES DE PROJET"
    title_cell.font = Font(bold=True, size=16, color="366092")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    instructions_text = [
        "",
        "INSTRUCTIONS D'UTILISATION :",
        "",
        "1. Ce fichier contient 4 feuilles pour saisir les données de votre projet :",
        "   • Temps de Travail : temps passé par les membres de l'équipe",
        "   • Dépenses Externes : dépenses externes du projet",
        "   • Autres Dépenses : autres dépenses diverses",
        "   • Recettes : recettes du projet",
        "",
        "2. Chaque feuille contient :",
        "   • Ligne 1 : En-têtes de colonnes (NE PAS MODIFIER)",
        "   • Ligne 2 : Descriptions et exemples de format",
        "   • Lignes 3+ : Exemples de données (À REMPLACER par vos données)",
        "",
        "3. IMPORTANT - Format des données :",
        "   • Année : format numérique (ex: 2024)",
        "   • Mois : nom du mois en français (ex: Janvier)",
        "   • Montants : format numérique (ex: 1500,50 ou 1500.50)",
        "   • Jours : format numérique (ex: 15,5 ou 15.5)",
        "",
        "4. Saisie des données :",
        "   • Supprimez les lignes d'exemple (lignes 3+)",
        "   • Saisissez vos données réelles à partir de la ligne 3",
        "   • Vous pouvez ajouter autant de lignes que nécessaire",
        "   • Ne laissez pas de lignes vides au milieu de vos données",
        "",
        "5. Une fois le fichier complété :",
        "   • Enregistrez le fichier",
        "   • Utilisez la fonction d'import dans le logiciel",
        "   • Sélectionnez le projet cible",
        "   • Les données seront ajoutées à la base de données du projet",
        "",
        "6. Notes :",
        "   • Les données d'exemple sont fournies à titre indicatif",
        "   • Respectez bien les formats indiqués pour éviter les erreurs d'import",
        "   • En cas de doute, conservez une copie de votre fichier avant l'import",
        "",
        f"Date de création du modèle : {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ]
    
    for row_idx, text in enumerate(instructions_text, 3):
        cell = ws_instructions.cell(row=row_idx, column=1, value=text)
        if "INSTRUCTIONS" in text or "IMPORTANT" in text:
            cell.font = Font(bold=True, size=12, color="366092")
        else:
            cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    ws_instructions.column_dimensions['A'].width = 100
    ws_instructions.row_dimensions[1].height = 30
    
    # Sauvegarder le fichier
    wb.save(nom_fichier)
    print(f"✓ Modèle Excel créé avec succès : {nom_fichier}")
    if projet_data:
        print(f"  - Projet : {projet_data['code']} - {projet_data['nom']}")
        print(f"  - Listes déroulantes ajoutées pour le temps de travail")
    print(f"  - Feuille 1 : Instructions")
    print(f"  - Feuille 2 : Temps de Travail")
    print(f"  - Feuille 3 : Dépenses Externes")
    print(f"  - Feuille 4 : Autres Dépenses")
    print(f"  - Feuille 5 : Recettes")
    return nom_fichier


if __name__ == "__main__":
    creer_modele_excel()
