import sqlite3
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, 
                            QTableWidgetItem, QPushButton, QComboBox, QGroupBox, 
                            QCheckBox, QListWidget, QListWidgetItem, QDateEdit,
                            QRadioButton, QButtonGroup, QGridLayout, QSpinBox,
                            QSplitter, QWidget, QMessageBox, QScrollArea, QFrame,
                            QFileDialog)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
import datetime
from compte_resultat_dialog import CompteResultatDialog

DB_PATH = 'gestion_budget.db'

class PrintConfigDialog(QDialog):
    def __init__(self, parent, projet_id=None):
        super().__init__(parent)
        self.parent = parent
        self.projet_id = projet_id
        self.setWindowTitle("Configuration du Compte de Résultat")
        self.setMinimumSize(600, 600)

        layout = QVBoxLayout()

        # Titre
        title = QLabel("Configuration du Compte de Résultat")
        title.setStyleSheet("font-size: 16pt; font-weight: bold; margin-bottom: 20px;")
        layout.addWidget(title)

        # Sélection du projet
        project_group = QGroupBox("Sélection du projet")
        project_layout = QVBoxLayout()

        # Options de sélection
        self.selection_type_group = QButtonGroup(self)

        self.radio_single_project = QRadioButton("Un projet spécifique")
        self.radio_all_projects = QRadioButton("Tous les projets")
        self.radio_multiple_projects = QRadioButton("Sélection de certains projets")
        self.radio_by_theme = QRadioButton("Sélection par thèmes")
        self.radio_by_main_theme = QRadioButton("Sélection par le thème principal")

        self.selection_type_group.addButton(self.radio_single_project)
        self.selection_type_group.addButton(self.radio_all_projects)
        self.selection_type_group.addButton(self.radio_multiple_projects)
        self.selection_type_group.addButton(self.radio_by_theme)
        self.selection_type_group.addButton(self.radio_by_main_theme)

        self.radio_single_project.setChecked(True)

        project_layout.addWidget(self.radio_single_project)
        project_layout.addWidget(self.radio_all_projects)
        project_layout.addWidget(self.radio_multiple_projects)
        project_layout.addWidget(self.radio_by_theme)
        project_layout.addWidget(self.radio_by_main_theme)

        # Widgets dynamiques pour les options
        self.project_combo = QComboBox()
        self.project_list_widget = QListWidget()
        self.theme_list_widget = QListWidget()
        self.main_theme_list_widget = QListWidget()

        self.load_projects()
        self.load_themes()
        self.load_main_themes()

        project_layout.addWidget(self.project_combo)
        project_layout.addWidget(self.project_list_widget)
        project_layout.addWidget(self.theme_list_widget)
        project_layout.addWidget(self.main_theme_list_widget)

        self.project_combo.hide()
        self.project_list_widget.hide()
        self.theme_list_widget.hide()
        self.main_theme_list_widget.hide()

        self.radio_single_project.toggled.connect(self.update_selection_widgets)
        self.radio_all_projects.toggled.connect(self.update_selection_widgets)
        self.radio_multiple_projects.toggled.connect(self.update_selection_widgets)
        self.radio_by_theme.toggled.connect(self.update_selection_widgets)
        self.radio_by_main_theme.toggled.connect(self.update_selection_widgets)
        
        # Connecter le changement de projet pour mettre à jour les années
        self.project_combo.currentIndexChanged.connect(self.on_project_changed)

        project_group.setLayout(project_layout)
        layout.addWidget(project_group)

        # Sélection de la période
        period_group = QGroupBox("Période")
        period_layout = QVBoxLayout()

        # Type de période
        self.period_type = QComboBox()
        self.period_type.addItems(["Année spécifique", "Plusieurs années", "Période globale du projet"])
        self.period_type.currentIndexChanged.connect(self.update_period_widgets)
        period_layout.addWidget(QLabel("Type de période:"))
        period_layout.addWidget(self.period_type)

        # Sélection d'une année spécifique
        self.year_combo = QComboBox()
        period_layout.addWidget(self.year_combo)

        # Sélection de plusieurs années
        self.years_list_widget = QListWidget()
        self.years_list_widget.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        for year in range(2000, 2101):
            item = QListWidgetItem(str(year))
            item.setCheckState(Qt.CheckState.Unchecked)
            self.years_list_widget.addItem(item)
        period_layout.addWidget(self.years_list_widget)

        # Granularité (par mois ou par an)
        granularity_group = QGroupBox("Granularité")
        granularity_layout = QHBoxLayout()

        self.radio_monthly = QRadioButton("Par mois")
        self.radio_yearly = QRadioButton("Par an")
        self.radio_yearly.setChecked(True)

        granularity_layout.addWidget(self.radio_monthly)
        granularity_layout.addWidget(self.radio_yearly)
        granularity_group.setLayout(granularity_layout)
        period_layout.addWidget(granularity_group)

        period_group.setLayout(period_layout)
        layout.addWidget(period_group)

        # Boutons
        buttons_layout = QHBoxLayout()

        generate_btn = QPushButton("Générer le Compte de Résultat")
        generate_btn.clicked.connect(self.generate_compte_resultat)
        generate_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }")

        cancel_btn = QPushButton("Annuler")
        cancel_btn.clicked.connect(self.reject)

        buttons_layout.addStretch()
        buttons_layout.addWidget(generate_btn)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        # Initialiser l'affichage
        self.update_period_widgets()
        self.update_selection_widgets()

    def load_projects(self):
        """Charge les projets depuis la base de données"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT id, code, nom FROM projets ORDER BY code")
        projects = cursor.fetchall()
        conn.close()

        self.project_combo.clear()
        self.project_list_widget.clear()

        for project_id, code, name in projects:
            self.project_combo.addItem(f"{code} - {name}", project_id)
            item = QListWidgetItem(f"{code} - {name}")
            item.setData(Qt.ItemDataRole.UserRole, project_id)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.project_list_widget.addItem(item)
        
        # Connecter le signal pour recharger les années quand on coche/décoche des projets
        self.project_list_widget.itemChanged.connect(self.on_project_selection_changed)

    def load_themes(self):
        """Charge les thèmes depuis la base de données"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT id, nom FROM themes ORDER BY nom")
        themes = cursor.fetchall()
        conn.close()

        self.theme_list_widget.clear()

        for theme_id, name in themes:
            item = QListWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, theme_id)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.theme_list_widget.addItem(item)
        
        # Connecter le signal pour recharger les années quand on coche/décoche des thèmes
        self.theme_list_widget.itemChanged.connect(self.on_project_selection_changed)
        
        # Connecter le signal pour recharger les années quand on coche/décoche des thèmes principaux
        self.main_theme_list_widget.itemChanged.connect(self.on_project_selection_changed)

    def load_main_themes(self):
        """Charge les thèmes principaux depuis la base de données"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT theme_principal FROM projets WHERE theme_principal IS NOT NULL AND theme_principal != '' ORDER BY theme_principal")
        main_themes = cursor.fetchall()
        conn.close()

        self.main_theme_list_widget.clear()

        for (theme_name,) in main_themes:
            item = QListWidgetItem(theme_name)
            item.setData(Qt.ItemDataRole.UserRole, theme_name)
            item.setCheckState(Qt.CheckState.Unchecked)
            self.main_theme_list_widget.addItem(item)

    def on_project_changed(self):
        """Appelée quand le projet sélectionné change dans le ComboBox."""
        # Mettre à jour les années si on est en mode "Projet spécifique"
        if self.radio_single_project.isChecked():
            if self.period_type.currentIndex() == 0:  # Année spécifique
                self.update_years_for_project()
            elif self.period_type.currentIndex() == 1:  # Plusieurs années
                self.update_years_for_single_project_multiple_years()

    def update_years_for_single_project_multiple_years(self):
        """Met à jour la liste des années cochables pour un projet spécifique en mode 'plusieurs années'."""
        project_id = self.project_combo.currentData()
        if not project_id:
            QMessageBox.warning(self, "Erreur", "Aucun projet sélectionné.")
            self.years_list_widget.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            result = cursor.fetchone()
            if not result:
                QMessageBox.warning(self, "Erreur", "Impossible de récupérer les dates pour le projet sélectionné.")
                self.years_list_widget.clear()
                return

            date_debut, date_fin = result
            if not date_debut or not date_fin:
                QMessageBox.warning(self, "Erreur", "Les dates de début ou de fin sont manquantes pour ce projet.")
                self.years_list_widget.clear()
                return

            # Extraire les années à partir des dates au format MM/AAAA
            try:
                debut_annee = int(date_debut.split("/")[1])
                fin_annee = int(date_fin.split("/")[1])
            except (IndexError, ValueError):
                QMessageBox.warning(self, "Erreur", "Les dates de début ou de fin sont mal formatées (attendu: MM/AAAA).")
                self.years_list_widget.clear()
                return

            # Remplir la liste des années cochables
            self.years_list_widget.clear()
            for year in range(debut_annee, fin_annee + 1):
                item = QListWidgetItem(str(year))
                item.setCheckState(Qt.CheckState.Unchecked)
                self.years_list_widget.addItem(item)

        finally:
            conn.close()

    def update_years_for_all_projects(self):
        """Met à jour la liste déroulante des années avec toutes les années où il y a au moins un projet."""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE date_debut IS NOT NULL AND date_fin IS NOT NULL")
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet avec des dates valides n'a été trouvé.")
                self.year_combo.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets.")
                self.year_combo.clear()
                return

            # Remplir la liste déroulante des années
            self.year_combo.clear()
            for year in years:
                self.year_combo.addItem(str(year))

        finally:
            conn.close()

    def update_years_for_selected_projects(self):
        """Met à jour la liste déroulante des années avec les années des projets sélectionnés."""
        # Récupérer les IDs des projets cochés
        selected_project_ids = []
        for i in range(self.project_list_widget.count()):
            item = self.project_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected_project_ids.append(item.data(Qt.ItemDataRole.UserRole))
        
        if not selected_project_ids:
            # Aucun projet sélectionné, vider la liste
            self.year_combo.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les dates des projets sélectionnés
            placeholders = ','.join('?' * len(selected_project_ids))
            cursor.execute(f"""
                SELECT date_debut, date_fin FROM projets 
                WHERE id IN ({placeholders}) 
                AND date_debut IS NOT NULL AND date_fin IS NOT NULL
            """, selected_project_ids)
            
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet sélectionné avec des dates valides n'a été trouvé.")
                self.year_combo.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets sélectionnés.")
                self.year_combo.clear()
                return

            # Remplir la liste déroulante des années
            self.year_combo.clear()
            for year in years:
                self.year_combo.addItem(str(year))

        finally:
            conn.close()

    def update_years_for_selected_themes(self):
        """Met à jour la liste déroulante des années avec les années des projets liés aux thèmes sélectionnés."""
        # Récupérer les IDs des thèmes cochés
        selected_theme_ids = []
        for i in range(self.theme_list_widget.count()):
            item = self.theme_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected_theme_ids.append(item.data(Qt.ItemDataRole.UserRole))
        
        if not selected_theme_ids:
            # Aucun thème sélectionné, vider la liste
            self.year_combo.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les projets liés aux thèmes sélectionnés
            placeholders = ','.join('?' * len(selected_theme_ids))
            cursor.execute(f"""
                SELECT DISTINCT p.date_debut, p.date_fin 
                FROM projets p
                JOIN projet_themes pt ON p.id = pt.projet_id
                WHERE pt.theme_id IN ({placeholders})
                AND p.date_debut IS NOT NULL AND p.date_fin IS NOT NULL
            """, selected_theme_ids)
            
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet lié aux thèmes sélectionnés avec des dates valides n'a été trouvé.")
                self.year_combo.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets liés aux thèmes sélectionnés.")
                self.year_combo.clear()
                return

            # Remplir la liste déroulante des années
            self.year_combo.clear()
            for year in years:
                self.year_combo.addItem(str(year))

        finally:
            conn.close()

    def update_years_for_selected_themes_multiple(self):
        """Met à jour la liste des années cochables avec les années des projets liés aux thèmes sélectionnés."""
        # Récupérer les IDs des thèmes cochés
        selected_theme_ids = []
        for i in range(self.theme_list_widget.count()):
            item = self.theme_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected_theme_ids.append(item.data(Qt.ItemDataRole.UserRole))
        
        if not selected_theme_ids:
            # Aucun thème sélectionné, vider la liste
            self.years_list_widget.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les projets liés aux thèmes sélectionnés
            placeholders = ','.join('?' * len(selected_theme_ids))
            cursor.execute(f"""
                SELECT DISTINCT p.date_debut, p.date_fin 
                FROM projets p
                JOIN projet_themes pt ON p.id = pt.projet_id
                WHERE pt.theme_id IN ({placeholders})
                AND p.date_debut IS NOT NULL AND p.date_fin IS NOT NULL
            """, selected_theme_ids)
            
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet lié aux thèmes sélectionnés avec des dates valides n'a été trouvé.")
                self.years_list_widget.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets liés aux thèmes sélectionnés.")
                self.years_list_widget.clear()
                return

            # Remplir la liste des années cochables
            self.years_list_widget.clear()
            for year in years:
                item = QListWidgetItem(str(year))
                item.setCheckState(Qt.CheckState.Unchecked)
                self.years_list_widget.addItem(item)

        finally:
            conn.close()

    def update_years_for_selected_main_themes(self):
        """Met à jour la liste déroulante des années avec les années des projets liés aux thèmes principaux sélectionnés."""
        # Récupérer les thèmes principaux cochés
        selected_main_themes = []
        for i in range(self.main_theme_list_widget.count()):
            item = self.main_theme_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected_main_themes.append(item.data(Qt.ItemDataRole.UserRole))
        
        if not selected_main_themes:
            # Aucun thème principal sélectionné, vider la liste
            self.year_combo.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les projets avec les thèmes principaux sélectionnés
            placeholders = ','.join('?' * len(selected_main_themes))
            cursor.execute(f"""
                SELECT date_debut, date_fin 
                FROM projets 
                WHERE theme_principal IN ({placeholders})
                AND date_debut IS NOT NULL AND date_fin IS NOT NULL
            """, selected_main_themes)
            
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet avec les thèmes principaux sélectionnés et des dates valides n'a été trouvé.")
                self.year_combo.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets avec les thèmes principaux sélectionnés.")
                self.year_combo.clear()
                return

            # Remplir la liste déroulante des années
            self.year_combo.clear()
            for year in years:
                self.year_combo.addItem(str(year))

        finally:
            conn.close()

    def update_years_for_selected_main_themes_multiple(self):
        """Met à jour la liste des années cochables avec les années des projets liés aux thèmes principaux sélectionnés."""
        # Récupérer les thèmes principaux cochés
        selected_main_themes = []
        for i in range(self.main_theme_list_widget.count()):
            item = self.main_theme_list_widget.item(i)
            if item.checkState() == Qt.CheckState.Checked:
                selected_main_themes.append(item.data(Qt.ItemDataRole.UserRole))
        
        if not selected_main_themes:
            # Aucun thème principal sélectionné, vider la liste
            self.years_list_widget.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les projets avec les thèmes principaux sélectionnés
            placeholders = ','.join('?' * len(selected_main_themes))
            cursor.execute(f"""
                SELECT date_debut, date_fin 
                FROM projets 
                WHERE theme_principal IN ({placeholders})
                AND date_debut IS NOT NULL AND date_fin IS NOT NULL
            """, selected_main_themes)
            
            results = cursor.fetchall()
            
            if not results:
                QMessageBox.warning(self, "Aucun projet trouvé", "Aucun projet avec les thèmes principaux sélectionnés et des dates valides n'a été trouvé.")
                self.years_list_widget.clear()
                return

            years_set = set()
            
            for date_debut, date_fin in results:
                try:
                    # Extraire les années à partir des dates au format MM/AAAA
                    debut_annee = int(date_debut.split("/")[1])
                    fin_annee = int(date_fin.split("/")[1])
                    
                    # Ajouter toutes les années entre début et fin (incluses)
                    for year in range(debut_annee, fin_annee + 1):
                        years_set.add(year)
                        
                except (IndexError, ValueError):
                    # Si la conversion échoue, on ignore cette entrée
                    continue

            years = sorted(list(years_set))
            
            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année valide n'a été trouvée dans les projets avec les thèmes principaux sélectionnés.")
                self.years_list_widget.clear()
                return

            # Remplir la liste des années cochables
            self.years_list_widget.clear()
            for year in years:
                item = QListWidgetItem(str(year))
                item.setCheckState(Qt.CheckState.Unchecked)
                self.years_list_widget.addItem(item)

        finally:
            conn.close()

    def on_project_selection_changed(self):
        """Appelée quand la sélection de projets ou thèmes change"""
        # Mettre à jour les années selon le mode et le type de période
        if self.radio_multiple_projects.isChecked() and self.period_type.currentIndex() == 0:
            # Mode "Sélection de certains projets" + "Année spécifique"
            self.update_years_for_selected_projects()
        elif self.radio_by_theme.isChecked() and self.period_type.currentIndex() == 0:
            # Mode "Sélection par thèmes" + "Année spécifique"
            self.update_years_for_selected_themes()
        elif self.radio_by_theme.isChecked() and self.period_type.currentIndex() == 1:
            # Mode "Sélection par thèmes" + "Plusieurs années"
            self.update_years_for_selected_themes_multiple()
        elif self.radio_by_main_theme.isChecked() and self.period_type.currentIndex() == 0:
            # Mode "Sélection par le thème principal" + "Année spécifique"
            self.update_years_for_selected_main_themes()
        elif self.radio_by_main_theme.isChecked() and self.period_type.currentIndex() == 1:
            # Mode "Sélection par le thème principal" + "Plusieurs années"
            self.update_years_for_selected_main_themes_multiple()
        elif self.period_type.currentIndex() == 1:
            # Mode "Plusieurs années" pour les autres cas
            self.load_years_with_projects()

    def get_selected_project_ids(self):
        """Retourne la liste des IDs des projets sélectionnés selon l'option choisie"""
        project_ids = []
        
        if self.radio_single_project.isChecked():
            # Un projet spécifique
            project_id = self.project_combo.currentData()
            if project_id:
                project_ids = [project_id]
                
        elif self.radio_all_projects.isChecked():
            # Tous les projets
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM projets")
            project_ids = [row[0] for row in cursor.fetchall()]
            conn.close()
            
        elif self.radio_multiple_projects.isChecked():
            # Projets sélectionnés dans la liste
            for i in range(self.project_list_widget.count()):
                item = self.project_list_widget.item(i)
                if item.checkState() == Qt.CheckState.Checked:
                    project_ids.append(item.data(Qt.ItemDataRole.UserRole))
                    
        elif self.radio_by_theme.isChecked():
            # Projets par thèmes sélectionnés
            selected_theme_ids = []
            for i in range(self.theme_list_widget.count()):
                item = self.theme_list_widget.item(i)
                if item.checkState() == Qt.CheckState.Checked:
                    selected_theme_ids.append(item.data(Qt.ItemDataRole.UserRole))
            
            if selected_theme_ids:
                conn = sqlite3.connect(DB_PATH)
                cursor = conn.cursor()
                placeholders = ','.join('?' * len(selected_theme_ids))
                cursor.execute(f"""
                    SELECT DISTINCT projet_id FROM projet_themes 
                    WHERE theme_id IN ({placeholders})
                """, selected_theme_ids)
                project_ids = [row[0] for row in cursor.fetchall()]
                conn.close()
        
        return project_ids

    def load_years_with_projects(self):
        """Charge les années où des projets sélectionnés existent depuis la base de données"""
        project_ids = self.get_selected_project_ids()
        
        if not project_ids:
            # Ne pas afficher de message d'erreur, juste vider la liste
            self.years_list_widget.clear()
            return
            
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            # Récupérer les projets sélectionnés avec leurs dates
            placeholders = ','.join('?' * len(project_ids))
            cursor.execute(f"""
                SELECT date_debut, date_fin FROM projets 
                WHERE id IN ({placeholders})
                AND date_debut IS NOT NULL AND date_debut LIKE '__/____'
            """, project_ids)
            
            years_set = set()
            
            for date_debut, date_fin in cursor.fetchall():
                # Extraire l'année de début
                if date_debut and len(date_debut) >= 7:
                    year_debut = int(date_debut[-4:])
                    years_set.add(year_debut)
                    
                    # Si il y a une date de fin, ajouter toutes les années entre début et fin
                    if date_fin and len(date_fin) >= 7 and date_fin.count('/') == 1:
                        try:
                            year_fin = int(date_fin[-4:])
                            # Ajouter toutes les années entre début et fin (incluses)
                            for year in range(year_debut, year_fin + 1):
                                years_set.add(year)
                        except (ValueError, IndexError):
                            # Si la conversion échoue, on ignore la date de fin
                            pass

            years = sorted(list(years_set))

            if not years:
                QMessageBox.warning(self, "Aucune année trouvée", "Aucune année avec des projets sélectionnés n'a été trouvée.")

            self.years_list_widget.clear()
            for year in years:
                item = QListWidgetItem(str(year))
                item.setCheckState(Qt.CheckState.Unchecked)
                self.years_list_widget.addItem(item)
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement des années : {str(e)}")
        finally:
            conn.close()

    def update_years_for_project(self):
        """Met à jour la liste des années disponibles pour le projet sélectionné."""
        project_id = self.project_combo.currentData()
        if not project_id:
            QMessageBox.warning(self, "Erreur", "Aucun projet sélectionné.")
            self.year_combo.clear()
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT date_debut, date_fin FROM projets WHERE id = ?", (project_id,))
            result = cursor.fetchone()
            if not result:
                QMessageBox.warning(self, "Erreur", "Impossible de récupérer les dates pour le projet sélectionné.")
                self.year_combo.clear()
                return

            date_debut, date_fin = result
            if not date_debut or not date_fin:
                QMessageBox.warning(self, "Erreur", "Les dates de début ou de fin sont manquantes pour ce projet.")
                self.year_combo.clear()
                return

            # Extraire les années à partir des dates au format MM/AAAA
            try:
                debut_annee = int(date_debut.split("/")[1])
                fin_annee = int(date_fin.split("/")[1])
            except (IndexError, ValueError):
                QMessageBox.warning(self, "Erreur", "Les dates de début ou de fin sont mal formatées (attendu: MM/AAAA).")
                self.year_combo.clear()
                return

            # Remplir la liste déroulante des années
            self.year_combo.clear()
            for year in range(debut_annee, fin_annee + 1):
                self.year_combo.addItem(str(year))

        finally:
            conn.close()

    def update_period_widgets(self):
        """Met à jour l'affichage selon le type de période sélectionné"""
        period_type = self.period_type.currentIndex()

        if period_type == 0:  # Année spécifique
            self.year_combo.show()
            self.years_list_widget.hide()
            # Mettre à jour les années selon le type de sélection
            if self.radio_single_project.isChecked():
                self.update_years_for_project()
            elif self.radio_all_projects.isChecked():
                self.update_years_for_all_projects()
            elif self.radio_multiple_projects.isChecked():
                self.update_years_for_selected_projects()
            elif self.radio_by_theme.isChecked():
                self.update_years_for_selected_themes()
            elif self.radio_by_main_theme.isChecked():
                self.update_years_for_selected_main_themes()
        elif period_type == 1:  # Plusieurs années
            self.year_combo.hide()
            self.years_list_widget.show()
            # Recharger les années selon le type de sélection
            if self.radio_single_project.isChecked():
                self.update_years_for_single_project_multiple_years()
            elif self.radio_by_theme.isChecked():
                self.update_years_for_selected_themes_multiple()
            elif self.radio_by_main_theme.isChecked():
                self.update_years_for_selected_main_themes_multiple()
            else:
                self.load_years_with_projects()
        else:  # Période globale du projet
            self.year_combo.hide()
            self.years_list_widget.hide()

    def update_selection_widgets(self):
        """Met à jour les widgets de sélection en fonction de l'option choisie"""
        if self.radio_single_project.isChecked():
            self.project_combo.show()
            self.project_list_widget.hide()
            self.theme_list_widget.hide()

            # Pré-sélectionner le projet courant si défini
            if self.projet_id is not None:
                index = self.project_combo.findData(self.projet_id)
                if index != -1:
                    self.project_combo.setCurrentIndex(index)

            # Mettre à jour les années si "Année spécifique" est sélectionné
            if self.period_type.currentIndex() == 0:  # Année spécifique
                self.update_years_for_project()

        elif self.radio_all_projects.isChecked():
            self.project_combo.hide()
            self.project_list_widget.hide()
            self.theme_list_widget.hide()
        elif self.radio_multiple_projects.isChecked():
            self.project_combo.hide()
            self.project_list_widget.show()
            self.theme_list_widget.hide()
        elif self.radio_by_theme.isChecked():
            self.project_combo.hide()
            self.project_list_widget.hide()
            self.theme_list_widget.show()
            self.main_theme_list_widget.hide()
        elif self.radio_by_main_theme.isChecked():
            self.project_combo.hide()
            self.project_list_widget.hide()
            self.theme_list_widget.hide()
            self.main_theme_list_widget.show()
        
        # Mettre à jour les années selon le type de sélection si "Tous les projets" est sélectionné
        if self.radio_all_projects.isChecked():
            # Mettre à jour les années si on est en mode "Année spécifique"
            if self.period_type.currentIndex() == 0:  # Année spécifique
                self.update_years_for_all_projects()

        # Recharger les années si on est en mode "Plusieurs années" et seulement pour "Tous les projets"
        if self.period_type.currentIndex() == 1 and self.radio_all_projects.isChecked():
            self.load_years_with_projects()

    def get_selected_project_id(self):
        """Retourne l'ID du projet sélectionné"""
        if self.radio_single_project.isChecked():
            return self.project_combo.currentData()
        elif self.radio_multiple_projects.isChecked():
            # Récupérer les projets sélectionnés dans la liste
            selected_items = self.project_list_widget.selectedItems()
            return [item.data(Qt.ItemDataRole.UserRole) for item in selected_items]
        else:
            return None

    def get_selected_theme_ids(self):
        """Retourne les IDs des thèmes sélectionnés"""
        if self.radio_by_theme.isChecked():
            # Récupérer les thèmes sélectionnés dans la liste
            selected_items = self.theme_list_widget.selectedItems()
            return [item.data(Qt.ItemDataRole.UserRole) for item in selected_items]
        else:
            return None

    def get_selected_period(self):
        """Retourne les informations de période sélectionnée"""
        project_id = self.get_selected_project_id()
        year = int(self.year_combo.currentText()) if self.year_combo.currentText() else None
        
        if self.period_type.currentIndex() == 0:  # Année spécifique
            return project_id, year, None
        elif self.period_type.currentIndex() == 1:  # Plusieurs années
            # Récupérer les années sélectionnées dans la liste
            selected_items = self.years_list_widget.selectedItems()
            years = [int(item.text()) for item in selected_items]
            return project_id, years, None
        else:  # Période globale du projet
            return project_id, None, "complete_project"
    
    def generate_compte_resultat(self):
        """Génère le compte de résultat"""
        project_id, year, month = self.get_selected_period()
        
        if project_id is None:
            QMessageBox.warning(self, "Avertissement", "Veuillez sélectionner un projet.")
            return
        
        # Déterminer le type de période
        if month == "complete_project":
            period_type = "complete_project"
            year = None
            month = None
        else:
            period_type = None
        
        # Créer et afficher la fenêtre de compte de résultat
        compte_resultat_dialog = CompteResultatDialog(self.parent, project_id, year, month, period_type)
        compte_resultat_dialog.exec()
        self.accept()

class CompteResultatDialog(QDialog):
    def __init__(self, parent, projet_id, year, month=None, period_type=None):
        super().__init__(parent)
        self.projet_id = projet_id
        self.year = year
        self.month = month
        self.period_type = period_type
        
        # Configuration de la fenêtre
        if period_type == "complete_project":
            period_str = "Période complète du projet"
        elif month:
            period_str = f"{month:02d}/{year}"
        else:
            period_str = str(year)
        self.setWindowTitle(f"Compte de Résultat - {period_str}")
        self.setMinimumSize(800, 600)
        
        layout = QVBoxLayout()
        
        # En-tête avec titre et boutons d'actions
        header_layout = QHBoxLayout()
        
        # Titre
        project_name = self.get_project_name()
        title = QLabel(f"COMPTE DE RÉSULTAT\n{project_name}\nPériode: {period_str}")
        title.setStyleSheet("font-size: 14pt; font-weight: bold; margin-bottom: 10px;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Boutons d'actions
        excel_btn = QPushButton("Export Excel")
        excel_btn.clicked.connect(self.export_to_excel)
        header_layout.addWidget(excel_btn)
        
        pdf_btn = QPushButton("Export PDF")
        pdf_btn.clicked.connect(self.export_to_pdf)
        header_layout.addWidget(pdf_btn)
        
        print_btn = QPushButton("Imprimer")
        print_btn.clicked.connect(self.print_compte_resultat)
        header_layout.addWidget(print_btn)
        
        layout.addLayout(header_layout)
        
        # Configuration du pourcentage d'amortissement
        config_layout = QHBoxLayout()
        config_layout.addWidget(QLabel("Pourcentage d'amortissement à appliquer:"))
        
        self.amortissement_percent = QSpinBox()
        self.amortissement_percent.setRange(0, 100)
        self.amortissement_percent.setValue(10)
        self.amortissement_percent.setSuffix("%")
        self.amortissement_percent.valueChanged.connect(self.refresh_compte_resultat)
        config_layout.addWidget(self.amortissement_percent)
        
        refresh_btn = QPushButton("Actualiser")
        refresh_btn.clicked.connect(self.refresh_compte_resultat)
        config_layout.addWidget(refresh_btn)
        config_layout.addStretch();
        
        layout.addLayout(config_layout)
        
        # Tableau du compte de résultat
        self.compte_table = QTableWidget(0, 2)
        self.compte_table.setHorizontalHeaderLabels(["Poste", "Montant (€)"])
        self.compte_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.compte_table)
        
        # Bouton fermer
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.accept)
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)
        
        # Générer le compte de résultat
        self.refresh_compte_resultat()
    
    def get_project_name(self):
        """Récupère le nom du projet"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT code, nom FROM projets WHERE id = ?", (self.projet_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            code, nom = result
            return f"{code} - {nom}"
        return "Projet inconnu"
    
    def refresh_compte_resultat(self):
        """Actualise le compte de résultat"""
        try:
            # Collecter les données
            donnees = self.collecter_donnees_compte_resultat()
            
            # Afficher dans le tableau
            self.display_compte_resultat(donnees)
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération du compte de résultat: {str(e)}")
    
    def collector_donnees_compte_resultat(self):
        """Collecte toutes les données nécessaires pour le compte de résultat"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        try:
            # Conditions de filtrage selon la période
            if self.period_type == "complete_project":
                # Période complète du projet - pas de filtre de date
                condition = "1=1"
                params = []
            elif self.month:
                # Filtre par mois et année - utiliser les noms de mois français
                months_fr = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", 
                           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
                month_name = months_fr[self.month - 1]
                
                condition = "annee = ? AND mois = ?"
                params = [self.year, month_name]
            else:
                # Filtre par année seulement
                condition = "annee = ?"
                params = [self.year]
            
            donnees = {
                'recettes': 0,
                'subventions': 0,
                'depenses_externes': 0,
                'autres_achats': 0,
                'cout_direct': 0,
                'dotation_amortissements': 0,
                'credit_impot': 0
            }
            
            # 1. RECETTES
            cursor.execute(f"""
                SELECT COALESCE(SUM(montant), 0) FROM recettes 
                WHERE projet_id = ? AND {condition}
            """, [self.projet_id] + params)
            donnees['recettes'] = cursor.fetchone()[0]
            
            # 2. SUBVENTIONS - cette table n'a pas de structure temporelle, on prend tout
            if self.period_type == "complete_project":
                cursor.execute("""
                    SELECT COALESCE(SUM(cd), 0) FROM subventions 
                    WHERE projet_id = ?
                """, [self.projet_id])
            else:
                # Pour les subventions, on considère qu'elles s'appliquent sur toute la période du projet
                cursor.execute("""
                    SELECT COALESCE(SUM(cd), 0) FROM subventions 
                    WHERE projet_id = ?
                """, [self.projet_id])
            donnees['subventions'] = cursor.fetchone()[0]
            
            # 3. DÉPENSES EXTERNES (table depenses)
            cursor.execute(f"""
                SELECT COALESCE(SUM(montant), 0) FROM depenses 
                WHERE projet_id = ? AND {condition}
            """, [self.projet_id] + params)
            donnees['depenses_externes'] = cursor.fetchone()[0]
            
            # 4. AUTRES ACHATS (table autres_depenses)
            cursor.execute(f"""
                SELECT COALESCE(SUM(montant), 0) FROM autres_depenses 
                WHERE projet_id = ? AND {condition}
            """, [self.projet_id] + params)
            donnees['autres_achats'] = cursor.fetchone()[0]
            
            # 5. COÛT DIRECT (temps_travail * cout_production)
            cursor.execute(f"""
                SELECT COALESCE(SUM(t.jours * c.cout_production), 0)
                FROM temps_travail t
                JOIN categorie_cout c ON t.categorie = c.categorie AND t.annee = c.annee
                WHERE t.projet_id = ? AND {condition.replace('annee', 't.annee').replace('mois', 't.mois')}
            """, [self.projet_id] + params)
            
            donnees['cout_direct'] = cursor.fetchone()[0]
            
            # 6. DOTATION AMORTISSEMENTS (investissements)
            # Pour les investissements, on regarde la date d'achat
            if self.period_type == "complete_project":
                cursor.execute("""
                    SELECT COALESCE(SUM(montant / duree), 0) FROM investissements 
                    WHERE projet_id = ?
                """, [self.projet_id])
            else:
                # Pour simplifier, on prend les amortissements de l'année
                cursor.execute("""
                    SELECT COALESCE(SUM(montant / duree), 0) FROM investissements 
                    WHERE projet_id = ? AND strftime('%Y', date_achat) = ?
                """, [self.projet_id, str(self.year)])
            
            donnees['dotation_amortissements'] = cursor.fetchone()[0]
            
            # 7. CRÉDIT D'IMPÔT - Calculé selon les coefficients CIR
            # Pour simplifier, on met 0 pour l'instant
            donnees['credit_impot'] = 0
            
            return donnees
            
        finally:
            conn.close()
    
    def display_compte_resultat(self, donnees):
        """Affiche le compte de résultat dans le tableau"""
        self.compte_table.setRowCount(0)
        
        # PRODUITS D'EXPLOITATION
        self.add_table_row("=== PRODUITS D'EXPLOITATION ===", "", bold=True)
        self.add_table_row("Recettes", f"{donnees['recettes']:.2f}")
        self.add_table_row("Subventions d'exploitation", f"{donnees['subventions']:.2f}")
        
        total_produits_exploitation = donnees['recettes'] + donnees['subventions']
        self.add_table_row("TOTAL PRODUITS D'EXPLOITATION", f"{total_produits_exploitation:.2f}", bold=True)
        
        self.add_table_row("", "")  # Ligne vide
        
        # CHARGES D'EXPLOITATION
        self.add_table_row("=== CHARGES D'EXPLOITATION ===", "", bold=True)
        self.add_table_row("Dépenses externes", f"{donnees['depenses_externes']:.2f}")
        self.add_table_row("Autres achats", f"{donnees['autres_achats']:.2f}")
        self.add_table_row("Coût direct du personnel", f"{donnees['cout_direct']:.2f}")
        self.add_table_row("Dotation aux amortissements", f"{donnees['dotation_amortissements']:.2f}")
        
        total_charges_exploitation = (donnees['depenses_externes'] + donnees['autres_achats'] + 
                                    donnees['cout_direct'] + donnees['dotation_amortissements'])
        self.add_table_row("TOTAL CHARGES D'EXPLOITATION", f"{total_charges_exploitation:.2f}", bold=True)
        
        self.add_table_row("", "")  # Ligne vide
        
        # RÉSULTAT D'EXPLOITATION
        resultat_exploitation = total_produits_exploitation - total_charges_exploitation
        self.add_table_row("RÉSULTAT D'EXPLOITATION", f"{resultat_exploitation:.2f}", bold=True, 
                          color="green" if resultat_exploitation >= 0 else "red")
        
        self.add_table_row("", "")  # Ligne vide
        
        # PRODUITS EXCEPTIONNELS
        if donnees['credit_impot'] > 0:
            self.add_table_row("=== PRODUITS EXCEPTIONNELS ===", "", bold=True)
            self.add_table_row("Crédit d'impôt recherche", f"{donnees['credit_impot']:.2f}")
            self.add_table_row("TOTAL PRODUITS EXCEPTIONNELS", f"{donnees['credit_impot']:.2f}", bold=True)
            self.add_table_row("", "")  # Ligne vide
        
        # RÉSULTAT NET
        resultat_net = resultat_exploitation + donnees['credit_impot']
        self.add_table_row("RÉSULTAT NET", f"{resultat_net:.2f}", bold=True, 
                          color="green" if resultat_net >= 0 else "red", size=16)
        
        # Ajuster les dimensions du tableau
        self.compte_table.resizeColumnsToContents()
        self.compte_table.resizeRowsToContents()
    
    def add_table_row(self, poste, montant, bold=False, color=None, size=None):
        """Ajoute une ligne au tableau"""
        row = self.compte_table.rowCount()
        self.compte_table.insertRow(row)
        
        # Cellule poste
        poste_item = QTableWidgetItem(poste)
        if bold:
            font = poste_item.font()
            font.setBold(True)
            if size:
                font.setPointSize(size)
            poste_item.setFont(font)
        if color:
            if color == "green":
                poste_item.setBackground(QColor(144, 238, 144))  # lightgreen
            elif color == "red":
                poste_item.setBackground(QColor(255, 182, 193))  # lightcoral
        
        # Cellule montant
        montant_item = QTableWidgetItem(montant)
        montant_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        if bold:
            font = montant_item.font()
            font.setBold(True)
            if size:
                font.setPointSize(size)
            montant_item.setFont(font)
        if color:
            if color == "green":
                montant_item.setBackground(QColor(144, 238, 144))  # lightgreen
            elif color == "red":
                montant_item.setBackground(QColor(255, 182, 193))  # lightcoral
        
        self.compte_table.setItem(row, 0, poste_item)
        self.compte_table.setItem(row, 1, montant_item)
    
    def export_to_excel(self):
        """Exporte le compte de résultat vers Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exporter le compte de résultat",
                f"compte_resultat_{self.year}_{self.month or 'annuel'}.xlsx",
                "Fichiers Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Compte de Résultat"
            
            # Styles
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            bold_font = Font(bold=True)
            
            # Titre
            project_name = self.get_project_name()
            period_str = f"{self.month:02d}/{self.year}" if self.month else str(self.year)
            
            ws['A1'] = f"COMPTE DE RÉSULTAT - {project_name}"
            ws['A1'].font = title_font
            ws['A2'] = f"Période: {period_str}"
            ws['A2'].font = header_font
            
            # En-têtes
            ws['A4'] = "Poste"
            ws['B4'] = "Montant (€)"
            ws['A4'].font = bold_font
            ws['B4'].font = bold_font
            
            # Données
            row = 5
            for i in range(self.compte_table.rowCount()):
                poste_item = self.compte_table.item(i, 0)
                montant_item = self.compte_table.item(i, 1)
                
                if poste_item and montant_item:
                    ws[f'A{row}'] = poste_item.text()
                    ws[f'B{row}'] = montant_item.text()
                    
                    # Appliquer le style gras si nécessaire
                    if poste_item.font().bold():
                        ws[f'A{row}'].font = bold_font
                        ws[f'B{row}'].font = bold_font
                
                row += 1
            
            # Ajuster les colonnes
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 15
            
            wb.save(file_path)
            QMessageBox.information(self, "Export réussi", f"Fichier exporté: {file_path}")
            
        except ImportError:
            QMessageBox.critical(self, "Erreur", "Le module openpyxl n'est pas installé.\nInstallez-le avec: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export Excel: {str(e)}")
    
    def export_to_pdf(self):
        """Exporte le compte de résultat vers PDF"""
        try:
            from PyQt6.QtPrintSupport import QPrinter
            from PyQt6.QtGui import QTextDocument
            from PyQt6.QtCore import QMarginsF
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Exporter le compte de résultat",
                f"compte_resultat_{self.year}_{self.month or 'annuel'}.pdf",
                "Fichiers PDF (*.pdf)"
            )
            
            if not file_path:
                return
            
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageMargins(QMarginsF(15, 15, 15, 15), QPrinter.Unit.Millimeter)
            
            # Générer le HTML
            html_content = self.generate_html()
            
            document = QTextDocument()
            document.setHtml(html_content)
            document.print(printer)
            
            QMessageBox.information(self, "Export réussi", f"Fichier exporté: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export PDF: {str(e)}")
    
    def print_compte_resultat(self):
        """Imprime le compte de résultat"""
        try:
            from PyQt6.QtPrintSupport import QPrintDialog, QPrinter
            from PyQt6.QtGui import QTextDocument
            
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            
            dialog = QPrintDialog(printer, self)
            if dialog.exec() != QPrintDialog.DialogCode.Accepted:
                return
            
            # Générer le HTML et imprimer
            html_content = self.generate_html()
            
            document = QTextDocument()
            document.setHtml(html_content)
            document.print(printer)
            
            QMessageBox.information(self, "Impression", "Document envoyé à l'imprimante.")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'impression", f"Erreur lors de l'impression: {str(e)}")
    
    def generate_html(self):
        """Génère le contenu HTML pour l'export PDF/impression"""
        project_name = self.get_project_name()
        period_str = f"{self.month:02d}/{self.year}" if self.month else str(self.year)
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ text-align: center; color: #333; }}
                h2 {{ text-align: center; color: #666; font-size: 14pt; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 10px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                .bold {{ font-weight: bold; }}
                .amount {{ text-align: right; }}
                .total {{ background-color: #e8f4f8; font-weight: bold; }}
                .result {{ background-color: #d4edda; font-weight: bold; font-size: 14pt; }}
            </style>
        </head>
        <body>
            <h1>COMPTE DE RÉSULTAT</h1>
            <h2>{project_name}</h2>
            <h2>Période: {period_str}</h2>
            
            <table>
                <tr>
                    <th>Poste</th>
                    <th>Montant (€)</th>
                </tr>
        """
        
        # Ajouter les données du tableau
        for i in range(self.compte_table.rowCount()):
            poste_item = self.compte_table.item(i, 0)
            montant_item = self.compte_table.item(i, 1)
            
            if poste_item and montant_item:
                poste = poste_item.text()
                montant = montant_item.text()
                
                # Déterminer le style CSS
                css_class = ""
                if poste_item.font().bold():
                    if "RÉSULTAT NET" in poste:
                        css_class = "result"
                    elif "TOTAL" in poste or "RÉSULTAT" in poste:
                        css_class = "total"
                    else:
                        css_class = "bold"
                
                html += f'<tr class="{css_class}"><td>{poste}</td><td class="amount">{montant}</td></tr>'
        
        html += """
            </table>
        </body>
        </html>
        """
        
        return html

def show_print_config_dialog(parent, projet_id=None):
    """Fonction d'entrée pour afficher la configuration d'impression"""
    dialog = PrintConfigDialog(parent, projet_id)
    return dialog.exec()