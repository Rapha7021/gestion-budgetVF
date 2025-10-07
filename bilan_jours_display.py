import sqlite3
from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, 
                            QTableWidgetItem, QPushButton, QMessageBox, 
                            QFileDialog, QHeaderView, QGroupBox, QGridLayout, 
                            QColorDialog, QLineEdit, QComboBox)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
import os

DB_PATH = 'gestion_budget.db'

class BilanJoursDisplay(QDialog):
    def __init__(self, parent, config_data):
        super().__init__(parent)
        self.parent = parent
        self.config_data = config_data
        
        # Extraction des paramètres de configuration
        self.project_ids = config_data.get('project_ids', [])
        self.years = config_data.get('years', [])
        self.period_type = config_data.get('period_type', 'yearly')  # 'yearly' ou 'monthly'
        self.granularity = config_data.get('granularity', 'yearly')  # 'yearly' ou 'monthly'
        
        self.setWindowTitle("Bilan des Jours")
        self.setMinimumSize(1000, 700)
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        """Initialise l'interface utilisateur"""
        layout = QVBoxLayout()
        
        # En-tête avec titre et informations
        header_layout = self.create_header()
        layout.addLayout(header_layout)
        
        # Tableau principal du bilan des jours
        self.table = QTableWidget()
        self.setup_table()
        layout.addWidget(self.table)
        
        # Boutons d'actions
        buttons_layout = self.create_buttons()
        layout.addLayout(buttons_layout)
        
        self.setLayout(layout)
    
    def create_header(self):
        """Crée l'en-tête avec titre et informations"""
        header_layout = QHBoxLayout()
        
        # Titre principal
        title = QLabel("BILAN DES JOURS")
        title.setStyleSheet("font-size: 18pt; font-weight: bold; color: #2c3e50; margin-bottom: 10px;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Informations sur la sélection
        info_text = self.get_selection_info()
        info_label = QLabel(info_text)
        info_label.setStyleSheet("font-size: 10pt; color: #7f8c8d;")
        header_layout.addWidget(info_label)
        
        return header_layout
    
    def get_selection_info(self):
        """Génère le texte d'information sur la sélection"""
        info_lines = []
        
        # Projets
        if len(self.project_ids) == 1:
            project_name = self.get_project_name(self.project_ids[0])
            info_lines.append(f"Projet: {project_name}")
        else:
            info_lines.append(f"Projets: {len(self.project_ids)} sélectionnés")
        
        # Années
        if len(self.years) == 1:
            info_lines.append(f"Année: {self.years[0]}")
        else:
            years_str = f"{min(self.years)}-{max(self.years)}" if len(self.years) > 1 else str(self.years[0])
            info_lines.append(f"Années: {years_str}")
        
        # Granularité
        granularity_text = "Mensuel" if self.granularity == 'monthly' else "Annuel"
        info_lines.append(f"Granularité: {granularity_text}")
        
        return " | ".join(info_lines)
    
    def get_project_name(self, project_id):
        """Récupère le nom d'un projet"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT code, nom FROM projets WHERE id = ?", (project_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            code, nom = result
            return f"{code} - {nom}"
        return "Projet inconnu"
    
    def setup_table(self):
        """Configure le tableau du bilan des jours"""
        # Le nombre de colonnes dépend de la granularité et des directions
        if self.granularity == 'monthly':
            # Mode mensuel : une colonne par direction et par mois
            directions = self.get_active_directions()
            columns = ["Catégorie"]
            for year in sorted(self.years):
                for month in range(1, 13):
                    month_name = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun",
                                "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"][month-1]
                    for direction in directions:
                        columns.append(f"{direction}\n{month_name} {year}")
            columns.append("TOTAL")
        else:
            # Mode annuel : une colonne par direction
            directions = self.get_active_directions()
            columns = ["Catégorie"] + directions + ["TOTAL"]
        
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        
        # Configuration de l'apparence
        # Définir une largeur minimum pour la première colonne (catégories)
        self.table.setColumnWidth(0, 150)  # Largeur fixe de 150 pixels pour les catégories
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        # Les autres colonnes s'adaptent automatiquement
        for col in range(1, len(columns)):
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # Rendre le tableau en lecture seule
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
    
    def get_active_directions(self):
        """Récupère les directions qui ont des données pour les projets/années sélectionnés"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        try:
            # Construire la requête pour récupérer les directions actives
            placeholders_projects = ','.join('?' * len(self.project_ids))
            placeholders_years = ','.join('?' * len(self.years))
            
            query = f"""
                SELECT DISTINCT tt.direction
                FROM temps_travail tt
                WHERE tt.projet_id IN ({placeholders_projects})
                AND tt.annee IN ({placeholders_years})
                AND tt.direction IS NOT NULL
                AND tt.direction != ''
                AND tt.jours > 0
                ORDER BY tt.direction
            """
            
            cursor.execute(query, self.project_ids + self.years)
            directions = [row[0] for row in cursor.fetchall()]
            
            # Si aucune direction n'est trouvée avec des données, retourner une liste vide
            # (on ne récupère plus toutes les directions par défaut)
            return directions
            
        finally:
            conn.close()
    
    def create_buttons(self):
        """Crée les boutons d'actions"""
        buttons_layout = QHBoxLayout()
        
        # Export Excel
        excel_btn = QPushButton("Export Excel")
        excel_btn.clicked.connect(self.export_to_excel)
        excel_btn.setStyleSheet("QPushButton { background-color: #27ae60; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(excel_btn)
        
        # Export PDF
        pdf_btn = QPushButton("Export PDF")
        pdf_btn.clicked.connect(self.export_to_pdf)
        pdf_btn.setStyleSheet("QPushButton { background-color: #e74c3c; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(pdf_btn)
        
        # Imprimer
        print_btn = QPushButton("Imprimer")
        print_btn.clicked.connect(self.print_bilan_jours)
        print_btn.setStyleSheet("QPushButton { background-color: #3498db; color: white; font-weight: bold; padding: 8px; }")
        buttons_layout.addWidget(print_btn)
        
        buttons_layout.addStretch()
        
        # Fermer
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.accept)
        buttons_layout.addWidget(close_btn)
        
        return buttons_layout
    
    def load_data(self):
        """Charge et affiche les données du bilan des jours"""
        try:
            data = self.collect_jours_data()
            self.populate_table(data)
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors du chargement des données : {str(e)}")
            self.reject()  # Fermer le dialogue avec un code d'échec
    
    def has_data(self, data):
        """Vérifie s'il y a des données de jours travaillés"""
        for period_key, period_data in data.items():
            for direction, direction_data in period_data.items():
                for categorie, jours in direction_data.items():
                    if jours > 0:
                        return True
        return False

    def collect_jours_data(self):
        """Collecte toutes les données de jours travaillés"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        try:
            data = {}
            directions = self.get_active_directions()
            categories = self.get_categories()
            
            if self.granularity == 'monthly':
                # Mode mensuel
                for year in sorted(self.years):
                    for month in range(1, 13):
                        month_name = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                                    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"][month-1]
                        period_key = f"{year}_{month:02d}"
                        data[period_key] = self.collect_period_jours_data(cursor, year, month, directions, categories)
            else:
                # Mode annuel
                for year in sorted(self.years):
                    data[str(year)] = self.collect_period_jours_data(cursor, year, None, directions, categories)
            
            return data
            
        finally:
            conn.close()
    
    def get_categories(self):
        """Récupère les catégories d'équipe actives"""
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        try:
            # D'abord, essayer de récupérer les catégories depuis temps_travail
            placeholders_projects = ','.join('?' * len(self.project_ids))
            placeholders_years = ','.join('?' * len(self.years))
            
            query = f"""
                SELECT DISTINCT tt.categorie
                FROM temps_travail tt
                WHERE tt.projet_id IN ({placeholders_projects})
                AND tt.annee IN ({placeholders_years})
                AND tt.categorie IS NOT NULL
                AND tt.categorie != ''
                AND tt.jours > 0
                ORDER BY tt.categorie
            """
            
            cursor.execute(query, self.project_ids + self.years)
            categories = [row[0] for row in cursor.fetchall()]
            
            # Si aucune catégorie n'est trouvée avec des données, retourner une liste vide
            # (on ne utilise plus les catégories par défaut)
            return categories
            
        finally:
            conn.close()
    
    def collect_period_jours_data(self, cursor, year, month, directions, categories):
        """Collecte les données de jours pour une période spécifique"""
        period_data = {}
        
        # Initialiser toutes les combinaisons direction/catégorie à 0
        for direction in directions:
            period_data[direction] = {}
            for categorie in categories:
                period_data[direction][categorie] = 0.0
        
        # Construire la requête selon la granularité
        placeholders_projects = ','.join('?' * len(self.project_ids))
        
        if month is not None:
            # Requête mensuelle
            month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
                          "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            query = f"""
                SELECT tt.direction, tt.categorie, SUM(tt.jours) as total_jours
                FROM temps_travail tt
                WHERE tt.projet_id IN ({placeholders_projects})
                AND tt.annee = ?
                AND tt.mois = ?
                AND tt.direction IS NOT NULL
                AND tt.categorie IS NOT NULL
                GROUP BY tt.direction, tt.categorie
            """
            cursor.execute(query, self.project_ids + [year, month_names[month-1]])
        else:
            # Requête annuelle
            query = f"""
                SELECT tt.direction, tt.categorie, SUM(tt.jours) as total_jours
                FROM temps_travail tt
                WHERE tt.projet_id IN ({placeholders_projects})
                AND tt.annee = ?
                AND tt.direction IS NOT NULL
                AND tt.categorie IS NOT NULL
                GROUP BY tt.direction, tt.categorie
            """
            cursor.execute(query, self.project_ids + [year])
        
        # Remplir les données
        for direction, categorie, total_jours in cursor.fetchall():
            if direction in period_data and categorie in categories:
                period_data[direction][categorie] = float(total_jours or 0)
        
        return period_data
    
    def populate_table(self, data):
        """Remplit le tableau avec les données"""
        categories = self.get_categories()
        directions = self.get_active_directions()
        
        # Définir le nombre de lignes (catégories + ligne total)
        self.table.setRowCount(len(categories) + 1)
        
        # Remplir les catégories (première colonne)
        for row, categorie in enumerate(categories):
            item = QTableWidgetItem(categorie)
            item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            self.table.setItem(row, 0, item)
        
        # Ligne total
        total_item = QTableWidgetItem("TOTAL")
        total_item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        total_item.setBackground(QColor("#ecf0f1"))
        self.table.setItem(len(categories), 0, total_item)
        
        # Remplir les données selon la granularité
        if self.granularity == 'monthly':
            self.populate_monthly_data(data, categories, directions)
        else:
            self.populate_yearly_data(data, categories, directions)
    
    def populate_yearly_data(self, data, categories, directions):
        """Remplit le tableau en mode annuel"""
        col_index = 1
        totals_by_direction = {}
        totals_by_category = {cat: 0.0 for cat in categories}
        grand_total = 0.0
        
        # Pour chaque direction
        for direction in directions:
            totals_by_direction[direction] = 0.0
            
            # Pour chaque catégorie
            for row, categorie in enumerate(categories):
                total_jours = 0.0
                
                # Sommer sur toutes les années
                for year_key in data.keys():
                    year_data = data[year_key]
                    if direction in year_data and categorie in year_data[direction]:
                        total_jours += year_data[direction][categorie]
                
                # Afficher la valeur
                if total_jours > 0:
                    item = QTableWidgetItem(f"{total_jours:.1f}")
                else:
                    item = QTableWidgetItem("")
                
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table.setItem(row, col_index, item)
                
                # Mise à jour des totaux
                totals_by_direction[direction] += total_jours
                totals_by_category[categorie] += total_jours
            
            # Total par direction (dernière ligne)
            total_item = QTableWidgetItem(f"{totals_by_direction[direction]:.1f}" if totals_by_direction[direction] > 0 else "")
            total_item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            total_item.setBackground(QColor("#ecf0f1"))
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(len(categories), col_index, total_item)
            
            grand_total += totals_by_direction[direction]
            col_index += 1
        
        # Colonne TOTAL (dernière colonne)
        for row, categorie in enumerate(categories):
            total_item = QTableWidgetItem(f"{totals_by_category[categorie]:.1f}" if totals_by_category[categorie] > 0 else "")
            total_item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            total_item.setBackground(QColor("#d5dbdb"))
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, col_index, total_item)
        
        # Grand total (coin inférieur droit)
        grand_total_item = QTableWidgetItem(f"{grand_total:.1f}" if grand_total > 0 else "")
        grand_total_item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        grand_total_item.setBackground(QColor("#bdc3c7"))
        grand_total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(len(categories), col_index, grand_total_item)
    
    def populate_monthly_data(self, data, categories, directions):
        """Remplit le tableau en mode mensuel"""
        col_index = 1
        totals_by_category = {cat: 0.0 for cat in categories}
        grand_total = 0.0
        
        # Pour chaque année et chaque mois
        for year in sorted(self.years):
            for month in range(1, 13):
                period_key = f"{year}_{month:02d}"
                
                # Pour chaque direction
                for direction in directions:
                    totals_by_direction_month = 0.0
                    
                    # Pour chaque catégorie
                    for row, categorie in enumerate(categories):
                        jours = 0.0
                        
                        if period_key in data:
                            year_data = data[period_key]
                            if direction in year_data and categorie in year_data[direction]:
                                jours = year_data[direction][categorie]
                        
                        # Afficher la valeur
                        if jours > 0:
                            item = QTableWidgetItem(f"{jours:.1f}")
                        else:
                            item = QTableWidgetItem("")
                        
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        self.table.setItem(row, col_index, item)
                        
                        # Mise à jour des totaux
                        totals_by_direction_month += jours
                        totals_by_category[categorie] += jours
                    
                    # Total par direction/mois (dernière ligne)
                    total_item = QTableWidgetItem(f"{totals_by_direction_month:.1f}" if totals_by_direction_month > 0 else "")
                    total_item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
                    total_item.setBackground(QColor("#ecf0f1"))
                    total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(len(categories), col_index, total_item)
                    
                    grand_total += totals_by_direction_month
                    col_index += 1
        
        # Colonne TOTAL (dernière colonne)
        for row, categorie in enumerate(categories):
            total_item = QTableWidgetItem(f"{totals_by_category[categorie]:.1f}" if totals_by_category[categorie] > 0 else "")
            total_item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            total_item.setBackground(QColor("#d5dbdb"))
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, col_index, total_item)
        
        # Grand total (coin inférieur droit)
        grand_total_item = QTableWidgetItem(f"{grand_total:.1f}" if grand_total > 0 else "")
        grand_total_item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        grand_total_item.setBackground(QColor("#bdc3c7"))
        grand_total_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(len(categories), col_index, grand_total_item)
    
    def export_to_excel(self):
        """Exporte le bilan des jours vers Excel"""
        filename = self.generate_filename("xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exporter vers Excel", filename, "Fichiers Excel (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Créer un nouveau workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bilan des Jours"
            
            # Styles
            header_font = Font(name='Arial', size=12, bold=True)
            category_font = Font(name='Arial', size=10, bold=True)
            data_font = Font(name='Arial', size=9)
            total_font = Font(name='Arial', size=10, bold=True)
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            category_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Titre
            ws.merge_cells('A1:' + chr(65 + self.table.columnCount() - 1) + '1')
            title_cell = ws['A1']
            title_cell.value = "BILAN DES JOURS"
            title_cell.font = Font(name='Arial', size=16, bold=True)
            title_cell.alignment = center_alignment
            
            # Informations
            info_text = self.get_selection_info()
            ws.merge_cells('A2:' + chr(65 + self.table.columnCount() - 1) + '2')
            info_cell = ws['A2']
            info_cell.value = info_text
            info_cell.font = Font(name='Arial', size=10)
            info_cell.alignment = center_alignment
            
            # En-têtes de colonnes
            start_row = 4
            for col in range(self.table.columnCount()):
                cell = ws.cell(row=start_row, column=col+1)
                header_item = self.table.horizontalHeaderItem(col)
                cell.value = header_item.text() if header_item else ""
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Données
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    cell = ws.cell(row=start_row + 1 + row, column=col+1)
                    table_item = self.table.item(row, col)
                    
                    if table_item:
                        cell.value = table_item.text()
                        
                        # Appliquer les styles
                        if col == 0:  # Colonne catégorie
                            if row == self.table.rowCount() - 1:  # Ligne total
                                cell.font = total_font
                                cell.fill = total_fill
                            else:
                                cell.font = category_font
                                cell.fill = category_fill
                        elif row == self.table.rowCount() - 1:  # Ligne total
                            cell.font = total_font
                            cell.fill = total_fill
                        elif col == self.table.columnCount() - 1:  # Colonne total
                            cell.font = total_font
                            cell.fill = total_fill
                        else:
                            cell.font = data_font
                        
                        cell.alignment = center_alignment
            
            # Ajuster la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder
            wb.save(file_path)
            QMessageBox.information(self, "Export réussi", f"Le bilan des jours a été exporté vers :\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export Excel :\n{str(e)}")
    
    def export_to_pdf(self):
        """Exporte le bilan des jours vers PDF"""
        filename = self.generate_filename("pdf")
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Exporter vers PDF", filename, "Fichiers PDF (*.pdf)"
        )
        
        if not file_path:
            return
        
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageMargins(20, 20, 20, 20, QPrinter.Unit.Millimeter)
            
            html_content = self.generate_html_content()
            
            from PyQt6.QtGui import QTextDocument
            document = QTextDocument()
            document.setHtml(html_content)
            document.print(printer)
            
            QMessageBox.information(self, "Export réussi", f"Le bilan des jours a été exporté vers :\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export", f"Erreur lors de l'export PDF :\n{str(e)}")
    
    def print_bilan_jours(self):
        """Imprime le bilan des jours"""
        try:
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            print_dialog = QPrintDialog(printer, self)
            
            if print_dialog.exec() == QPrintDialog.DialogCode.Accepted:
                html_content = self.generate_html_content()
                
                from PyQt6.QtGui import QTextDocument
                document = QTextDocument()
                document.setHtml(html_content)
                document.print(printer)
                
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'impression", f"Erreur lors de l'impression :\n{str(e)}")
    
    def generate_html_content(self):
        """Génère le contenu HTML pour l'export PDF et l'impression"""
        html = """
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                h1 { text-align: center; color: #2c3e50; margin-bottom: 10px; }
                .info { text-align: center; color: #7f8c8d; margin-bottom: 20px; font-size: 12px; }
                table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                th, td { border: 1px solid #bdc3c7; padding: 8px; text-align: center; }
                th { background-color: #366092; color: white; font-weight: bold; }
                .category { background-color: #D9E2F3; font-weight: bold; }
                .total-row { background-color: #ecf0f1; font-weight: bold; }
                .total-col { background-color: #d5dbdb; font-weight: bold; }
                .grand-total { background-color: #bdc3c7; font-weight: bold; }
            </style>
        </head>
        <body>
        """
        
        html += "<h1>BILAN DES JOURS</h1>"
        html += f"<div class='info'>{self.get_selection_info()}</div>"
        
        html += "<table>"
        
        # En-têtes
        html += "<tr>"
        for col in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(col)
            header_text = header_item.text() if header_item else ""
            html += f"<th>{header_text}</th>"
        html += "</tr>"
        
        # Données
        for row in range(self.table.rowCount()):
            html += "<tr>"
            for col in range(self.table.columnCount()):
                table_item = self.table.item(row, col)
                cell_text = table_item.text() if table_item else ""
                
                # Déterminer la classe CSS
                css_class = ""
                if col == 0 and row < self.table.rowCount() - 1:  # Colonne catégorie (sauf total)
                    css_class = "category"
                elif row == self.table.rowCount() - 1:  # Ligne total
                    if col == self.table.columnCount() - 1:  # Grand total
                        css_class = "grand-total"
                    else:
                        css_class = "total-row"
                elif col == self.table.columnCount() - 1:  # Colonne total
                    css_class = "total-col"
                
                html += f"<td class='{css_class}'>{cell_text}</td>"
            html += "</tr>"
        
        html += "</table>"
        html += "</body></html>"
        
        return html
    
    def generate_filename(self, extension):
        """Génère un nom de fichier pour l'export"""
        base_name = "bilan_jours"
        
        # Ajouter les années
        if len(self.years) == 1:
            base_name += f"_{self.years[0]}"
        elif len(self.years) > 1:
            base_name += f"_{min(self.years)}-{max(self.years)}"
        
        # Ajouter la granularité
        if self.granularity == 'monthly':
            base_name += "_mensuel"
        else:
            base_name += "_annuel"
        
        # Ajouter l'extension
        return f"{base_name}.{extension}"


def show_bilan_jours(parent, config_data):
    """Fonction pour afficher le bilan des jours"""
    dialog = BilanJoursDisplay(parent, config_data)
    return dialog.exec()